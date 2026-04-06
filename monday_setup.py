#!/usr/bin/env python3
"""
Monday Setup — Run this each Monday to generate your weekly gym plan.

Reads your exercise tracking spreadsheet, analyzes the last 2 weeks of data,
applies progression rules, and outputs:
  1. gym_tracker.html  — phone-friendly tracker for the gym
  2. Weekly_Plan_<dates>.xlsx — printable plan with progression notes

Usage:
  python3 monday_setup.py <Weight2026_DEXA_calibratedc.xlsx> [output_dir]

Output goes to output_dir (default: current directory).
"""

import sys, os, json
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# COLUMN MAP: 0-indexed columns in the summary sheet
# ─────────────────────────────────────────────
COL = {
    "run": (5, 6), "vest walk": (7, 8), "hang": (9, None), "plank": (10, None),
    "knee lift": (11, None),
    "bench": (12, 13), "incl. bench": (14, 15), "mil press": (16, 17),
    "HS sh press": (18, 19), "pec fly": (20, 21), "Lat lift": (22, 23),
    "push ups": (24, None), "tric pull": (25, 26), "skull crush": (27, 28),
    "deadlift": (29, 30), "Row": (31, 32), "bent row": (33, 34),
    "lat pull dn": (35, 36), "TRX pullup": (37, None), "neg pull up": (38, None),
    "face pull": (39, 40), "cable curl": (41, 42), "curl": (43, 44),
    "hammer curl": (45, 46), "back ext": (47, 48),
    "squats": (49, 50), "land mine squat": (51, 52), "lunge walk": (53, 54),
    "leg press": (55, 56), "hip thrust": (57, 58), "Hstring curl": (59, 60),
    "weight step": (61, 62),
}

# ─────────────────────────────────────────────
# WEEKLY PLAN TEMPLATE
# Each entry: (display_name, logName, plan_sets, plan_reps_target, base_weight,
#              progression_rule, rest, is_dumbbell)
#   progression_rule: (threshold_reps, threshold_sets, increment)
#     e.g. (12, 3, 5) = "when you hit 12 reps for 3 sets, add 5 lbs"
#     For dumbbells, increment is per hand. None = no auto-progression.
# ─────────────────────────────────────────────
PUSH_A = [  # Tuesday
    ("Flat bench press (BB)", "bench",       3, "8-10",  95,  (10, 3, 5),    120, False),
    ("Incline bench (DB)",    "incl. bench", 3, "10-12", 25,  (12, 3, 2.5),  90,  True),
    ("Military press (DB)",   "mil press",   3, "10-12", 20,  (12, 3, 2.5),  90,  True),
    ("Pec fly (cable)",       "pec fly",     3, "12",    110, (12, 3, 5),    60,  False),
    ("Lateral raises (DB)",   "Lat lift",    3, "12-15", 10,  (15, 3, 2.5),  60,  True),
    ("Skull crushers (DB)",   "skull crush", 3, "10-12", 17.5,(12, 3, 2.5),  60,  True),
    ("Tricep pulldown",       "tric pull",   3, "12",    50,  (12, 3, 5),    60,  False),
]

PULL_A = [  # Wednesday
    ("Deadlift",              "deadlift",    3, "8-10",  135, (12, 3, 10),   120, False),
    ("Cable row",             "Row",         3, "10-12", 60,  (12, 3, 10),   90,  False),
    ("Lat pulldown",          "lat pull dn", 3, "10-12", 50,  (12, 3, 10),   90,  False),
    ("Face pulls (cable)",    "face pull",   3, "15-20", 30,  (20, 3, 5),    45,  False),
    ("Bicep curl (DB)",       "curl",        3, "10-12", 22.5,(12, 3, 2.5),  60,  True),
    ("Cable curl",            "cable curl",  3, "12",    30,  (12, 3, 5),    60,  False),
    ("Back extension",        "back ext",    3, "12",    50,  (12, 3, 10),   60,  False),
]

LEGS_A = [  # Thursday
    ("Barbell squats",        "squats",        3, "10-12", 70,  (12, 3, 5),    90,  False),
    ("Landmine squat (T-bar)","land mine squat",3,"10-12", 70,  (12, 3, 5),    90,  False),
    ("Leg press",             "leg press",     3, "10-12", 170, (12, 3, 10),   90,  False),
    ("Hip thrust",            "hip thrust",    3, "10",    140, (12, 3, 10),   90,  False),
    ("Hanging knee raises",   "knee lift",     3, "12",    0,   None,          45,  False),
    ("Dead hang",             "hang",          2, "max",   0,   None,          60,  False),
]

PUSH_B = [  # Friday
    ("HS shoulder press",     "HS sh press", 3, "10-12", 70,  (12, 3, 5),    90,  False),
    ("Incline bench (DB)",    "incl. bench", 3, "10-12", 25,  (12, 3, 2.5),  90,  True),
    ("Pec fly (cable)",       "pec fly",     3, "12",    110, (12, 3, 5),    60,  False),
    ("Lateral raises (DB)",   "Lat lift",    3, "12-15", 10,  (15, 3, 2.5),  60,  True),
    ("Push-ups",              "push ups",    2, "max",   0,   None,          60,  False),
    ("Tricep pulldown",       "tric pull",   3, "12",    50,  (12, 3, 5),    60,  False),
]

PULL_B = [  # Saturday
    ("Bent-over row (BB)",    "bent row",    3, "8-10",  85,  (10, 3, 5),    90,  False),
    ("Lat pulldown",          "lat pull dn", 3, "10-12", 60,  (12, 3, 10),   60,  False),
    ("TRX pull-ups",          "TRX pullup",  3, "12",    0,   None,          90,  False),
    ("Negative pull-ups",     "neg pull up", 3, "3-5",   0,   None,          90,  False),
    ("Face pulls (cable)",    "face pull",   3, "15-20", 30,  (20, 3, 5),    45,  False),
    ("Hammer curl (DB)",      "hammer curl", 3, "10-12", 22.5,(12, 3, 2.5),  60,  True),
    ("Plank",                 "plank",       2, "90s",   0,   None,          45,  False),
]

LEGS_B = [  # Sunday
    ("Lunge walk (DB)",       "lunge walk",  2, "12-15", 17.5,(15, 2, 2.5),  90,  True),
    ("Hip thrust",            "hip thrust",  3, "10-12", 140, (12, 3, 10),   90,  False),
    ("Hamstring curl (cable)","Hstring curl", 3, "12",   70,  (12, 3, 5),    60,  False),
    ("Back extension",        "back ext",    3, "12",    50,  (12, 3, 10),   60,  False),
    ("Weighted step-ups (DB)","weight step", 2, "12",    12,  (12, 2, 2.5),  60,  True),
    ("Hanging knee raises",   "knee lift",   3, "12",    0,   None,          45,  False),
    ("Dead hang",             "hang",        2, "max",   0,   None,          60,  False),
]

CARDIO = {
    "tue": ("1-mile run (warm-up)", "run", "1mi"),
    "wed": ("1-mile run (warm-up)", "run", "1mi"),
    "thu": ("1-mile run (warm-up)", "run", "1mi"),
    "fri": ("1-mile run (warm-up)", "run", "1mi"),
    "sat": ("5K run", "run", "5k"),
    "sun": ("1-mile run (warm-up)", "run", "1mi"),
}

VEST_WALK_DAYS = {"mon", "sat", "sun"}

WEEK = [
    ("mon", "MONDAY",    "Active Recovery \u2014 Rest Day",                    "#9B59B6", []),
    ("tue", "TUESDAY",   "PUSH A \u2014 Chest + Shoulders + Triceps (Heavy)", "#E94560", PUSH_A),
    ("wed", "WEDNESDAY", "PULL A \u2014 Back + Biceps + Rear Delts",          "#4A90D9", PULL_A),
    ("thu", "THURSDAY",  "LEGS A \u2014 Quads + Glutes + Core (PT)",          "#F5A623", LEGS_A),
    ("fri", "FRIDAY",    "PUSH B \u2014 Shoulders + Chest (Volume)",          "#E94560", PUSH_B),
    ("sat", "SATURDAY",  "PULL B \u2014 5K + Back + Biceps (Long Session)",   "#4A90D9", PULL_B),
    ("sun", "SUNDAY",    "LEGS B \u2014 Posterior Chain + Glutes + Core",     "#F5A623", LEGS_B),
]


# ─────────────────────────────────────────────
# READ RECENT DATA
# ─────────────────────────────────────────────
def read_recent_data(ws, lookback_days=14):
    today = datetime.now().date()
    cutoff = today - timedelta(days=lookback_days)
    data = {}
    for row in range(3, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=3).value
        if date_cell is None:
            continue
        if isinstance(date_cell, datetime):
            d = date_cell.date()
        else:
            continue
        if d < cutoff or d > today:
            continue
        for log_name, (wt_col, reps_col) in COL.items():
            wt_val = ws.cell(row=row, column=wt_col + 1).value
            reps_val = ws.cell(row=row, column=reps_col + 1).value if reps_col is not None else None
            if wt_val is not None or reps_val is not None:
                if log_name not in data:
                    data[log_name] = []
                data[log_name].append((d, str(wt_val) if wt_val else "", str(reps_val) if reps_val else ""))
    for k in data:
        data[k].sort(key=lambda x: x[0], reverse=True)
    return data


def parse_weight(val_str, is_db):
    if not val_str:
        return None
    s = str(val_str).strip()
    if s.startswith("2x") or s.startswith("2X"):
        try:
            return float(s[2:])
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_reps(reps_str):
    if not reps_str:
        return None, None
    s = str(reps_str).strip()
    if "x" in s.lower():
        parts = s.lower().split("x")
        try:
            return int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            return None, None
    if "," in s:
        parts = [p.strip() for p in s.split(",")]
        try:
            nums = [int(p) for p in parts if p]
            return min(nums), len(nums)
        except ValueError:
            return None, None
    try:
        return int(s), 1
    except ValueError:
        return None, None


def compute_target(log_name, is_db, base_weight, progression_rule, recent_data):
    entries = recent_data.get(log_name, [])
    if not entries:
        wt = base_weight
        fmt = f"2x{wt}" if is_db else str(wt)
        return wt, fmt, f"No recent data \u2014 start at plan weight ({fmt})", False, "\u2014"

    last_date, last_wt_str, last_reps_str = entries[0]
    last_session = f"{last_wt_str}: {last_reps_str}" if last_reps_str else last_wt_str
    last_date_str = last_date.strftime("%-m/%-d")
    last_session_display = f"{last_session} ({last_date_str})"

    last_weight = parse_weight(last_wt_str, is_db)
    last_reps, last_sets = parse_reps(last_reps_str)

    if last_weight is None:
        last_weight = base_weight

    if progression_rule and last_reps is not None and last_sets is not None:
        thresh_reps, thresh_sets, increment = progression_rule
        if last_reps >= thresh_reps and last_sets >= thresh_sets:
            new_weight = last_weight + increment
            fmt_old = f"2x{last_weight}" if is_db else str(last_weight)
            fmt_new = f"2x{new_weight}" if is_db else str(new_weight)
            note = f"\u2191 PROGRESS \u2014 hit {last_reps}\u00d7{last_sets} at {fmt_old}. Go to {fmt_new}."
            return new_weight, fmt_new, note, True, last_session_display

    fmt = f"2x{last_weight}" if is_db else str(last_weight)
    if last_reps and last_sets:
        note = f"Hold at {fmt}. Last: {last_reps}\u00d7{last_sets}."
    else:
        note = f"Hold at {fmt}."
    return last_weight, fmt, note, False, last_session_display


def get_last_cardio(log_name, recent_data):
    entries = recent_data.get(log_name, [])
    for d, wt, reps in entries:
        if wt:
            time_str = reps if reps else ""
            return f"{wt}" + (f" / {time_str}" if time_str else ""), d.strftime("%-m/%-d")
    return "\u2014", ""


# ─────────────────────────────────────────────
# STRETCHES (from stretches.json)
# ─────────────────────────────────────────────
def load_stretches(script_dir):
    """Load stretches.json from the same directory as this script."""
    path = os.path.join(script_dir, "stretches.json")
    if not os.path.exists(path):
        print(f"  ⚠ stretches.json not found at {path} — skipping stretches")
        return None
    with open(path) as f:
        data = json.load(f)
    print(f"  ✓ Loaded stretches.json (balance_level={data.get('balance_level', 1)})")
    return data


def build_stretch_exercises_js(stretches, day_id):
    """Build JS object literals for stretch exercises to prepend to a day."""
    if not stretches:
        return []
    items = []
    balance_level = stretches.get("balance_level", 1)

    for entry in stretches.get("core_routine", []):
        if "section" in entry and "name" not in entry:
            continue
        items.append(_stretch_js(entry))

    for entry in stretches.get("balance_block", []):
        if "section" in entry and "name" not in entry:
            continue
        if entry.get("level", 1) > balance_level:
            continue
        items.append(_stretch_js(entry))

    for entry in stretches.get("day_bonus", {}).get(day_id, []):
        items.append(_stretch_js(entry, bonus=True))

    return items


def _stretch_js(entry, bonus=False):
    """Convert a stretch JSON entry into a JS object literal string."""
    name = entry["name"].replace('"', '\\"')
    log = entry["logName"].replace('"', '\\"')
    plan = entry["plan"].replace('"', '\\"')
    note = entry.get("note", "").replace('"', '\\"')
    sets = entry.get("sets", "2")
    reps = entry.get("reps", "30s")
    rest = entry.get("rest", 30)
    bonus_flag = ",bonus:true" if bonus else ""
    return (
        '{type:"stretch",name:"%s",logName:"%s",plan:"%s",note:"%s",'
        'target:{sets:"%s",reps:"%s"},rest:%d%s}' %
        (name, log, plan, note, sets, reps, rest, bonus_flag)
    )


# ─────────────────────────────────────────────
# MODIFICATIONS (injury / deload overrides)
# ─────────────────────────────────────────────
def load_modifications(script_dir):
    """Load modifications.json from the same directory as this script."""
    path = os.path.join(script_dir, "modifications.json")
    if not os.path.exists(path):
        return None
    with open(path) as f:
        data = json.load(f)
    if not data.get("active", True):
        print("  ℹ modifications.json found but active=false — skipping")
        return None
    print(f"  ⚠ Modifications active: {data.get('label', 'unnamed')}")
    return data


def _mod_exercise_js(entry):
    """Convert a modification add_exercises entry into a JS object literal string."""
    etype = entry.get("type", "strength")
    name = entry["name"].replace('"', '\\"')
    log = entry["logName"].replace('"', '\\"')
    plan = entry.get("plan", "").replace('"', '\\"')
    note = entry.get("note", "").replace('"', '\\"')
    if etype == "cardio":
        fields = entry.get("fields", ["distance", "time"])
        return ('{type:"cardio",name:"%s",logName:"%s",fields:["%s"],'
                'defaults:{},plan:"%s",last:""}' %
                (name, log, '","'.join(fields), plan))
    else:
        t = entry.get("target", {})
        wt = t.get("weight", "")
        sets = t.get("sets", "3")
        reps = t.get("reps", "10")
        rest = entry.get("rest", 60)
        return ('{name:"%s",logName:"%s",plan:"%s",note:"%s",last:"",'
                'target:{weight:"%s",sets:"%s",reps:"%s"},rest:%d}' %
                (name, log, plan, note, wt, sets, reps, rest))


# ─────────────────────────────────────────────
# AUTO-DISCOVER EXERCISES FROM SUMMARY SHEET
# ─────────────────────────────────────────────
# Known exercise logNames from COL map + hardcoded extras + stretches
_KNOWN_LOG_NAMES = set(COL.keys()) | {"elliptical", "Ellip", "5k run", "leg ext", "calf raise", "erps"}

def scan_extra_exercises(ws, exclude=None):
    """Scan summary sheet headers for exercises not in the static COL map.
    Returns a list of dicts: {name, logName, type} for each discovered exercise."""
    extras = []
    seen = set(_KNOWN_LOG_NAMES)
    if exclude:
        seen |= exclude
    for c in range(6, ws.max_column + 1):
        h = ws.cell(row=2, column=c).value
        if h is None:
            continue
        h_str = str(h).strip()
        h_lower = h_str.lower()
        # Skip reps columns, notes column, and empty headers
        if 'reps' in h_lower or h_lower == 'notes' or not h_str:
            continue
        if h_str in seen:
            continue
        seen.add(h_str)
        # Determine type: check if next column is a reps column
        next_h = ws.cell(row=2, column=c + 1).value
        has_reps = next_h is not None and "reps" in str(next_h).lower()
        # Guess type from name
        cardio_hints = {'run', 'walk', 'bike', 'cycle', 'swim', 'row', 'ellip', 'cardio', 'jog'}
        ex_type = 'cardio' if any(hint in h_lower for hint in cardio_hints) else 'strength'
        extras.append({'name': h_str, 'logName': h_str, 'type': ex_type,
                       'has_reps': has_reps, 'col': c})
    return extras


# ─────────────────────────────────────────────
# GENERATE HTML
# ─────────────────────────────────────────────
def generate_html(week_dates, recent_data, output_path, stretches=None, mods=None, extra_exercises=None):
    date_map = {}
    for i, (day_id, _, _, _, _) in enumerate(WEEK):
        date_map[day_id] = week_dates[i]

    js_days = []
    for day_id, day_name, subtitle, color, exercises in WEEK:
        d = date_map[day_id]
        date_str = d.strftime("%B %-d")
        date_short = d.strftime("%-m/%-d")
        date_iso = d.strftime("%Y-%m-%d")

        # Check for day-level modifications
        day_mods = mods.get(day_id, {}) if mods else {}
        remove_set = set(day_mods.get("remove", []))
        wt_overrides = day_mods.get("weight_overrides", {})
        if day_mods.get("subtitle_override"):
            subtitle = day_mods["subtitle_override"]

        ex_json = []

        # Prepend stretch exercises from stretches.json
        stretch_items = build_stretch_exercises_js(stretches, day_id)
        ex_json.extend(stretch_items)

        # Cardio (skip if removed by modifications)
        if day_id in CARDIO and CARDIO[day_id][1] not in remove_set:
            cname, clog, cdist = CARDIO[day_id]
            clast, _ = get_last_cardio(clog, recent_data)
            ex_json.append(
                '{type:"cardio",name:"%s",logName:"%s",fields:["distance","time"],'
                'defaults:{distance:"%s"},plan:"%s",last:"%s"}' %
                (cname, clog, cdist,
                 "Warm-up" if cdist == "1mi" else "Morning run",
                 clast.replace('"', '\\"')))

        if day_id in VEST_WALK_DAYS and "vest walk" not in remove_set:
            vlast, _ = get_last_cardio("vest walk", recent_data)
            ex_json.append(
                '{type:"cardio",name:"Vest walk",logName:"vest walk",fields:["distance","time"],'
                'defaults:{},plan:"5-8 mi \\u2022 Recovery pace",last:"%s"}' %
                vlast.replace('"', '\\"'))

        for display, log_name, sets, reps_target, base_wt, prog_rule, rest, is_db in exercises:
            if log_name in remove_set:
                continue
            # Apply weight overrides from modifications
            if log_name in wt_overrides:
                base_wt = wt_overrides[log_name]["weight"]
            target_wt, wt_fmt, note, is_prog, last_str = compute_target(
                log_name, is_db, base_wt, prog_rule, recent_data)
            # Override note if modification specifies one
            if log_name in wt_overrides and "note" in wt_overrides[log_name]:
                note = wt_overrides[log_name]["note"]
                is_prog = False
            plan_str = "%d\\u00d7%s @ %s" % (sets, reps_target, wt_fmt)
            if is_prog:
                plan_str += " \\u2191"
            wt_field = wt_fmt if target_wt > 0 else ""
            ex_json.append(
                '{name:"%s",logName:"%s",plan:"%s",note:"%s",last:"%s",'
                'target:{weight:"%s",sets:"%d",reps:"%s"},rest:%d%s}' %
                (display, log_name, plan_str,
                 note.replace('"', '\\"'),
                 last_str.replace('"', '\\"'),
                 wt_field, sets, reps_target.split("-")[0],
                 rest,
                 ",progress:true" if is_prog else ""))

        # Append modification add_exercises (injury rehab, substitutions, etc.)
        for add_ex in day_mods.get("add_exercises", []):
            ex_json.append(_mod_exercise_js(add_ex))

        js_days.append(
            '{id:"%s",label:"%s",date:"%s",isoDate:"%s",'
            'title:"%s \\u2014 %s",'
            'subtitle:"%s",'
            'color:"%s",'
            'exercises:[%s]}' %
            (day_id, day_id.capitalize()[:3], date_short, date_iso,
             day_name, date_str, subtitle, color,
             ",\n      ".join(ex_json)))

    days_js = "[\n  " + ",\n  ".join(js_days) + "\n]"

    # Build JS for auto-discovered exercises from summary sheet
    discovered_js_items = []
    for ex in (extra_exercises or []):
        etype = ex['type']
        rest_val = 0 if etype == 'cardio' else 60
        discovered_js_items.append(
            "{ name: '%s', logName: '%s', type: '%s', rest: %d }" %
            (ex['name'].replace("'", "\\'"), ex['logName'].replace("'", "\\'"),
             etype, rest_val))
    discovered_js = ",\n  ".join(discovered_js_items) if discovered_js_items else ""

    html = HTML_TEMPLATE.replace("__DAYS_DATA__", days_js)
    html = html.replace("__DISCOVERED_EXERCISES__", discovered_js)
    with open(output_path, "w") as f:
        f.write(html)
    if discovered_js_items:
        print(f"  ✓ {output_path} (+ {len(discovered_js_items)} auto-discovered exercises)")
    else:
        print(f"  ✓ {output_path}")

HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="theme-color" content="#1A1A2E">
<link rel="manifest" href="manifest.json">
<title>Gym Tracker</title>
<style>
  :root { --bg:#1A1A2E;--card:#2D2D44;--green:#2ECC71;--text:#E8E8E8;--dim:#888;--red:#E94560; }
  * { box-sizing:border-box; margin:0; padding:0; }
  body { font-family:-apple-system,BlinkMacSystemFont,'SF Pro',system-ui,sans-serif;
    background:var(--bg); color:var(--text); padding:0 0 100px 0; -webkit-text-size-adjust:100%; }
  .tabs { display:flex; position:sticky; top:0; z-index:100; background:var(--bg);
    border-bottom:2px solid #333; overflow-x:auto; -webkit-overflow-scrolling:touch; }
  .tab { flex:1; min-width:52px; padding:12px 4px; text-align:center; font-size:13px;
    font-weight:600; color:var(--dim); cursor:pointer; border-bottom:3px solid transparent; }
  .tab.active { color:var(--text); }
  .tab .day-label { font-size:11px; display:block; margin-top:2px; }
  .day-panel { display:none; padding:16px; }
  .day-panel.active { display:block; }
  .day-title { font-size:18px; font-weight:700; margin-bottom:4px; }
  .day-subtitle { font-size:13px; color:var(--dim); margin-bottom:16px; }
  .exercise-card { background:var(--card); border-radius:12px; padding:14px 14px 14px 14px;
    margin-bottom:12px; border-left:4px solid var(--dim); position:relative; }
  .exercise-card.cardio { border-left-color:var(--green); }
  .exercise-card.stretch { border-left-color:#1ABC9C; }
  .section-divider { font-size:12px; font-weight:700; color:#1ABC9C; text-transform:uppercase;
    letter-spacing:1px; margin:20px 0 12px 0; padding-top:8px; border-top:1px solid rgba(26,188,156,0.2); }
  .section-divider:first-child { margin-top:0; border-top:none; }
  .section-divider.workout { color:var(--text); margin-top:24px; border-top-color:rgba(255,255,255,0.15); }

  .exercise-header { display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; }
  .exercise-name { font-size:15px; font-weight:600; }
  .exercise-plan { font-size:13px; color:var(--dim); margin-bottom:10px; }
  .exercise-plan .prog { color:var(--green); font-weight:600; }
  .input-row { display:flex; gap:8px; align-items:center; }
  .input-group { flex:1; }
  .input-group label { display:block; font-size:11px; color:var(--dim); margin-bottom:3px;
    text-transform:uppercase; letter-spacing:0.5px; }
  .input-group input { width:100%; background:rgba(255,255,255,0.1); border:1px solid rgba(255,255,255,0.15);
    border-radius:8px; color:var(--text); padding:10px; font-size:16px; font-family:inherit; -webkit-appearance:none; }
  .input-group input:focus { outline:none; border-color:rgba(255,255,255,0.4); background:rgba(255,255,255,0.15); }
  .input-group input::placeholder { color:rgba(255,255,255,0.25); }
  .cardio-inputs { display:flex; gap:8px; }
  .rest-timer { background:rgba(255,255,255,0.08); border:1px solid rgba(255,255,255,0.15);
    border-radius:8px; padding:8px 14px; font-size:14px; color:var(--text); cursor:pointer;
    font-family:'SF Mono',monospace; min-width:80px; text-align:center; }
  .rest-timer.running { color:var(--green); border-color:var(--green); }
  .rest-timer.done { color:var(--red); border-color:var(--red); animation:pulse 0.5s ease 3; }
  @keyframes pulse { 0%,100%{opacity:1;} 50%{opacity:0.5;} }
  .done-check { width:28px; height:28px; border-radius:50%; border:2px solid rgba(255,255,255,0.2);
    display:flex; align-items:center; justify-content:center; cursor:pointer; flex-shrink:0; }
  .done-check.checked { background:var(--green); border-color:var(--green); }
  .done-check.checked::after { content:'\2713'; color:white; font-weight:700; font-size:14px; }
  .exercise-card.completed { opacity:0.5; }
  .bottom-bar { position:fixed; bottom:0; left:0; right:0; background:var(--card);
    border-top:1px solid #444; padding:12px 16px; display:flex; gap:10px; z-index:100; }
  .btn { flex:1; padding:12px; border-radius:10px; border:none; font-size:14px;
    font-weight:600; cursor:pointer; font-family:inherit; }
  .btn-primary { background:var(--green); color:#fff; }
  .btn-secondary { background:rgba(255,255,255,0.1); color:var(--text); }
  .toast { position:fixed; top:60px; left:50%; transform:translateX(-50%); background:var(--green);
    color:#fff; padding:10px 24px; border-radius:8px; font-size:14px; font-weight:600;
    opacity:0; transition:opacity 0.3s; z-index:200; pointer-events:none; }
  .toast.show { opacity:1; }
  .progress-bar { height:3px; background:rgba(255,255,255,0.1); margin:8px 0 0; border-radius:2px; overflow:hidden; }
  .progress-fill { height:100%; background:var(--green); transition:width 0.3s; border-radius:2px; }
  .notes-area { width:100%; background:rgba(255,255,255,0.08); border:1px solid rgba(255,255,255,0.15);
    border-radius:8px; color:var(--text); padding:10px; font-size:14px; font-family:inherit;
    min-height:44px; resize:vertical; margin-top:8px; }
  .notes-area::placeholder { color:rgba(255,255,255,0.25); }
  .add-activity-btn { display:block; width:100%; padding:14px; margin-top:8px; border:2px dashed rgba(255,255,255,0.15);
    border-radius:12px; background:transparent; color:var(--dim); font-size:14px; font-weight:600;
    font-family:inherit; cursor:pointer; text-align:center; }
  .add-activity-btn:active { background:rgba(255,255,255,0.05); }
  .modal-overlay { display:none; position:fixed; top:0; left:0; right:0; bottom:0;
    background:rgba(0,0,0,0.7); z-index:300; align-items:center; justify-content:center; padding:20px; }
  .modal-overlay.show { display:flex; }
  .modal { background:var(--card); border-radius:16px; padding:20px; width:100%; max-width:340px; }
  .modal h3 { font-size:16px; margin-bottom:16px; }
  .modal .input-group { margin-bottom:12px; }
  .modal-btns { display:flex; gap:10px; margin-top:16px; }
  .modal-btns .btn { padding:10px; font-size:14px; }
  .remove-btn { position:absolute; top:8px; right:40px; width:22px; height:22px; border-radius:50%;
    border:1px solid rgba(255,255,255,0.15); background:transparent; color:var(--dim); font-size:12px;
    cursor:pointer; display:none; align-items:center; justify-content:center; }
  .exercise-card.added .remove-btn { display:flex; }
  .repeat-btn { width:26px; height:26px; border-radius:50%;
    border:1px solid rgba(255,255,255,0.15); background:transparent; color:var(--dim); font-size:15px;
    cursor:pointer; display:flex; align-items:center; justify-content:center; flex-shrink:0; margin-left:6px; }
  .repeat-btn:active { background:rgba(255,255,255,0.15); }
  .occurrence-label { font-size:11px; color:var(--dim); font-weight:600; margin-left:6px; }
  .swap-bar { display:flex; align-items:center; justify-content:space-between; padding:8px 16px;
    background:rgba(255,255,255,0.05); border-bottom:1px solid #333; }
  .swap-btn { background:rgba(255,255,255,0.1); border:1px solid rgba(255,255,255,0.15); border-radius:8px;
    color:var(--text); padding:6px 12px; font-size:12px; font-weight:600; font-family:inherit; cursor:pointer; }
  .swap-btn:active { background:rgba(255,255,255,0.2); }
  .swap-info { font-size:12px; color:var(--green); font-weight:600; flex:1; margin-left:10px; }
  .swap-undo { background:transparent; border:1px solid var(--red); border-radius:8px;
    color:var(--red); padding:4px 10px; font-size:11px; font-weight:600; font-family:inherit; cursor:pointer; margin-left:8px; }
  .modal select { width:100%; background:rgba(255,255,255,0.1); border:1px solid rgba(255,255,255,0.15);
    border-radius:8px; color:var(--text); padding:10px; font-size:16px; font-family:inherit; -webkit-appearance:none; }
  .swap-arrow { text-align:center; font-size:20px; color:var(--dim); padding:4px 0; }
</style>
</head>
<body>
<div class="tabs" id="tabs"></div>
<div class="swap-bar" id="swapBar">
  <button class="swap-btn" onclick="openSwapModal()">&#8644; Swap Days</button>
  <span class="swap-info" id="swapInfo"></span>
</div>
<div id="panels"></div>
<div class="toast" id="toast">Copied to clipboard!</div>

<!-- Add Activity Modal -->
<div class="modal-overlay" id="addModal">
  <div class="modal">
    <h3 id="modalTitle">Add Activity</h3>
    <div class="input-group">
      <label>Exercise</label>
      <select id="addExercise" onchange="onAddExerciseChange()"></select>
    </div>
    <div id="customFields" style="display:none">
      <div class="input-group">
        <label>Type</label>
        <select id="addType">
          <option value="strength">Strength</option>
          <option value="cardio">Cardio</option>
          <option value="stretch">Stretch</option>
        </select>
      </div>
      <div class="input-group">
        <label>Exercise Name</label>
        <input type="text" id="addName" placeholder="e.g. Vest walk, Bench press" inputmode="text">
      </div>
      <div class="input-group">
        <label>Log Name (for Excel)</label>
        <input type="text" id="addLogName" placeholder="e.g. vest walk, bench" inputmode="text">
      </div>
    </div>
    <div class="input-group" id="addRestGroup">
      <label>Rest Timer (seconds)</label>
      <input type="text" id="addRest" placeholder="60" inputmode="numeric">
    </div>
    <div class="modal-btns">
      <button class="btn btn-secondary" onclick="closeAddModal()">Cancel</button>
      <button class="btn btn-primary" onclick="confirmAdd()">Add</button>
    </div>
  </div>
</div>

<!-- Swap Days Modal -->
<div class="modal-overlay" id="swapModal">
  <div class="modal">
    <h3>Swap Exercise Plans</h3>
    <div class="input-group">
      <label>Swap this day...</label>
      <select id="swapDay1"></select>
    </div>
    <div class="swap-arrow">&#8597;</div>
    <div class="input-group">
      <label>...with this day</label>
      <select id="swapDay2"></select>
    </div>
    <div class="modal-btns">
      <button class="btn btn-secondary" onclick="closeSwapModal()">Cancel</button>
      <button class="btn btn-primary" onclick="confirmSwap()">Swap</button>
    </div>
  </div>
</div>

<div class="bottom-bar">
  <button class="btn btn-secondary" onclick="copyResults()">Copy Results</button>
  <button class="btn btn-primary" onclick="copyForExcel()">Copy for Excel</button>
</div>

<script>
// ============================================================
// DATA
// ============================================================
var DAYS = __DAYS_DATA__;

// ============================================================
// EXERCISE CATALOG — built from DAYS so dropdown always matches Excel
// ============================================================
var EX_CATALOG = [];
var _seenLog = {};
DAYS.forEach(function(day) {
  day.exercises.forEach(function(ex) {
    if (_seenLog[ex.logName]) return;
    _seenLog[ex.logName] = true;
    EX_CATALOG.push({ name: ex.name, logName: ex.logName, type: ex.type || 'strength', rest: ex.rest || 60 });
  });
});
// Additional cardio exercises not in any day's plan but available to add
var EXTRA_CARDIO = [
  { name: 'Elliptical', logName: 'elliptical', type: 'cardio', rest: 0 },
  { name: '5K run', logName: '5k run', type: 'cardio', rest: 0 }
];
var EXTRA_STRENGTH = [
  { name: 'Leg extension (machine)', logName: 'leg ext', type: 'strength', rest: 60 },
  { name: 'Seated calf raise', logName: 'calf raise', type: 'strength', rest: 45 }
];
// Auto-discovered exercises from summary sheet columns (added by monday_setup.py)
var DISCOVERED_EXERCISES = [
  __DISCOVERED_EXERCISES__
];
EXTRA_STRENGTH.forEach(function(ex) {
  if (!_seenLog[ex.logName]) { _seenLog[ex.logName] = true; EX_CATALOG.push(ex); }
});
EXTRA_CARDIO.forEach(function(ex) {
  if (!_seenLog[ex.logName]) { _seenLog[ex.logName] = true; EX_CATALOG.push(ex); }
});
DISCOVERED_EXERCISES.forEach(function(ex) {
  if (ex && ex.logName && !_seenLog[ex.logName]) { _seenLog[ex.logName] = true; EX_CATALOG.push(ex); }
});
var _typeOrder = {cardio:0, stretch:1, strength:2};
EX_CATALOG.sort(function(a, b) {
  var ta = _typeOrder[a.type] !== undefined ? _typeOrder[a.type] : 2;
  var tb = _typeOrder[b.type] !== undefined ? _typeOrder[b.type] : 2;
  if (ta !== tb) return ta - tb;
  return a.name.localeCompare(b.name);
});

// ============================================================
// STORAGE KEY — changes weekly so old data doesn't persist forever
// ============================================================
var STORE_KEY = 'gym_' + DAYS[0].isoDate;

// Keep original exercises for swap/undo
var ORIG_EXERCISES = {};
DAYS.forEach(function(day) { ORIG_EXERCISES[day.id] = day.exercises.slice(); });
var activeSwap = null; // { day1: id, day2: id }

function captureState() {
  var state = { checked: {}, inputs: {}, added: {}, swap: activeSwap };
  document.querySelectorAll('.exercise-card').forEach(function(card) {
    var id = card.id.replace('card-', '');
    var chk = card.querySelector('.done-check');
    if (chk && chk.classList.contains('checked')) state.checked[id] = true;
    card.querySelectorAll('input, textarea').forEach(function(inp) {
      if (inp.id && inp.value) state.inputs[inp.id] = inp.value;
    });
  });
  DAYS.forEach(function(day) {
    var extras = day._added || [];
    if (extras.length) state.added[day.id] = extras;
  });
  return state;
}

function restoreState(state) {
  Object.keys(state.inputs || {}).forEach(function(id) {
    var el = document.getElementById(id);
    if (el && !el.value) el.value = state.inputs[id];
  });
  Object.keys(state.checked || {}).forEach(function(cid) {
    var card = document.getElementById('card-' + cid);
    if (card) {
      var chk = card.querySelector('.done-check');
      if (chk && !chk.classList.contains('checked')) { chk.classList.add('checked'); card.classList.add('completed'); }
    }
  });
}

function saveState() {
  var state = captureState();
  try { localStorage.setItem(STORE_KEY, JSON.stringify(state)); } catch(e) {}
}

function loadState() {
  try {
    var raw = localStorage.getItem(STORE_KEY);
    if (!raw) return null;
    return JSON.parse(raw);
  } catch(e) { return null; }
}

// ============================================================
// RENDERING
// ============================================================
var tabsEl = document.getElementById('tabs');
var panelsEl = document.getElementById('panels');

function countOccurrences(day, logName) {
  var allEx = day.exercises.concat(day._added || []);
  var count = 0;
  for (var i = 0; i < allEx.length; i++) { if (allEx[i].logName === logName) count++; }
  return count;
}
function getOccurrenceNum(day, logName, idx) {
  var allEx = day.exercises.concat(day._added || []);
  var n = 0;
  for (var i = 0; i <= idx && i < allEx.length; i++) { if (allEx[i].logName === logName) n++; }
  return n;
}

function buildExerciseHTML(day, ex, idx, isAdded) {
  var cid = day.id + '-' + idx;
  var total = countOccurrences(day, ex.logName);
  var occNum = total > 1 ? getOccurrenceNum(day, ex.logName, idx) : 0;
  var occLabel = occNum > 0 ? '<span class="occurrence-label">#' + occNum + '</span>' : '';
  var repeatBtn = '<button class="repeat-btn" onclick="repeatExercise(\'' + day.id + '\',' + idx + ')" title="Add another">+</button>';
  var html = '';
  if (ex.type === 'cardio') {
    html += '<div class="exercise-card cardio' + (isAdded ? ' added' : '') + '" id="card-' + cid + '">' +
      (isAdded ? '<button class="remove-btn" onclick="removeAdded(\'' + day.id + '\',' + idx + ')">\u00d7</button>' : '') +
      '<div class="exercise-header"><span class="exercise-name" style="color:var(--green)">' + ex.name + occLabel + '</span>' +
      '<div style="display:flex;align-items:center">' + '<div class="done-check" onclick="toggleDone(\'' + cid + '\')"></div>' + repeatBtn + '</div></div>' +
      (ex.plan ? '<div class="exercise-plan">' + ex.plan + (ex.last ? ' &bull; Last: ' + ex.last : '') + '</div>' : '') +
      '<div class="cardio-inputs">' +
      '<div class="input-group"><label>Distance</label><input type="text" id="in-' + cid + '-dist" placeholder="' +
      ((ex.defaults && ex.defaults.distance) || 'distance') + '" value="' + ((ex.defaults && ex.defaults.distance) || '') + '" inputmode="text" oninput="saveState()"></div>' +
      '<div class="input-group"><label>Time</label><input type="text" id="in-' + cid + '-time" placeholder="mm:ss" inputmode="text" oninput="saveState()"></div>' +
      '</div>' +
      '<textarea class="notes-area" id="in-' + cid + '-note" placeholder="Notes (optional)" rows="1" oninput="saveState()"></textarea></div>';
  } else if (ex.type === 'stretch') {
    html += '<div class="exercise-card stretch' + (isAdded ? ' added' : '') + '" id="card-' + cid + '">' +
      (isAdded ? '<button class="remove-btn" onclick="removeAdded(\'' + day.id + '\',' + idx + ')">\u00d7</button>' : '') +
      '<div class="exercise-header"><span class="exercise-name" style="color:#1ABC9C">' + ex.name + occLabel + '</span>' +
      '<div style="display:flex;align-items:center">' + '<div class="done-check" onclick="toggleDone(\'' + cid + '\')"></div>' + repeatBtn + '</div></div>' +
      '<div class="exercise-plan">' + ex.plan +
      (ex.note ? ' &bull; <span style="color:var(--dim)">' + ex.note + '</span>' : '') + '</div>' +
      '<div class="input-row" style="justify-content:space-between;align-items:center">' +
      '<div style="padding-top:0"><div class="rest-timer" onclick="startTimer(this,' + (ex.rest||30) + ')">' +
      (ex.rest||30) + 's</div></div>' +
      '</div>' +
      '<textarea class="notes-area" id="in-' + cid + '-note" placeholder="Notes (optional)" rows="1" oninput="saveState()"></textarea></div>';
  } else {
    var pb = ex.progress ? '<span style="color:var(--green);font-size:11px;font-weight:700"> &#8593; UP</span>' : '';
    html += '<div class="exercise-card' + (isAdded ? ' added' : '') + '" id="card-' + cid + '" style="border-left-color:' + day.color + '">' +
      (isAdded ? '<button class="remove-btn" onclick="removeAdded(\'' + day.id + '\',' + idx + ')">\u00d7</button>' : '') +
      '<div class="exercise-header"><span class="exercise-name">' + ex.name + occLabel + pb + '</span>' +
      '<div style="display:flex;align-items:center">' + '<div class="done-check" onclick="toggleDone(\'' + cid + '\')"></div>' + repeatBtn + '</div></div>';
    if (ex.plan) {
      html += '<div class="exercise-plan">Target: ' + ex.plan + (ex.last ? ' &bull; Last: ' + ex.last : '') +
        (ex.note ? '<br><span class="' + (ex.progress ? 'prog' : '') + '">' + ex.note + '</span>' : '') + '</div>';
    }
    html += '<div class="input-row">' +
      '<div class="input-group" style="flex:1.2"><label>Weight</label>' +
      '<input type="text" id="in-' + cid + '-wt" placeholder="' + ((ex.target && ex.target.weight) || 'BW') + '"' +
      ' value="' + ((ex.target && ex.target.weight) || '') + '" inputmode="text" oninput="saveState()"></div>' +
      '<div class="input-group" style="flex:0.8"><label>Reps</label>' +
      '<input type="text" id="in-' + cid + '-reps" placeholder="e.g. 12x3" inputmode="text" oninput="saveState()"></div>' +
      '<div style="padding-top:18px"><div class="rest-timer" onclick="startTimer(this,' + (ex.rest || 60) + ')">' +
      (ex.rest || 60) + 's</div></div></div>' +
      '<textarea class="notes-area" id="in-' + cid + '-note" placeholder="Notes (optional)" rows="1" oninput="saveState()"></textarea></div>';
  }
  return html;
}

function renderDay(day) {
  var panel = document.getElementById('panel-' + day.id);
  if (!panel) {
    panel = document.createElement('div');
    panel.className = 'day-panel';
    panel.id = 'panel-' + day.id;
    panelsEl.appendChild(panel);
  }
  var allEx = day.exercises.concat(day._added || []);
  var html = '<div class="day-title" style="color:' + day.color + '">' + day.title + '</div>' +
    '<div class="day-subtitle">' + day.subtitle + '</div>' +
    '<div class="progress-bar"><div class="progress-fill" id="progress-' + day.id + '" style="width:0%"></div></div>' +
    '<div style="height:12px"></div>';
  var hasStretches = allEx.some(function(e) { return e.type === 'stretch'; });
  var stretchShown = false, workoutShown = false;
  allEx.forEach(function(ex, i) {
    if (hasStretches && ex.type === 'stretch' && !stretchShown) {
      html += '<div class="section-divider">MORNING STRETCHES (~20-30 min)</div>';
      stretchShown = true;
    }
    if (hasStretches && ex.type !== 'stretch' && !workoutShown) {
      html += '<div class="section-divider workout">WORKOUT</div>';
      workoutShown = true;
    }
    var isAdded = i >= day.exercises.length;
    html += buildExerciseHTML(day, ex, i, isAdded);
  });
  html += '<button class="add-activity-btn" onclick="openAddModal(\'' + day.id + '\')">+ Add Activity</button>';
  panel.innerHTML = html;
}

// Build tabs
DAYS.forEach(function(day) {
  day._added = [];
  var tab = document.createElement('div');
  tab.className = 'tab';
  tab.setAttribute('data-day', day.id);
  tab.innerHTML = day.label + '<span class="day-label">' + day.date + '</span>';
  tab.style.borderBottomColor = 'transparent';
  tab.onclick = (function(d) { return function() { switchDay(d); }; })(day.id);
  tabsEl.appendChild(tab);
});

// Load saved state, restore swap and added exercises first, then render
var savedState = loadState();
if (savedState && savedState.swap) {
  activeSwap = savedState.swap;
  applySwap(activeSwap.day1, activeSwap.day2);
}
if (savedState && savedState.added) {
  DAYS.forEach(function(day) {
    if (savedState.added[day.id]) day._added = savedState.added[day.id];
  });
}

DAYS.forEach(function(day) { renderDay(day); });
updateSwapInfo();

// Restore input values and checked states
if (savedState) {
  Object.keys(savedState.inputs || {}).forEach(function(id) {
    var el = document.getElementById(id);
    if (el) el.value = savedState.inputs[id];
  });
  Object.keys(savedState.checked || {}).forEach(function(cid) {
    var card = document.getElementById('card-' + cid);
    if (card) {
      var chk = card.querySelector('.done-check');
      if (chk) { chk.classList.add('checked'); card.classList.add('completed'); }
    }
  });
}

// ============================================================
// ADD ACTIVITY
// ============================================================
var addTargetDay = null;

function buildCurrentCatalog() {
  // Rebuild catalog from DAYS exercises + extras + any custom _added exercises
  var catalog = [];
  var seen = {};
  DAYS.forEach(function(day) {
    var allEx = day.exercises.concat(day._added || []);
    allEx.forEach(function(ex) {
      if (seen[ex.logName]) return;
      seen[ex.logName] = true;
      catalog.push({ name: ex.name, logName: ex.logName, type: ex.type || 'strength', rest: ex.rest || 60 });
    });
  });
  EXTRA_CARDIO.forEach(function(ex) {
    if (!seen[ex.logName]) { seen[ex.logName] = true; catalog.push(ex); }
  });
  EXTRA_STRENGTH.forEach(function(ex) {
    if (!seen[ex.logName]) { seen[ex.logName] = true; catalog.push(ex); }
  });
  DISCOVERED_EXERCISES.forEach(function(ex) {
    if (ex && ex.logName && !seen[ex.logName]) { seen[ex.logName] = true; catalog.push(ex); }
  });
  catalog.sort(function(a, b) {
    if (a.type === 'cardio' && b.type !== 'cardio') return -1;
    if (a.type !== 'cardio' && b.type === 'cardio') return 1;
    return a.name.localeCompare(b.name);
  });
  return catalog;
}

function openAddModal(dayId) {
  addTargetDay = dayId;
  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { day = DAYS[d]; break; } }
  document.getElementById('modalTitle').textContent = 'Add to ' + (day ? day.label : dayId);
  // Build dropdown with current catalog (includes custom-added exercises)
  EX_CATALOG = buildCurrentCatalog();
  var sel = document.getElementById('addExercise');
  sel.innerHTML = '<option value="">-- Select exercise --</option>';
  var lastType = '';
  EX_CATALOG.forEach(function(ex, i) {
    var groupLabel = ex.type === 'cardio' ? 'Cardio' : (ex.type === 'stretch' ? 'Stretch' : 'Strength');
    if (groupLabel !== lastType) {
      if (lastType) sel.innerHTML += '</optgroup>';
      sel.innerHTML += '<optgroup label="' + groupLabel + '">';
      lastType = groupLabel;
    }
    sel.innerHTML += '<option value="' + i + '">' + ex.name + '</option>';
  });
  if (lastType) sel.innerHTML += '</optgroup>';
  sel.innerHTML += '<optgroup label="Other"><option value="custom">Custom exercise...</option></optgroup>';
  sel.value = '';
  document.getElementById('customFields').style.display = 'none';
  document.getElementById('addRest').value = '60';
  document.getElementById('addRestGroup').style.display = '';
  document.getElementById('addModal').classList.add('show');
}

function onAddExerciseChange() {
  var val = document.getElementById('addExercise').value;
  var customFields = document.getElementById('customFields');
  var restGroup = document.getElementById('addRestGroup');
  if (val === 'custom') {
    customFields.style.display = 'block';
    restGroup.style.display = '';
    document.getElementById('addName').value = '';
    document.getElementById('addLogName').value = '';
    document.getElementById('addType').value = 'strength';
    document.getElementById('addRest').value = '60';
  } else if (val !== '') {
    customFields.style.display = 'none';
    var ex = EX_CATALOG[parseInt(val)];
    if (ex.type === 'cardio') {
      restGroup.style.display = 'none';
    } else {
      restGroup.style.display = '';
      document.getElementById('addRest').value = ex.rest || 60;
    }
  } else {
    customFields.style.display = 'none';
    restGroup.style.display = '';
  }
}

function closeAddModal() {
  document.getElementById('addModal').classList.remove('show');
  addTargetDay = null;
}

function confirmAdd() {
  var selVal = document.getElementById('addExercise').value;
  var type, name, logName, rest;
  if (selVal === 'custom') {
    type = document.getElementById('addType').value;
    name = document.getElementById('addName').value.trim();
    logName = document.getElementById('addLogName').value.trim() || name.toLowerCase();
    rest = parseInt(document.getElementById('addRest').value) || 60;
    if (!name) { document.getElementById('addName').focus(); return; }
  } else if (selVal !== '') {
    var catEx = EX_CATALOG[parseInt(selVal)];
    type = catEx.type || 'strength';
    name = catEx.name;
    logName = catEx.logName;
    rest = parseInt(document.getElementById('addRest').value) || catEx.rest || 60;
  } else {
    return; // Nothing selected
  }

  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === addTargetDay) { day = DAYS[d]; break; } }
  if (!day) return;

  var newEx;
  if (type === 'cardio') {
    newEx = { type: 'cardio', name: name, logName: logName, fields: ['distance', 'time'], defaults: {} };
  } else if (type === 'stretch') {
    newEx = { type: 'stretch', name: name, logName: logName, plan: '', note: '', target: { sets: '2', reps: '30s' }, rest: rest || 30 };
  } else {
    newEx = { name: name, logName: logName, target: { weight: '', sets: '', reps: '' }, rest: rest };
  }
  day._added.push(newEx);
  var liveState = captureState();
  renderDay(day);
  restoreState(liveState);
  updateProgress(day.id);
  saveState();
  closeAddModal();
}

function removeAdded(dayId, idx) {
  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { day = DAYS[d]; break; } }
  if (!day) return;
  var addedIdx = idx - day.exercises.length;
  if (addedIdx >= 0 && addedIdx < day._added.length) {
    var liveState = captureState();
    day._added.splice(addedIdx, 1);
    renderDay(day);
    restoreState(liveState);
    updateProgress(dayId);
    saveState();
  }
}

// ============================================================
// REPEAT EXERCISE (add another occurrence)
// ============================================================
function repeatExercise(dayId, idx) {
  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { day = DAYS[d]; break; } }
  if (!day) return;
  var allEx = day.exercises.concat(day._added || []);
  var ex = allEx[idx];
  if (!ex) return;
  var newEx;
  if (ex.type === 'cardio') {
    newEx = { type: 'cardio', name: ex.name, logName: ex.logName, fields: ex.fields || ['distance', 'time'], defaults: {} };
  } else if (ex.type === 'stretch') {
    newEx = { type: 'stretch', name: ex.name, logName: ex.logName, plan: ex.plan || '', note: ex.note || '', target: { sets: (ex.target && ex.target.sets) || '2', reps: (ex.target && ex.target.reps) || '30s' }, rest: ex.rest || 30 };
  } else {
    newEx = { name: ex.name, logName: ex.logName, target: { weight: (ex.target && ex.target.weight) || '', sets: (ex.target && ex.target.sets) || '', reps: (ex.target && ex.target.reps) || '' }, rest: ex.rest || 60 };
  }
  day._added.push(newEx);
  var liveState = captureState();
  renderDay(day);
  restoreState(liveState);
  updateProgress(dayId);
  saveState();
}

// ============================================================
// SWAP DAYS
// ============================================================
function openSwapModal() {
  var sel1 = document.getElementById('swapDay1');
  var sel2 = document.getElementById('swapDay2');
  sel1.innerHTML = ''; sel2.innerHTML = '';
  DAYS.forEach(function(day) {
    sel1.innerHTML += '<option value="' + day.id + '">' + day.label + ' \u2014 ' + day.subtitle.split('\u2014')[0].trim() + '</option>';
    sel2.innerHTML += '<option value="' + day.id + '">' + day.label + ' \u2014 ' + day.subtitle.split('\u2014')[0].trim() + '</option>';
  });
  // Default: Wed and Sat
  sel1.value = 'wed'; sel2.value = 'sat';
  document.getElementById('swapModal').classList.add('show');
}

function closeSwapModal() {
  document.getElementById('swapModal').classList.remove('show');
}

function applySwap(day1Id, day2Id) {
  var d1 = null, d2 = null;
  for (var i = 0; i < DAYS.length; i++) {
    if (DAYS[i].id === day1Id) d1 = DAYS[i];
    if (DAYS[i].id === day2Id) d2 = DAYS[i];
  }
  if (!d1 || !d2 || d1 === d2) return;
  // Swap exercises, subtitles, and colors
  var tmpEx = d1.exercises; d1.exercises = d2.exercises; d2.exercises = tmpEx;
  var tmpSub = d1.subtitle; d1.subtitle = d2.subtitle; d2.subtitle = tmpSub;
  var tmpColor = d1.color; d1.color = d2.color; d2.color = tmpColor;
  // Clear added exercises for swapped days
  d1._added = []; d2._added = [];
}

function confirmSwap() {
  var day1Id = document.getElementById('swapDay1').value;
  var day2Id = document.getElementById('swapDay2').value;
  if (day1Id === day2Id) { closeSwapModal(); return; }
  // Undo any existing swap first
  if (activeSwap) undoSwap(true);
  activeSwap = { day1: day1Id, day2: day2Id };
  applySwap(day1Id, day2Id);
  // Re-render both days and update UI
  DAYS.forEach(function(day) { renderDay(day); });
  updateSwapInfo();
  saveState();
  closeSwapModal();
  // Switch to first swapped day
  switchDay(day1Id);
}

function undoSwap(silent) {
  if (!activeSwap) return;
  // Swap back
  applySwap(activeSwap.day1, activeSwap.day2);
  activeSwap = null;
  DAYS.forEach(function(day) { renderDay(day); });
  updateSwapInfo();
  if (!silent) saveState();
  switchDay(getTodayTab());
}

function updateSwapInfo() {
  var info = document.getElementById('swapInfo');
  if (activeSwap) {
    var n1 = '', n2 = '';
    for (var i = 0; i < DAYS.length; i++) {
      if (DAYS[i].id === activeSwap.day1) n1 = DAYS[i].label;
      if (DAYS[i].id === activeSwap.day2) n2 = DAYS[i].label;
    }
    info.innerHTML = n1 + ' \u21c4 ' + n2 + ' swapped <button class="swap-undo" onclick="undoSwap()">Undo</button>';
  } else {
    info.innerHTML = '';
  }
}

// ============================================================
// CORE FUNCTIONS
// ============================================================
function getTodayTab() {
  var dow = new Date().getDay();
  var map = { 1: 'mon', 2: 'tue', 3: 'wed', 4: 'thu', 5: 'fri', 6: 'sat', 0: 'sun' };
  return map[dow] || 'mon';
}

function switchDay(dayId) {
  var tabs = document.querySelectorAll('.tab');
  for (var t = 0; t < tabs.length; t++) { tabs[t].classList.remove('active'); tabs[t].style.borderBottomColor = 'transparent'; }
  var panels = document.querySelectorAll('.day-panel');
  for (var p = 0; p < panels.length; p++) { panels[p].classList.remove('active'); }
  var tab = document.querySelector('.tab[data-day="' + dayId + '"]');
  var dayData = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { dayData = DAYS[d]; break; } }
  if (tab && dayData) { tab.classList.add('active'); tab.style.borderBottomColor = dayData.color; }
  var panel = document.getElementById('panel-' + dayId);
  if (panel) panel.classList.add('active');
  updateProgress(dayId);
}

function toggleDone(cid) {
  var card = document.getElementById('card-' + cid);
  var chk = card.querySelector('.done-check');
  chk.classList.toggle('checked');
  card.classList.toggle('completed');
  updateProgress(cid.split('-')[0]);
  saveState();
}

function updateProgress(dayId) {
  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { day = DAYS[d]; break; } }
  if (!day) return;
  var allEx = day.exercises.concat(day._added || []);
  var total = allEx.length, done = 0;
  for (var i = 0; i < total; i++) {
    var c = document.querySelector('#card-' + dayId + '-' + i + ' .done-check');
    if (c && c.classList.contains('checked')) done++;
  }
  var bar = document.getElementById('progress-' + dayId);
  if (bar) bar.style.width = (total > 0 ? done / total * 100 : 0) + '%';
}

var activeTimers = {};
function startTimer(el, sec) {
  var key = el.id || (el.id = 't' + Math.random());
  if (activeTimers[key]) {
    clearInterval(activeTimers[key]); delete activeTimers[key];
    el.textContent = sec + 's'; el.className = 'rest-timer'; return;
  }
  var rem = sec;
  el.textContent = rem + 's'; el.className = 'rest-timer running';
  activeTimers[key] = setInterval(function() {
    rem--;
    el.textContent = rem + 's';
    if (rem <= 0) {
      clearInterval(activeTimers[key]); delete activeTimers[key];
      el.textContent = 'GO!'; el.className = 'rest-timer done';
      if (navigator.vibrate) navigator.vibrate([200, 100, 200]);
      setTimeout(function() { el.textContent = sec + 's'; el.className = 'rest-timer'; }, 3000);
    }
  }, 1000);
}

// ============================================================
// COPY / EXPORT
// ============================================================
function gv(id) { var el = document.getElementById(id); return el ? el.value.trim() : ''; }

function getExForCard(dayId, idx) {
  var day = null;
  for (var d = 0; d < DAYS.length; d++) { if (DAYS[d].id === dayId) { day = DAYS[d]; break; } }
  if (!day) return null;
  var allEx = day.exercises.concat(day._added || []);
  return idx < allEx.length ? allEx[idx] : null;
}

function copyResults() {
  var text = '';
  DAYS.forEach(function(day) {
    var dt = '';
    var grouped = {};
    var order = [];
    var allEx = day.exercises.concat(day._added || []);
    allEx.forEach(function(ex, i) {
      var cid = day.id + '-' + i;
      var chk = document.querySelector('#card-' + cid + ' .done-check');
      if (!chk || !chk.classList.contains('checked')) return;
      if (!grouped[ex.logName]) { grouped[ex.logName] = { entries: [], type: ex.type }; order.push(ex.logName); }
      if (ex.type === 'cardio') {
        var ds = gv('in-' + cid + '-dist'), tm = gv('in-' + cid + '-time');
        if (ds || tm) grouped[ex.logName].entries.push(ds + (tm ? ' / ' + tm : ''));
      } else {
        var wt = gv('in-' + cid + '-wt'), rp = gv('in-' + cid + '-reps'), nt = gv('in-' + cid + '-note');
        if (wt || rp) grouped[ex.logName].entries.push(wt + (rp ? ' \u2014 ' + rp : '') + (nt ? ' (' + nt + ')' : ''));
      }
    });
    order.forEach(function(logName) {
      var g = grouped[logName];
      if (g.entries.length) dt += '  ' + logName + ': ' + g.entries.join('  +  ') + '\n';
    });
    if (dt) text += day.title + '\n' + dt + '\n';
  });
  if (!text) text = 'No exercises logged yet.';
  copyText(text);
}

function copyForExcel() {
  var lines = [];
  DAYS.forEach(function(day) {
    // Group checked exercises by logName to accumulate multiple occurrences
    var grouped = {};  // logName -> { vals: [{col3,col4,col5}], type }
    var order = [];    // preserve first-seen order
    var allEx = day.exercises.concat(day._added || []);
    allEx.forEach(function(ex, i) {
      var cid = day.id + '-' + i;
      var chk = document.querySelector('#card-' + cid + ' .done-check');
      if (!chk || !chk.classList.contains('checked')) return;
      var col3, col4, col5;
      if (ex.type === 'stretch') {
        col3 = '\u2713'; col4 = ''; col5 = gv('in-' + cid + '-note');
      } else if (ex.type === 'cardio') {
        col3 = gv('in-' + cid + '-dist'); col4 = gv('in-' + cid + '-time'); col5 = gv('in-' + cid + '-note');
      } else {
        col3 = gv('in-' + cid + '-wt'); col4 = gv('in-' + cid + '-reps'); col5 = gv('in-' + cid + '-note');
      }
      if (!col3 && !col4) return;
      if (!grouped[ex.logName]) { grouped[ex.logName] = { vals: [], type: ex.type }; order.push(ex.logName); }
      grouped[ex.logName].vals.push({ c3: col3, c4: col4, c5: col5 });
    });
    order.forEach(function(logName) {
      var g = grouped[logName];
      var c3, c4, c5;
      if (g.vals.length === 1) {
        c3 = g.vals[0].c3; c4 = g.vals[0].c4; c5 = g.vals[0].c5;
      } else {
        c3 = g.vals.map(function(v) { return v.c3; }).filter(Boolean).join('; ');
        c4 = g.vals.map(function(v) { return v.c4; }).filter(Boolean).join('; ');
        var notes = g.vals.map(function(v) { return v.c5; }).filter(Boolean);
        c5 = notes.join('; ');
      }
      lines.push(day.isoDate + '|' + logName + '|' + c3 + '|' + c4 + '|' + c5);
    });
  });
  if (!lines.length) { copyText('No exercises logged yet.'); return; }
  copyText(lines.join('\n'));
}

function copyText(text) {
  if (navigator.clipboard && navigator.clipboard.writeText)
    navigator.clipboard.writeText(text).then(showToast).catch(fb);
  else fb();
  function fb() {
    var ta = document.createElement('textarea'); ta.value = text;
    ta.style.position = 'fixed'; ta.style.left = '-9999px';
    document.body.appendChild(ta); ta.select(); ta.setSelectionRange(0, 99999);
    document.execCommand('copy'); document.body.removeChild(ta); showToast();
  }
}

function showToast() {
  var t = document.getElementById('toast'); t.classList.add('show');
  setTimeout(function() { t.classList.remove('show'); }, 2000);
}

// ============================================================
// INIT
// ============================================================
switchDay(getTodayTab());
DAYS.forEach(function(day) { updateProgress(day.id); });

// Register service worker (works when hosted on HTTPS)
if ('serviceWorker' in navigator) {
  navigator.serviceWorker.register('sw.js').catch(function() {});
}
</script>
</body>
</html>
'''


# ─────────────────────────────────────────────
# GENERATE XLSX
# ─────────────────────────────────────────────
def generate_xlsx(week_dates, recent_data, output_path):
    wb = Workbook()
    HEADER_BG = "2D2D44"
    PROGRESS_UP = "27AE60"
    HOLD_COLOR = "F39C12"
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    date_map = {}
    for i, (day_id, _, _, _, _) in enumerate(WEEK):
        date_map[day_id] = week_dates[i]

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Week Summary"
    ws_sum.column_dimensions['A'].width = 4
    ws_sum.column_dimensions['B'].width = 26
    ws_sum.column_dimensions['C'].width = 16
    ws_sum.column_dimensions['D'].width = 16
    ws_sum.column_dimensions['E'].width = 34

    d0, d1 = week_dates[0], week_dates[-1]
    ws_sum.merge_cells('A1:E1')
    ws_sum['A1'].value = f"Week of {d0.strftime('%B %-d')}\u2013{d1.strftime('%-d, %Y')} \u2014 Progression Summary"
    ws_sum['A1'].font = Font(name="Arial", bold=True, size=14)

    for j, h in enumerate(["", "Exercise", "Previous", "This Week", "Change"], 1):
        c = ws_sum.cell(row=3, column=j, value=h)
        c.font = hdr_font
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    sum_row = 4
    seen = set()
    for _, _, _, _, exercises in WEEK:
        for display, log_name, sets, reps_target, base_wt, prog_rule, rest, is_db in exercises:
            if log_name in seen:
                continue
            target_wt, wt_fmt, note, is_prog, last_str = compute_target(
                log_name, is_db, base_wt, prog_rule, recent_data)
            if not is_prog and "Hold" not in note:
                continue
            seen.add(log_name)
            arrow = "\u2191" if is_prog else "\u2192"
            color = PROGRESS_UP if is_prog else HOLD_COLOR
            ws_sum.cell(row=sum_row, column=1, value=arrow).font = Font(name="Arial", bold=True, color=color)
            ws_sum.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=sum_row, column=2, value=display).font = Font(name="Arial", size=10, bold=True)
            ws_sum.cell(row=sum_row, column=3, value=last_str).font = Font(name="Arial", size=10)
            ws_sum.cell(row=sum_row, column=3).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=sum_row, column=4, value=f"{wt_fmt} \u00d7 {sets}\u00d7{reps_target}").font = Font(name="Arial", size=10, bold=True)
            ws_sum.cell(row=sum_row, column=4).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=sum_row, column=5, value=note).font = Font(name="Arial", size=10, color=color)
            for j in range(1, 6):
                ws_sum.cell(row=sum_row, column=j).border = thin_border
            sum_row += 1

    # Day sheets
    for day_id, day_name, subtitle, color, exercises in WEEK:
        d = date_map[day_id]
        ws = wb.create_sheet(f"{day_id.capitalize()[:3]} {d.strftime('%-m-%-d')}")
        ws.sheet_properties.tabColor = color.lstrip("#")
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 36

        ws.merge_cells('A1:G1')
        ws['A1'].value = f"{day_name} \u2014 {d.strftime('%B %-d')}"
        ws['A1'].font = Font(name="Arial", bold=True, size=14, color=color.lstrip("#"))
        ws.merge_cells('A2:G2')
        ws['A2'].value = subtitle
        ws['A2'].font = Font(name="Arial", size=10, italic=True, color="666666")

        headers = ["#", "Exercise", "Sets \u00d7 Reps", "Weight Target", "Last Session", "Rest", "Notes / Progression"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=j, value=h)
            c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=color.lstrip("#"))
            c.alignment = Alignment(horizontal="center")

        row = 4
        if day_id in CARDIO:
            cname, clog, cdist = CARDIO[day_id]
            clast, _ = get_last_cardio(clog, recent_data)
            ws.cell(row=row, column=2, value=cname).font = Font(name="Arial", bold=True, color="2ECC71")
            ws.cell(row=row, column=3, value=cdist)
            ws.cell(row=row, column=5, value=clast).font = Font(name="Arial", size=9, italic=True, color="888888")
            for j in range(1, 8):
                ws.cell(row=row, column=j).border = thin_border
                ws.cell(row=row, column=j).alignment = Alignment(horizontal="center")
            row += 1

        if day_id in VEST_WALK_DAYS:
            vlast, _ = get_last_cardio("vest walk", recent_data)
            ws.cell(row=row, column=2, value="Vest walk").font = Font(name="Arial", bold=True, color="2ECC71")
            ws.cell(row=row, column=3, value="5-8 mi")
            ws.cell(row=row, column=5, value=vlast).font = Font(name="Arial", size=9, italic=True, color="888888")
            for j in range(1, 8):
                ws.cell(row=row, column=j).border = thin_border
                ws.cell(row=row, column=j).alignment = Alignment(horizontal="center")
            row += 1

        for idx, (display, log_name, sets, reps_target, base_wt, prog_rule, rest, is_db) in enumerate(exercises, 1):
            target_wt, wt_fmt, note, is_prog, last_str = compute_target(
                log_name, is_db, base_wt, prog_rule, recent_data)
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=display).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3, value=f"{sets} \u00d7 {reps_target}").font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            wt_display = wt_fmt + (" \u2191" if is_prog else "")
            ws.cell(row=row, column=4, value=wt_display if target_wt > 0 else "Bodyweight")
            ws.cell(row=row, column=4).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=last_str).font = Font(name="Arial", size=9, italic=True, color="888888")
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            rest_str = f"{rest // 60} min" if rest >= 120 else f"{rest}s"
            ws.cell(row=row, column=6, value=rest_str).alignment = Alignment(horizontal="center")
            note_cell = ws.cell(row=row, column=7, value=note)
            note_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            note_cell.font = Font(name="Arial", size=9, bold=True, color=PROGRESS_UP) if is_prog else Font(name="Arial", size=9, color=HOLD_COLOR)
            for j in range(1, 8):
                ws.cell(row=row, column=j).border = thin_border
            row += 1

    wb.save(output_path)
    print(f"  \u2713 {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 monday_setup.py <Weight2026_DEXA_calibratedc.xlsx> [output_dir]")
        print("\nGenerates gym_tracker.html and Weekly_Plan_<dates>.xlsx")
        print("Output goes to output_dir (default: current directory).")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.getcwd()
    os.makedirs(out_dir, exist_ok=True)

    print("Reading recent exercise data...")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['summary']
    recent = read_recent_data(ws)
    print(f"  Found data for {len(recent)} exercises in the last 14 days.")

    # Load stretches early so we can exclude their logNames from auto-discovery
    script_dir = os.path.dirname(os.path.abspath(__file__))
    stretches = load_stretches(script_dir)
    stretch_names = set()
    if stretches:
        for entry in stretches.get("core_routine", []):
            if "logName" in entry: stretch_names.add(entry["logName"])
        for entry in stretches.get("balance_block", []):
            if "logName" in entry: stretch_names.add(entry["logName"])

    discovered = scan_extra_exercises(ws, exclude=stretch_names)
    if discovered:
        print(f"  Auto-discovered {len(discovered)} new exercise(s) from summary columns:")
        for ex in discovered:
            print(f"    + {ex['logName']} ({ex['type']})")
    print()

    today = datetime.now().date()
    monday = today - timedelta(days=today.weekday())
    week_dates = [monday + timedelta(days=i) for i in range(7)]

    print(f"Generating plan for {monday.strftime('%B %-d')} \u2013 {week_dates[-1].strftime('%B %-d, %Y')}")
    print()

    html_path = os.path.join(out_dir, "index.html")
    sw_path = os.path.join(out_dir, "sw.js")
    xlsx_path = os.path.join(out_dir, f"Weekly_Plan_{monday.strftime('%b%-d')}-{week_dates[-1].strftime('%-d')}.xlsx")

    mods = load_modifications(script_dir)

    print("Generating files:")
    generate_html(week_dates, recent, html_path, stretches=stretches, mods=mods, extra_exercises=discovered)
    generate_xlsx(week_dates, recent, xlsx_path)

    # Generate sw.js with date-stamped cache version so the service worker
    # invalidates the old cache each time we push a new plan.
    cache_version = f"gym-tracker-{monday.strftime('%Y%m%d')}"
    sw_js = f"""var CACHE_NAME = '{cache_version}';
var URLS_TO_CACHE = ['./index.html', './manifest.json'];

self.addEventListener('install', function(e) {{
  e.waitUntil(caches.open(CACHE_NAME).then(function(cache) {{ return cache.addAll(URLS_TO_CACHE); }}));
  self.skipWaiting();
}});

self.addEventListener('activate', function(e) {{
  e.waitUntil(caches.keys().then(function(names) {{
    return Promise.all(names.filter(function(n) {{ return n !== CACHE_NAME; }}).map(function(n) {{ return caches.delete(n); }}));
  }}));
  self.clients.claim();
}});

self.addEventListener('fetch', function(e) {{
  e.respondWith(fetch(e.request).catch(function() {{ return caches.match(e.request); }}));
}});
"""
    with open(sw_path, "w") as f:
        f.write(sw_js)
    print(f"  ✓ {sw_path} (cache: {cache_version})")

    print(f"\nDone! To deploy to GitHub Pages:")
    print(f"  git add index.html sw.js && git commit -m 'Update plan {monday.strftime('%b %-d')}' && git push")
    print(f"\nAfter each session, use 'Copy for Excel' then run:")
    print(f"  python3 import_gym_log.py <paste_file.txt> {excel_path}")
