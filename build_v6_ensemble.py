#!/usr/bin/env python3
"""
Build V6: Dual-scale ensemble calibration.
  Oxiline + Fitindex → independent EMA → independent DEXA calibration
  → metric-specific weighted blend → Weight decorrelation → Change from baseline

Ensemble strategy (from error-correlation analysis):
  FAT metrics (Fat%, F_mas, V_fat, all limb/trunk fat): Fitindex-only or FI-heavy
  MUSCLE metrics (Lb_mas, limb muscle): weighted Oxi+FI ensemble (errors uncorrelated)
  Weight: either scale (near-identical)

Usage:
  python3 build_v6_ensemble.py <input.xlsx> [output.xlsx]

If output is omitted, writes to <input_basename>_built.xlsx in the same directory.
"""
import sys
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Parse command-line arguments
if len(sys.argv) >= 2 and not sys.argv[1].startswith('-'):
    _input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        _output_file = sys.argv[2]
    else:
        _base_name = os.path.splitext(_input_file)[0]
        _output_file = f"{_base_name}_built.xlsx"
    SRC = _input_file
    CAL = _input_file
    import tempfile as _tf_arg
    DST = os.path.join(_tf_arg.mkdtemp(), '_build_temp.xlsx')
else:
    print("Usage: python3 build_v5.py <input.xlsx> [output.xlsx]")
    print("  input.xlsx  - The Weight2026 DEXA calibrated spreadsheet")
    print("  output.xlsx - Output path (default: <input>_built.xlsx)")
    sys.exit(1)

# Use the CALIBRATED file as base (it has the user's custom charts, scatter plots,
# data labels, etc.). The raw file has simple default charts that are useless.
# We'll patch any newer BIA data from the raw file INTO the calibrated copy.
_base = CAL if os.path.exists(CAL) else SRC
os.system(f'cp "{_base}" "{DST}" && chmod 644 "{DST}"')
print(f"Base file: {'calibrated' if _base == CAL else 'raw'}")

# ============================================================
# STEP 0: Patch BIA data — ensure DST has the latest from BOTH files
# ============================================================
# DST is a copy of CAL (to preserve charts). Patch any BIA data from SRC (raw)
# that CAL might be missing (e.g., if raw file has newer entries).
from datetime import datetime as _dt
_wb_src = load_workbook(SRC, data_only=True)
_wb_dst = load_workbook(DST)
# --- Patch oxiline sheet from SRC (raw) ---
_ws_src = _wb_src['oxiline']
_ws_dst = _wb_dst['oxiline']
BIA_COLS = list(range(2, 55))  # cols 2-54: all raw BIA measurement columns
patched_oxi = 0
for r in range(4, 400):
    dst_wt = _ws_dst.cell(r, 4).value
    src_wt = _ws_src.cell(r, 4).value
    if dst_wt is None and src_wt is not None and isinstance(src_wt, (int, float)):
        for c in BIA_COLS:
            v = _ws_src.cell(r, c).value
            if v is not None:
                _ws_dst.cell(r, c, v)
        patched_oxi += 1
# --- Patch summary sheet (exercise data) from SRC ---
_ws_src_s = _ws_src_s = _wb_src['summary']
_ws_dst_s = _wb_dst['summary']
# Build date→row mapping for SRC
src_date_rows = {}
for r in range(3, _ws_src_s.max_row + 1):
    d = _ws_src_s.cell(r, 3).value
    if isinstance(d, _dt):
        src_date_rows[d.date()] = r
patched_sum = 0
for r in range(3, _ws_dst_s.max_row + 1):
    d = _ws_dst_s.cell(r, 3).value
    if not isinstance(d, _dt):
        continue
    src_r = src_date_rows.get(d.date())
    if not src_r:
        continue
    # Check if dst row is empty but src has data (check exercise cols)
    dst_has = any(_ws_dst_s.cell(r, c).value for c in range(10, 60))
    src_has = any(_ws_src_s.cell(src_r, c).value for c in range(10, 60))
    if not dst_has and src_has:
        for c in range(1, _ws_src_s.max_column + 1):
            v = _ws_src_s.cell(src_r, c).value
            if v is not None:
                _ws_dst_s.cell(r, c, v)
        patched_sum += 1
_wb_src.close()
if patched_oxi or patched_sum:
    # Save patched workbook to a TEMP file (openpyxl save destroys charts)
    import tempfile as _tf0
    _patch_temp = os.path.join(_tf0.mkdtemp(), 'patched.xlsx')
    _wb_dst.save(_patch_temp)
    _wb_dst.close()
    # Inject only the modified sheet XMLs back into DST via ZIP surgery
    _patch_dir = _tf0.mkdtemp()
    import zipfile as _zf0
    with _zf0.ZipFile(DST, 'r') as z:
        z.extractall(_patch_dir)
    # Find which sheets to inject (oxiline and summary)
    _p_wb_xml = open(os.path.join(_patch_dir, 'xl', 'workbook.xml'), 'r').read()
    _p_rels = open(os.path.join(_patch_dir, 'xl', '_rels', 'workbook.xml.rels'), 'r').read()
    for sheet_name in ['oxiline', 'summary']:
        _p_rid_m = re.search(rf'name="{sheet_name}"[^>]*r:id="(rId\d+)"', _p_wb_xml) or \
                   re.search(rf'name="{sheet_name}"[^>]*id="(rId\d+)"', _p_wb_xml, re.IGNORECASE)
        if not _p_rid_m: continue
        _p_rid = _p_rid_m.group(1)
        _p_tgt_m = re.search(f'Id="{_p_rid}"[^>]*Target="([^"]+)"', _p_rels) or \
                   re.search(f'Target="([^"]+)"[^>]*Id="{_p_rid}"', _p_rels)
        if not _p_tgt_m: continue
        _p_sheet_path = _p_tgt_m.group(1).lstrip('/')
        if _p_sheet_path.startswith('xl/'): _p_sheet_path = _p_sheet_path[3:]
        # Get the same sheet from the temp file
        with _zf0.ZipFile(_patch_temp, 'r') as zt:
            _t_wb_xml = zt.read('xl/workbook.xml').decode('utf-8')
            _t_rels = zt.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            _t_rid_m = re.search(rf'name="{sheet_name}"[^>]*r:id="(rId\d+)"', _t_wb_xml) or \
                       re.search(rf'name="{sheet_name}"[^>]*id="(rId\d+)"', _t_wb_xml, re.IGNORECASE)
            if not _t_rid_m: continue
            _t_rid = _t_rid_m.group(1)
            _t_tgt_m = re.search(f'Id="{_t_rid}"[^>]*Target="([^"]+)"', _t_rels) or \
                       re.search(f'Target="([^"]+)"[^>]*Id="{_t_rid}"', _t_rels)
            if not _t_tgt_m: continue
            _t_path = _t_tgt_m.group(1).lstrip('/')
            if not _t_path.startswith('xl/'): _t_path = 'xl/' + _t_path
            _new_xml = zt.read(_t_path)
        # Write into original structure
        with open(os.path.join(_patch_dir, 'xl', _p_sheet_path), 'wb') as f:
            f.write(_new_xml)
    # Repackage DST
    with _zf0.ZipFile(DST, 'w', _zf0.ZIP_DEFLATED) as zout:
        for root, dirs, files in os.walk(_patch_dir):
            for fn in files:
                fp = os.path.join(root, fn)
                arcname = os.path.relpath(fp, _patch_dir)
                zout.write(fp, arcname)
    print(f"Patched {patched_oxi} oxiline rows, {patched_sum} summary rows from raw file (charts preserved)")
else:
    _wb_dst.close()
    print("No patching needed (CAL already has all data)")

# ============================================================
# STEP 1: Load BOTH scales, compute independent DEXA calibrations
# ============================================================

# --- 1a: Load Oxiline (primary scale) ---
df_raw = pd.read_excel(DST, 'oxiline', header=None)
data = df_raw.iloc[3:].copy().reset_index(drop=True)
data[2] = pd.to_datetime(data[2], errors='coerce')
for c in [3,5,6,7,8,9,12,13,19,20,21,22,23,24,25,26,27,28]:
    data[c] = pd.to_numeric(data[c], errors='coerce')

valid = data.dropna(subset=[3])

PCOLS = {
    'Weight': 3, 'Fat_pct': 7, 'F_mas': 8, 'Lb_mas': 9, 'V_fat': 13,
    'R_a_f': 19, 'L_a_f': 20, 'T_f': 21, 'R_l_f': 22, 'L_l_f': 23,
    'R_a_m': 24, 'L_a_m': 25, 'T_m': 26, 'R_l_m': 27, 'L_l_m': 28,
}

# --- 1b: Load Fitindex (second scale) ---
df_fi_raw = pd.read_excel(DST, 'Fitindex', header=None)
fi_data = df_fi_raw.iloc[1:].copy().reset_index(drop=True)  # row 1 = headers, data starts row 2
fi_data[2] = pd.to_datetime(fi_data[2], errors='coerce')
for c in [3,5,6,7,8,9,12,13,19,20,21,22,23,24,25,26,27,28]:
    fi_data[c] = pd.to_numeric(fi_data[c], errors='coerce')
fi_valid = fi_data.dropna(subset=[3])

# Fitindex uses same column layout as Oxiline (verified from sheet headers)
FI_PCOLS = dict(PCOLS)  # same column indices

print(f"Loaded Oxiline: {len(valid)} rows, Fitindex: {len(fi_valid)} rows")

# --- 1c: Build date-aligned index ---
# Both scales aligned to Oxiline's date index (primary)
oxi_dates = valid[2].values
fi_date_map = {}  # date -> fi_valid index
for i in range(len(fi_valid)):
    d = fi_valid[2].iloc[i]
    if pd.notna(d):
        fi_date_map[d] = i

weight = valid[3]
mean_weight = weight.mean()

# --- 1d: DEXA reference values ---
DEXA_DEC = {
    'Weight': 189.4, 'Fat_pct': 0.242, 'F_mas': 45.9, 'Lb_mas': 136.3,
    'V_fat': 2.44, 'R_a_f': 2.2, 'L_a_f': 2.5, 'T_f': 25.5, 'R_l_f': 6.7, 'L_l_f': 6.7,
    'R_a_m': 8.7, 'L_a_m': 7.9, 'T_m': 66.0, 'R_l_m': 23.3, 'L_l_m': 22.4,
}
DEXA_MAR = {
    'Weight': 191.1, 'Fat_pct': 0.236, 'F_mas': 45.1, 'Lb_mas': 138.6,
    'V_fat': 2.21, 'R_a_f': 2.2, 'L_a_f': 2.3, 'T_f': 23.6, 'R_l_f': 7.8, 'L_l_f': 7.1,
    'R_a_m': 8.6, 'L_a_m': 8.6, 'T_m': 66.0, 'R_l_m': 24.2, 'L_l_m': 23.7,
}

# --- 1e: Compute two-point calibration for OXILINE ---
# Read validated coefficients from original file if available
cal_oxi = {}
if os.path.exists(CAL):
    _wb_cal2 = load_workbook(CAL, data_only=True)
    _ws_cal2 = _wb_cal2['oxiline']
    metrics_order = list(PCOLS.keys())
    for i, m in enumerate(metrics_order):
        col = 71 + i
        a_val = _ws_cal2.cell(1, col).value
        b_val = _ws_cal2.cell(2, col).value
        if isinstance(a_val, (int, float)) and isinstance(b_val, (int, float)):
            cal_oxi[m] = (a_val, b_val)
        else:
            cal_oxi[m] = (1.0, 0.0)
    _wb_cal2.close()
    print(f"Loaded Oxiline DEXA calibration from original file ({len(cal_oxi)} metrics)")
else:
    fw_mask_oxi = (valid[2] >= '2025-12-13') & (valid[2] <= '2025-12-19')
    fw_idx_oxi = fw_mask_oxi[fw_mask_oxi].index
    mar7_idx_oxi = (valid[2] == '2026-03-07')[valid[2] == '2026-03-07'].index[0]
    for m in PCOLS:
        fw_avg = np.nanmean(valid[PCOLS[m]].loc[fw_idx_oxi].values)
        m7_val = valid[PCOLS[m]].loc[mar7_idx_oxi]
        dd, dm = DEXA_DEC[m], DEXA_MAR[m]
        if abs(m7_val - fw_avg) > 1e-6:
            a = (dm - dd) / (m7_val - fw_avg)
            b = dd - a * fw_avg
        else:
            a, b = 0.0, dd
        cal_oxi[m] = (round(a, 6), round(b, 4))
    print("Computed Oxiline DEXA calibration (two-point method)")

# --- 1f: Compute two-point calibration for FITINDEX ---
cal_fi = {}
fw_mask_fi = (fi_valid[2] >= '2025-12-13') & (fi_valid[2] <= '2025-12-19')
fw_idx_fi = fw_mask_fi[fw_mask_fi].index
m7_candidates = fi_valid[2] == '2026-03-07'
if m7_candidates.any():
    mar7_idx_fi = m7_candidates[m7_candidates].index[0]
    for m in FI_PCOLS:
        fw_vals = fi_valid[FI_PCOLS[m]].loc[fw_idx_fi]
        fw_avg = np.nanmean(fw_vals.values)
        m7_val = fi_valid[FI_PCOLS[m]].loc[mar7_idx_fi]
        dd, dm = DEXA_DEC[m], DEXA_MAR[m]
        if abs(m7_val - fw_avg) > 1e-6:
            a = (dm - dd) / (m7_val - fw_avg)
            b = dd - a * fw_avg
        else:
            a, b = 0.0, dd
        cal_fi[m] = (round(a, 6), round(b, 4))
    print(f"Computed Fitindex DEXA calibration (two-point method, {len(cal_fi)} metrics)")
else:
    print("WARNING: Fitindex has no Mar 7 data — falling back to Oxiline-only")
    cal_fi = dict(cal_oxi)

# --- 1g: Ensemble weights (from error-correlation analysis) ---
# w_oxi = fraction of Oxiline, w_fi = 1 - w_oxi
# Fat metrics: FI dominates (lower post-calibration noise)
# Muscle metrics: ensemble (uncorrelated errors)
# Weight: equal (both excellent)
ENSEMBLE_WEIGHTS = {
    # Metric:         w_oxi  (w_fi = 1 - w_oxi)
    'Weight':          0.50,
    'Fat_pct':         0.05,   # FI 95%
    'F_mas':           0.14,   # FI 86%
    'Lb_mas':          0.65,   # Oxi-heavy ensemble
    'V_fat':           0.03,   # FI 97%
    'R_a_f':           0.50,   # Both zero-change, equal
    'L_a_f':           0.03,   # FI 97%
    'T_f':             0.12,   # FI 88%
    'R_l_f':           0.08,   # FI 92%
    'L_l_f':           0.09,   # FI 91%
    'R_a_m':           0.72,   # Oxi-heavy ensemble (error r=-0.36)
    'L_a_m':           0.77,   # Oxi-only (FI not sensitive)
    'T_m':             0.50,   # Both zero-change, equal
    'R_l_m':           0.47,   # Near-equal ensemble (error r=+0.07)
    'L_l_m':           0.85,   # Oxi-heavy ensemble
}

print("\nEnsemble weights (w_Oxi / w_FI):")
for m in PCOLS:
    w = ENSEMBLE_WEIGHTS[m]
    print(f"  {m:<10s}: {w:.0%} Oxi / {1-w:.0%} FI")

# Keep cal as alias for Oxiline calibration (used by spreadsheet formula sections)
cal = dict(cal_oxi)

# ============================================================
# STEP 2: Compute EMA on BOTH scales, DEXA-calibrate independently,
#          then blend using ensemble weights, then regress on EMA weight
# ============================================================
alpha = 2 / (20 + 1)  # span=20, alpha≈0.0952

def compute_ema(values, alpha):
    e = np.full_like(values, np.nan, dtype=float)
    for i in range(len(values)):
        if np.isnan(values[i]): continue
        if i == 0 or np.isnan(e[i-1]): e[i] = values[i]
        else: e[i] = alpha * values[i] + (1-alpha) * e[i-1]
    return e

# --- 2a: Oxiline EMA + DEXA calibration ---
ema_oxi = {}
dx_oxi = {}
for m in PCOLS:
    ema_oxi[m] = compute_ema(valid[PCOLS[m]].values, alpha)
    a, b = cal_oxi[m]
    dx_oxi[m] = a * ema_oxi[m] + b

# --- 2b: Fitindex EMA + DEXA calibration (aligned to Oxiline dates) ---
ema_fi = {}
dx_fi = {}
for m in FI_PCOLS:
    # Build Fitindex values aligned to Oxiline's date array
    fi_aligned = np.full(len(valid), np.nan, dtype=float)
    for i in range(len(valid)):
        d = valid[2].iloc[i]
        if d in fi_date_map:
            fi_idx = fi_date_map[d]
            fi_aligned[i] = fi_valid[FI_PCOLS[m]].iloc[fi_idx]
    ema_fi[m] = compute_ema(fi_aligned, alpha)
    a, b = cal_fi[m]
    dx_fi[m] = a * ema_fi[m] + b

# --- 2c: Blend: dx_uncorrected = w_oxi * dx_oxi + w_fi * dx_fi ---
# When Fitindex data is missing (NaN), fall back to Oxiline-only
dx_uncorrected = {}
for m in PCOLS:
    w_oxi = ENSEMBLE_WEIGHTS[m]
    w_fi = 1.0 - w_oxi
    blended = np.full(len(valid), np.nan, dtype=float)
    for i in range(len(valid)):
        oxi_v = dx_oxi[m][i]
        fi_v = dx_fi[m][i]
        if np.isnan(oxi_v):
            blended[i] = np.nan
        elif np.isnan(fi_v):
            # Fitindex not available — use Oxiline only
            blended[i] = oxi_v
        else:
            blended[i] = w_oxi * oxi_v + w_fi * fi_v
    dx_uncorrected[m] = blended

# Also keep ema_raw referencing Oxiline EMA (used by spreadsheet formulas)
ema_raw = dict(ema_oxi)

# EMA of weight for regression (blended weight)
ema_weight = ema_oxi['Weight']  # Weight is near-identical between scales
mean_ema_weight = np.nanmean(ema_weight)

# --- 2d: Weight regression in DEXA space (Mprime-style decorrelation) ---
weight_slopes_dx = {}
for m in PCOLS:
    if m == 'Weight':
        weight_slopes_dx[m] = 0.0
        continue
    y = dx_uncorrected[m]
    x = ema_weight
    mask = ~(np.isnan(y) | np.isnan(x))
    slope = np.polyfit(x[mask], y[mask], 1)[0]
    weight_slopes_dx[m] = round(slope, 6)

print("\nWeight decorrelation slopes (DEXA space, ensemble):")
for m, s in weight_slopes_dx.items():
    if s != 0:
        print(f"  {m}: {s}")

# Corrected = dx - slope * (ema_weight - mean_weight)
dx_corrected = {}
for m in PCOLS:
    dx_corrected[m] = dx_uncorrected[m] - weight_slopes_dx[m] * (ema_weight - mean_ema_weight)

# --- 2e: Noise comparison: Oxiline-only vs Ensemble ---
print("\n=== Noise comparison: Oxi-only vs Ensemble (daily diff std in DEXA space) ===")
print(f"{'Metric':<10} {'Oxi-only':>12} {'Ensemble':>12} {'improvement':>12}")
for m in PCOLS:
    if m == 'Weight': continue
    oxi_only = np.nanstd(np.diff(dx_oxi[m][~np.isnan(dx_oxi[m])]))
    ens = np.nanstd(np.diff(dx_uncorrected[m][~np.isnan(dx_uncorrected[m])]))
    pct = (1 - ens/oxi_only) * 100 if oxi_only > 0 else 0
    print(f"  {m:<10} {oxi_only:12.4f} {ens:12.4f} {pct:11.1f}%")

# Verify correlation removal
print("\n=== Correlation with EMA weight (ensemble) ===")
print(f"{'Metric':<10} {'before':>10} {'after':>10}")
for m in ['R_a_m','L_a_m','T_m','R_l_m','L_l_m','Lb_mas','F_mas']:
    mask = ~(np.isnan(dx_uncorrected[m]) | np.isnan(ema_weight))
    r_before = np.corrcoef(dx_uncorrected[m][mask], ema_weight[mask])[0,1]
    r_after = np.corrcoef(dx_corrected[m][mask], ema_weight[mask])[0,1]
    print(f"  {m:<10} {r_before:10.3f} {r_after:10.3f}")

# Aggregate muscle change (ensemble DEXA-calibrated)
muscle_raw = dx_uncorrected['R_a_m'] + dx_uncorrected['L_a_m'] + dx_uncorrected['R_l_m'] + dx_uncorrected['L_l_m']
chg_raw = muscle_raw - 62.3
print(f"\n=== chg_Muscle (ensemble) ===")
print(f"  Range: {np.nanmin(chg_raw):+.2f} to {np.nanmax(chg_raw):+.2f}")
print(f"  Daily diff std: {np.nanstd(np.diff(chg_raw[~np.isnan(chg_raw)])):.3f}")
print(f"  Final: {chg_raw[-1]:+.2f}, DEXA actual: +2.80")

# ============================================================
# STEP 3: Write spreadsheet
# ============================================================
# !! IMPORTANT — DO NOT HARDCODE ALPHA !!
# The user may have customized alpha in BC2 of the oxiline tab.
# We MUST read the existing value and preserve it. This bug has been
# reintroduced twice — see _read_alpha_from_sheet() below.
ALPHA_DEFAULT = 2 / (20 + 1)  # span=20, fallback ONLY if BC2 is empty
HEIGHT_M = 1.778
FIRST_DATA_ROW = 4
MAX_ROW = 388
metrics = list(PCOLS.keys())

RAW = {
    'Weight': 'D', 'Fat_pct': 'H', 'F_mas': 'I', 'Lb_mas': 'J', 'V_fat': 'N',
    'R_a_f': 'T', 'L_a_f': 'U', 'T_f': 'V', 'R_l_f': 'W', 'L_l_f': 'X',
    'R_a_m': 'Y', 'L_a_m': 'Z', 'T_m': 'AA', 'R_l_m': 'AB', 'L_l_m': 'AC',
}

wb = load_workbook(DST)
ws = wb['oxiline']


def _read_alpha_from_sheet(ws, default):
    """Read alpha from BC2 (col 55, row 2). If the user has set a custom
    value, preserve it. Only fall back to `default` if the cell is empty.
    !! This function exists because hardcoding alpha has caused a
    regression TWICE. Do not bypass it. !!"""
    val = ws.cell(row=2, column=55).value
    if isinstance(val, (int, float)) and 0 < val < 1:
        print(f"  Preserving existing alpha from BC2: {val}")
        return val
    print(f"  No existing alpha in BC2 — using default: {round(default, 4)}")
    return default


ALPHA = _read_alpha_from_sheet(ws, ALPHA_DEFAULT)

# Styles
HDR_FILL = PatternFill('solid', fgColor='2E75B6')
EMA_FILL = PatternFill('solid', fgColor='E2EFDA')
DX_FILL = PatternFill('solid', fgColor='D6E4F0')
COR_FILL = PatternFill('solid', fgColor='F2DCDB')
CHG_FILL = PatternFill('solid', fgColor='FCE4D6')
PARAM_FILL = PatternFill('solid', fgColor='FFF2CC')
BASELINE_FILL = PatternFill('solid', fgColor='DDEBF7')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=9)
PARAM_FONT = Font(name='Arial', size=7, color='808080', italic=True)
BASELINE_FONT = Font(name='Arial', size=8, bold=True, color='1F4E79')
DATA_FONT = Font(name='Arial', size=9)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
)
center = Alignment(horizontal='center')

# Global params in row 1 of dedicated columns before EMA section
# Alpha in col 55 (BC) row 1, mean_wt in col 55 (BC) row 2
# EMA columns will start at col 56 (BD) instead
PARAMS_COL = 55  # BC
c = ws.cell(row=1, column=PARAMS_COL, value='alpha')
c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center
# Write ALPHA to BC2 — this value came from _read_alpha_from_sheet(),
# which preserves whatever the user had. Do NOT replace ALPHA with a
# hardcoded constant here.
c = ws.cell(row=2, column=PARAMS_COL, value=round(ALPHA, 4))
c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center; c.number_format = '0.0000'
c = ws.cell(row=3, column=PARAMS_COL, value='mean_wt')
c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center
# mean_wt as AVERAGE formula referencing ema_Weight range so it auto-updates
# We'll fill this after creating EMA columns
ALPHA_REF = f'${get_column_letter(PARAMS_COL)}$2'
MEAN_WT_REF = None  # will be set after EMA section

# ============================================================
# SECTION 1: EMA of raw BIA (cols 55-69)
# All EMA columns reference the single alpha in BB2
# ============================================================
EMA_START = 56  # BD, since BC is params
ema_letters = {}
for i, m in enumerate(metrics):
    col = EMA_START + i
    ema_letters[m] = get_column_letter(col)
    ws.column_dimensions[get_column_letter(col)].width = 9

    c = ws.cell(row=3, column=col, value=f'ema_{m}')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = thin_border

    raw_letter = RAW[m]

    for row in range(FIRST_DATA_ROW, MAX_ROW + 1):
        raw_ref = f'{raw_letter}{row}'
        if row == FIRST_DATA_ROW:
            formula = f'=IF({raw_ref}="","",{raw_ref})'
        else:
            prev = f'{ema_letters[m]}{row-1}'
            formula = f'=IF({raw_ref}="","",IF({prev}="",{raw_ref},{ALPHA_REF}*{raw_ref}+(1-{ALPHA_REF})*{prev}))'
        c = ws.cell(row=row, column=col, value=formula)
        c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = EMA_FILL
        c.number_format = '0.0%' if m == 'Fat_pct' else '0.0'

# Now set mean_wt as AVERAGE of ema_Weight column so it auto-updates with alpha
ema_wt_letter = ema_letters['Weight']
mean_wt_formula = f'=AVERAGE({ema_wt_letter}{FIRST_DATA_ROW}:{ema_wt_letter}{MAX_ROW})'
c = ws.cell(row=4, column=PARAMS_COL, value=mean_wt_formula)
c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center; c.number_format = '0.00'
MEAN_WT_REF = f'${get_column_letter(PARAMS_COL)}$4'

print(f"\nSection 1 (EMA): cols {EMA_START}-{EMA_START+len(metrics)-1}")
print(f"  Alpha: {get_column_letter(PARAMS_COL)}2, Mean_wt: {get_column_letter(PARAMS_COL)}4 (formula)")

# ============================================================
# SECTION 2: DEXA-calibrated (ENSEMBLE-BLENDED VALUES)
# Writes numpy-computed ensemble blend (w_oxi * oxi_dx + w_fi * fi_dx)
# as static values. Params rows show ensemble weight for reference.
# ============================================================
DX_START = EMA_START + len(metrics)
dx_letters = {}

# Map valid dataframe index -> spreadsheet row
# valid was created from df_raw.iloc[3:] with reset_index, so valid index 0 = sheet row 4
_valid_dates = valid[2].values  # for checking which rows have data
_n_valid = len(valid)

ENS_FILL = PatternFill('solid', fgColor='BDD7EE')  # distinct blue for ensemble

for i, m in enumerate(metrics):
    col = DX_START + i
    dx_letters[m] = get_column_letter(col)
    ws.column_dimensions[get_column_letter(col)].width = 9

    # Row 1: ensemble weight (w_oxi)
    w_oxi = ENSEMBLE_WEIGHTS[m]
    c = ws.cell(row=1, column=col, value=f'w_oxi={w_oxi:.0%}')
    c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center
    # Row 2: ensemble label
    c = ws.cell(row=2, column=col, value='ensemble')
    c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center
    # Header
    c = ws.cell(row=3, column=col, value=f'dx_{m}')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = thin_border

    # Write ensemble-blended values from dx_uncorrected numpy array
    for row in range(FIRST_DATA_ROW, MAX_ROW + 1):
        vi = row - FIRST_DATA_ROW  # valid index
        if vi < _n_valid:
            val = dx_uncorrected[m][vi]
            if not np.isnan(val):
                c = ws.cell(row=row, column=col, value=round(float(val), 6))
                c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = ENS_FILL
                c.number_format = '0.0%' if m == 'Fat_pct' else '0.0'
            else:
                c = ws.cell(row=row, column=col, value='')
                c.border = thin_border; c.fill = ENS_FILL
        else:
            # Beyond data range - write EMA-based formula as fallback for future data
            ema_ref = f'{ema_letters[m]}{row}'
            a_val, b_val = cal_oxi[m]
            a_ref_val = a_val
            b_ref_val = b_val
            formula = f'=IF({ema_ref}="","",{a_ref_val}*{ema_ref}+{b_ref_val})'
            c = ws.cell(row=row, column=col, value=formula)
            c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = DX_FILL
            c.number_format = '0.0%' if m == 'Fat_pct' else '0.0'

print(f"Section 2 (DEXA ensemble): cols {DX_START}-{DX_START+len(metrics)-1}")

# ============================================================
# SECTION 3: Weight-decorrelated DEXA (cols 85-99)
# cdx = dx - wt_slope * (ema_weight - mean_weight)
# ============================================================
CDX_START = DX_START + len(metrics)
cdx_letters = {}
for i, m in enumerate(metrics):
    col = CDX_START + i
    cdx_letters[m] = get_column_letter(col)
    ws.column_dimensions[get_column_letter(col)].width = 10

    # Row 1: wt_slope label
    c = ws.cell(row=1, column=col, value='wt_slope')
    c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center
    # Row 2: wt_slope value
    c = ws.cell(row=2, column=col, value=weight_slopes_dx[m])
    c.font = PARAM_FONT; c.fill = PARAM_FILL; c.alignment = center; c.number_format = '0.000000'
    # Header
    c = ws.cell(row=3, column=col, value=f'cdx_{m}')
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = thin_border

    wt_slope_ref = f'${cdx_letters[m]}$2'
    ema_wt_letter = ema_letters['Weight']

    for row in range(FIRST_DATA_ROW, MAX_ROW + 1):
        dx_ref = f'{dx_letters[m]}{row}'

        if m == 'Weight':
            formula = f'=IF({dx_ref}="","",{dx_ref})'
        else:
            ema_wt_ref = f'{ema_wt_letter}{row}'
            formula = f'=IF({dx_ref}="","",{dx_ref}-{wt_slope_ref}*({ema_wt_ref}-{MEAN_WT_REF}))'

        c = ws.cell(row=row, column=col, value=formula)
        c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = COR_FILL
        c.number_format = '0.0%' if m == 'Fat_pct' else '0.0'

print(f"Section 3 (Weight-decorrelated): cols {CDX_START}-{CDX_START+len(metrics)-1}")

# ============================================================
# SECTION 4: Derived + Change
# ============================================================
RESULT_START = CDX_START + len(metrics)

DEC_BASELINES = {
    'chg_Muscle': 62.3, 'chg_ArmMuscle': 16.6, 'chg_LegMuscle': 45.7,
    'chg_ArmMass': 2.2+2.5+8.7+7.9, 'chg_LegMass': 6.7+6.7+23.3+22.4,
    'chg_TrunkMuscle': 66.0, 'chg_TrunkMass': 25.5+66.0,
    'chg_BodyFat': 45.9, 'chg_BF_pct': 0.242, 'chg_ViscFat': 2.44,
    'chg_ALMI': round(62.3/2.20462/(HEIGHT_M**2), 4),
}

derived_defs = [
    ('dx_Muscle',    lambda r: f'{dx_letters["R_a_m"]}{r}+{dx_letters["L_a_m"]}{r}+{dx_letters["R_l_m"]}{r}+{dx_letters["L_l_m"]}{r}', '0.0'),
    ('dx_ArmMuscle', lambda r: f'{dx_letters["R_a_m"]}{r}+{dx_letters["L_a_m"]}{r}', '0.0'),
    ('dx_LegMuscle', lambda r: f'{dx_letters["R_l_m"]}{r}+{dx_letters["L_l_m"]}{r}', '0.0'),
    ('dx_ArmMass',   lambda r: f'{dx_letters["R_a_f"]}{r}+{dx_letters["L_a_f"]}{r}+{dx_letters["R_a_m"]}{r}+{dx_letters["L_a_m"]}{r}', '0.0'),
    ('dx_LegMass',   lambda r: f'{dx_letters["R_l_f"]}{r}+{dx_letters["L_l_f"]}{r}+{dx_letters["R_l_m"]}{r}+{dx_letters["L_l_m"]}{r}', '0.0'),
    ('dx_TrunkMass', lambda r: f'{dx_letters["T_f"]}{r}+{dx_letters["T_m"]}{r}', '0.0'),
    ('dx_ALMI',      None, '0.00'),
    ('dx_ALMass',    'DEFERRED_ALMASS', '0.0'),
]

chg_defs = [
    ('chg_Muscle',      'dx_Muscle',      '+0.0;-0.0;0.0'),
    ('chg_ArmMuscle',   'dx_ArmMuscle',   '+0.0;-0.0;0.0'),
    ('chg_LegMuscle',   'dx_LegMuscle',   '+0.0;-0.0;0.0'),
    ('chg_ArmMass',     'dx_ArmMass',     '+0.0;-0.0;0.0'),
    ('chg_LegMass',     'dx_LegMass',     '+0.0;-0.0;0.0'),
    ('chg_TrunkMuscle', 'dx_T_m',         '+0.0;-0.0;0.0'),
    ('chg_TrunkMass',   'dx_TrunkMass',   '+0.0;-0.0;0.0'),
    ('chg_BodyFat',     'dx_F_mas',       '+0.0;-0.0;0.0'),
    ('chg_BF_pct',      'dx_Fat_pct',     '+0.0%;-0.0%;0.0%'),
    ('chg_ViscFat',     'dx_V_fat',       '+0.00;-0.00;0.00'),
    ('chg_ALMI',        'dx_ALMI',        '+0.00;-0.00;0.00'),
]

result_letters = {}
all_cols = derived_defs + [(c[0], None, c[2]) for c in chg_defs]

for i, item in enumerate(all_cols):
    col = RESULT_START + i
    name = item[0]
    result_letters[name] = get_column_letter(col)
    ws.column_dimensions[get_column_letter(col)].width = 13
    c = ws.cell(row=3, column=col, value=name)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = thin_border

for i, (name, src, nfmt) in enumerate(chg_defs):
    col = RESULT_START + len(derived_defs) + i
    baseline = DEC_BASELINES[name]
    c = ws.cell(row=1, column=col, value='Dec6 DEXA')
    c.font = PARAM_FONT; c.fill = BASELINE_FILL; c.alignment = center
    c = ws.cell(row=2, column=col, value=baseline)
    c.font = BASELINE_FONT; c.fill = BASELINE_FILL; c.alignment = center
    c.number_format = '0.0%' if 'pct' in name else '0.00'

# EMA-sum limb components for direct ALMass calibration
_al_ema_components = ['R_a_f', 'L_a_f', 'R_l_f', 'L_l_f', 'R_a_m', 'L_a_m', 'R_l_m', 'L_l_m']
DEXA_ALMASS = DEC_BASELINES['chg_ArmMass'] + DEC_BASELINES['chg_LegMass']  # 80.4

for i in range(len(derived_defs)):
    col = RESULT_START + i
    name = derived_defs[i][0]
    c = ws.cell(row=1, column=col, value='derived')
    c.font = PARAM_FONT; c.fill = DX_FILL; c.alignment = center

    if name == 'dx_ALMass':
        # Row 2: offset = DEXA_ALMass - EMA_sum at first data row (auto-adjusts with alpha)
        ema_sum_row4 = '+'.join(f'{ema_letters[m]}{FIRST_DATA_ROW}' for m in _al_ema_components)
        c2 = ws.cell(row=2, column=col, value=f'={DEXA_ALMASS}-({ema_sum_row4})')
        c2.font = PARAM_FONT; c2.fill = PARAM_FILL; c2.alignment = center; c2.number_format = '0.000'

check_col = dx_letters['Weight']

for row in range(FIRST_DATA_ROW, MAX_ROW + 1):
    cr = f'{check_col}{row}'

    for j, (name, formula_fn, nfmt) in enumerate(derived_defs):
        col = RESULT_START + j
        if name == 'dx_ALMI':
            muscle_ref = f'{result_letters["dx_Muscle"]}{row}'
            formula = f'=IF({cr}="","",{muscle_ref}/2.20462/{HEIGHT_M}^2)'
        elif name == 'dx_ALMass':
            # ALMass = EMA sum of 8 limb components + DEXA offset (avoids amplified leg fat calibration)
            ema_sum = '+'.join(f'{ema_letters[m]}{row}' for m in _al_ema_components)
            offset_ref = f'${result_letters["dx_ALMass"]}$2'
            formula = f'=IF({cr}="","",{ema_sum}+{offset_ref})'
        else:
            expr = formula_fn(row)
            formula = f'=IF({cr}="","",{expr})'
        c = ws.cell(row=row, column=col, value=formula)
        c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = DX_FILL; c.number_format = nfmt

    for j, (name, src, nfmt) in enumerate(chg_defs):
        col = RESULT_START + len(derived_defs) + j
        baseline_ref = f'${get_column_letter(col)}$2'

        if src == 'dx_T_m':
            src_ref = f'{dx_letters["T_m"]}{row}'
        elif src == 'dx_F_mas':
            src_ref = f'{dx_letters["F_mas"]}{row}'
        elif src == 'dx_Fat_pct':
            src_ref = f'{dx_letters["Fat_pct"]}{row}'
        elif src == 'dx_V_fat':
            src_ref = f'{dx_letters["V_fat"]}{row}'
        elif src == 'dx_ALMI':
            src_ref = f'{result_letters["dx_ALMI"]}{row}'
        elif src in result_letters:
            src_ref = f'{result_letters[src]}{row}'
        else:
            continue

        formula = f'=IF({cr}="","",{src_ref}-{baseline_ref})'
        c = ws.cell(row=row, column=col, value=formula)
        c.font = DATA_FONT; c.alignment = center; c.border = thin_border; c.fill = CHG_FILL; c.number_format = nfmt

total_end = RESULT_START + len(all_cols) - 1
print(f"Section 4 (Derived+Change): cols {RESULT_START}-{total_end}")

# Save to a TEMP file to extract the NEW formulas from openpyxl output.
# We UPDATE formulas IN-PLACE within existing cells, preserving cell structure,
# styles, and shared-formula XML structure. Only ADD truly new cells for
# rows/columns that don't exist in the original.
import tempfile as _tf
import zipfile
from lxml import etree as _ET

_oxi_temp = os.path.join(_tf.mkdtemp(), 'oxi_temp.xlsx')
wb.save(_oxi_temp)
wb.close()
print(f"\nSaved DEXA columns to temp. Total new columns: {total_end - EMA_START + 1}")

# --- Helpers ---
def _col_num(col_str):
    n = 0
    for ch in col_str:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def _cell_col_letter(cell_ref):
    return ''.join(c for c in cell_ref if c.isalpha())

import re as _re

# --- Extract DST into temp dir for surgery ---
_inject_dir = _tf.mkdtemp()
with zipfile.ZipFile(DST, 'r') as z:
    z.extractall(_inject_dir)

# Find oxiline sheet path in DST
_wb_xml = open(os.path.join(_inject_dir, 'xl', 'workbook.xml'), 'r').read()
_rels_xml = open(os.path.join(_inject_dir, 'xl', '_rels', 'workbook.xml.rels'), 'r').read()
_oxi_rid_match = _re.search(r'name="oxiline"[^>]*r:id="(rId\d+)"', _wb_xml) or \
    _re.search(r'name="oxiline"[^>]*id="(rId\d+)"', _wb_xml, _re.IGNORECASE)
_oxi_rid = _oxi_rid_match.group(1)
_oxi_target_match = _re.search(f'Id="{_oxi_rid}"[^>]*Target="([^"]+)"', _rels_xml) or \
    _re.search(f'Target="([^"]+)"[^>]*Id="{_oxi_rid}"', _rels_xml)
_oxi_sheet_path = _oxi_target_match.group(1).lstrip('/')
if _oxi_sheet_path.startswith('xl/'): _oxi_sheet_path = _oxi_sheet_path[3:]
print(f"Oxiline sheet file: {_oxi_sheet_path} (rId={_oxi_rid})")

# --- Find oxiline in temp file and parse with lxml ---
with zipfile.ZipFile(_oxi_temp, 'r') as z_temp:
    _temp_wb_xml = z_temp.read('xl/workbook.xml').decode('utf-8')
    _temp_rels = z_temp.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    _temp_rid_match = _re.search(r'name="oxiline"[^>]*id="(rId\d+)"', _temp_wb_xml, _re.IGNORECASE) or \
        _re.search(r'name="oxiline"[^>]*r:id="(rId\d+)"', _temp_wb_xml)
    _temp_rid = _temp_rid_match.group(1)
    _temp_target_match = _re.search(f'Id="{_temp_rid}"[^>]*Target="([^"]+)"', _temp_rels) or \
        _re.search(f'Target="([^"]+)"[^>]*Id="{_temp_rid}"', _temp_rels)
    _temp_oxi_path = _temp_target_match.group(1).lstrip('/')
    if not _temp_oxi_path.startswith('xl/'): _temp_oxi_path = 'xl/' + _temp_oxi_path
    _temp_oxi_bytes = z_temp.read(_temp_oxi_path)

# Parse temp oxiline with lxml to get new formulas by cell ref
_temp_tree = _ET.fromstring(_temp_oxi_bytes)
_SS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NEW_COL_MIN = PARAMS_COL  # 55 = BC

# Build dict: cell_ref -> formula_text from openpyxl temp
_new_formulas = {}  # e.g. 'BD4' -> 'IF(D4="","",D4)'
_new_values = {}    # for cells with values but no formula (e.g. calibration params)
for temp_row in _temp_tree.iter(f'{{{_SS}}}row'):
    for c_elem in temp_row.findall(f'{{{_SS}}}c'):
        ref = c_elem.get('r', '')
        col_letter = _cell_col_letter(ref)
        if not col_letter or _col_num(col_letter) < NEW_COL_MIN:
            continue
        f_elem = c_elem.find(f'{{{_SS}}}f')
        v_elem = c_elem.find(f'{{{_SS}}}v')
        if f_elem is not None and f_elem.text:
            _new_formulas[ref] = f_elem.text
        elif v_elem is not None and v_elem.text:
            _new_values[ref] = (v_elem.text, c_elem.get('t'))

print(f"Extracted {len(_new_formulas)} formulas + {len(_new_values)} values (cols >= {NEW_COL_MIN})")

# --- Parse original oxiline sheet with lxml ---
_orig_path = os.path.join(_inject_dir, 'xl', _oxi_sheet_path)
with open(_orig_path, 'rb') as f:
    _orig_tree = _ET.parse(f)
_orig_root = _orig_tree.getroot()
_ns = _orig_root.nsmap.get(None, _SS)

# --- Pre-scan: build shared formula group map for cols >= 55 ---
# We need this BEFORE modifying anything so we can compute explicit formulas
# for slave cells that aren't in our pipeline output.
_sheet_data = _orig_root.find(f'{{{_ns}}}sheetData')
_orig_rows = {int(r.get('r')): r for r in _sheet_data.findall(f'{{{_ns}}}row')}

# Map: si_number -> {'master_ref': 'BS4', 'master_row': 4, 'formula': 'IF(...)', 'col': 'BS'}
_shared_groups = {}
for row_num, row_elem in _orig_rows.items():
    for c_elem in row_elem.findall(f'{{{_ns}}}c'):
        ref = c_elem.get('r', '')
        col_letter = _cell_col_letter(ref)
        if not col_letter or _col_num(col_letter) < NEW_COL_MIN:
            continue
        f_elem = c_elem.find(f'{{{_ns}}}f')
        if f_elem is not None and f_elem.get('t') == 'shared':
            si = f_elem.get('si')
            if si and f_elem.text:  # This is a master (has formula text)
                cell_row = int(''.join(c for c in ref if c.isdigit()))
                _shared_groups[si] = {
                    'master_ref': ref,
                    'master_row': cell_row,
                    'formula': f_elem.text,
                    'col': col_letter
                }

print(f"Found {len(_shared_groups)} shared formula groups in cols >= {NEW_COL_MIN}")

# --- Helper: adjust formula row references for slave offset ---
def _adjust_formula_row(formula, row_delta):
    """Shift relative row references in an Excel formula by row_delta.
    $A$1 -> no change (both absolute)
    $A1  -> $A(1+delta)  (row is relative)
    A$1  -> A$1          (row is absolute)
    A1   -> A(1+delta)   (row is relative)
    """
    if row_delta == 0:
        return formula
    # Match cell references: optional $, col letters, optional $, row digits
    # Negative lookbehind for alphanumeric to avoid matching inside function names
    def _adjust_match(m):
        prefix = m.group(1)       # everything before the col
        col_part = m.group(2)     # $?[A-Z]+
        dollar = m.group(3)       # $ or empty (before row number)
        row_str = m.group(4)      # digits
        if dollar == '$':
            # Absolute row reference — don't adjust
            return prefix + col_part + dollar + row_str
        new_row = int(row_str) + row_delta
        if new_row < 1:
            new_row = 1
        return prefix + col_part + str(new_row)
    # Pattern: cell references like A1, $A1, A$1, $A$1, AA123, etc.
    # Use negative lookbehind for word chars to avoid matching inside names
    result = _re.sub(
        r'((?:^|[^A-Za-z0-9_]))(\$?[A-Z]{1,3})(\$?)(\d+)',
        _adjust_match,
        formula
    )
    return result

# --- Update existing cells in-place, preserving structure ---
# Strategy (confirmed by binary search testing):
# 1. Cells in _new_formulas: set explicit formula, strip shared attrs, remove <v>
# 2. Cells in _new_values: update value, but SKIP t="s" (shared string) cells
# 3. Unmatched cells with shared formulas: compute explicit formula from master
#    (master keeps its text, slaves get row-adjusted copy of master formula)
# 4. Unmatched cells without shared formulas: leave completely untouched

_updated_f = 0  # formulas updated in existing cells
_updated_v = 0  # values updated in existing cells
_added = 0      # new cells added
_shared_converted = 0  # shared formulas converted to regular
_slaves_expanded = 0   # slave formulas expanded from master
_skipped_ss = 0        # shared string cells skipped
_processed_refs = set()

# Phase 1: Update formulas/values in existing cells (keep cell r, s, etc. intact)
for row_num, row_elem in _orig_rows.items():
    for c_elem in row_elem.findall(f'{{{_ns}}}c'):
        ref = c_elem.get('r', '')
        col_letter = _cell_col_letter(ref)
        if not col_letter or _col_num(col_letter) < NEW_COL_MIN:
            continue

        if ref in _new_formulas:
            _processed_refs.add(ref)
            new_formula = _new_formulas[ref]

            # Get or create <f> element
            f_elem = c_elem.find(f'{{{_ns}}}f')
            if f_elem is not None:
                # STRIP all shared formula attributes — convert to regular formula
                for attr in ['t', 'ref', 'si']:
                    if attr in f_elem.attrib:
                        del f_elem.attrib[attr]
                        _shared_converted += 1
                # Set the explicit formula text for THIS cell
                f_elem.text = new_formula
            else:
                # Cell had no formula — add one
                f_elem = _ET.SubElement(c_elem, f'{{{_ns}}}f')
                f_elem.text = new_formula

            # Remove cached value entirely (Excel will recalculate)
            v_elem = c_elem.find(f'{{{_ns}}}v')
            if v_elem is not None:
                c_elem.remove(v_elem)
            _updated_f += 1

        elif ref in _new_values:
            val_text, val_type = _new_values[ref]
            if val_type == 's':
                # SKIP shared-string cells — the shared string table indices
                # differ between temp and original workbooks.
                _processed_refs.add(ref)
                _skipped_ss += 1
            else:
                # UPDATE numeric/boolean value cells (e.g. decorrelation slopes,
                # calibration params, ensemble-blended DX values). These are
                # computed fresh each run and must be written even if the cell
                # already exists in the original.
                _processed_refs.add(ref)

                # CRITICAL: Remove any existing formula element so Excel treats
                # this as a value cell, not a formula cell. This is essential for
                # ensemble DX columns which replace formulas with static values.
                f_elem = c_elem.find(f'{{{_ns}}}f')
                if f_elem is not None:
                    # Strip shared formula attrs before removing
                    for attr in ['t', 'ref', 'si']:
                        if attr in f_elem.attrib:
                            _shared_converted += 1
                    c_elem.remove(f_elem)

                v_elem = c_elem.find(f'{{{_ns}}}v')
                if v_elem is not None:
                    v_elem.text = val_text
                else:
                    v_elem = _ET.SubElement(c_elem, f'{{{_ns}}}v')
                    v_elem.text = val_text
                # Update type attribute if needed
                if val_type:
                    c_elem.set('t', val_type)
                elif 't' in c_elem.attrib:
                    del c_elem.attrib['t']
                _updated_v += 1

        else:
            # Cell exists in original but NOT in our pipeline output.
            # If it has a shared formula, convert to regular with explicit text.
            # Leave non-shared-formula cells completely untouched.
            f_elem = c_elem.find(f'{{{_ns}}}f')
            if f_elem is not None and f_elem.get('t') == 'shared':
                si = f_elem.get('si')
                if f_elem.text:
                    # MASTER cell — has formula text, just strip shared attrs
                    for attr in ['t', 'ref', 'si']:
                        if attr in f_elem.attrib:
                            del f_elem.attrib[attr]
                    _shared_converted += 1
                elif si and si in _shared_groups:
                    # SLAVE cell — compute explicit formula from master
                    group = _shared_groups[si]
                    cell_row = int(''.join(c for c in ref if c.isdigit()))
                    row_delta = cell_row - group['master_row']
                    expanded = _adjust_formula_row(group['formula'], row_delta)
                    # Strip shared attrs and set explicit formula
                    for attr in ['t', 'ref', 'si']:
                        if attr in f_elem.attrib:
                            del f_elem.attrib[attr]
                    f_elem.text = expanded
                    # Remove cached value (Excel will recalculate)
                    v_elem = c_elem.find(f'{{{_ns}}}v')
                    if v_elem is not None:
                        c_elem.remove(v_elem)
                    _slaves_expanded += 1
                    _shared_converted += 1
                else:
                    # Orphaned slave with unknown group — strip shared attrs,
                    # remove formula, leave cell as styled-only
                    for attr in ['t', 'ref', 'si']:
                        if attr in f_elem.attrib:
                            del f_elem.attrib[attr]
                    c_elem.remove(f_elem)
                    _shared_converted += 1

# Phase 2: Add cells that don't exist in original
_unprocessed_f = set(_new_formulas.keys()) - _processed_refs
_unprocessed_v = set(_new_values.keys()) - _processed_refs
_unprocessed = _unprocessed_f | _unprocessed_v

if _unprocessed:
    # Group by row
    _add_by_row = {}
    for ref in _unprocessed:
        row_num = int(''.join(c for c in ref if c.isdigit()))
        _add_by_row.setdefault(row_num, []).append(ref)

    for row_num, refs in sorted(_add_by_row.items()):
        if row_num in _orig_rows:
            row_elem = _orig_rows[row_num]
        else:
            row_elem = _ET.SubElement(_sheet_data, f'{{{_ns}}}row')
            row_elem.set('r', str(row_num))

        for ref in sorted(refs, key=lambda r: _col_num(_cell_col_letter(r))):
            new_c = _ET.SubElement(row_elem, f'{{{_ns}}}c')
            new_c.set('r', ref)
            if ref in _new_formulas:
                f_el = _ET.SubElement(new_c, f'{{{_ns}}}f')
                f_el.text = _new_formulas[ref]
            elif ref in _new_values:
                val_text, val_type = _new_values[ref]
                if val_type:
                    new_c.set('t', val_type)
                v_el = _ET.SubElement(new_c, f'{{{_ns}}}v')
                v_el.text = val_text
            _added += 1

        row_elem.set('spans', f'1:{total_end}')

print(f"Updated {_updated_f} formulas + {_updated_v} values in-place, added {_added} new cells")
print(f"Converted {_shared_converted} shared attrs, expanded {_slaves_expanded} slaves from masters")
print(f"Skipped {_skipped_ss} shared-string cells")

# --- Replace AQ (col 43) with EMA of AK (col 37) ---
# EMA formula: row 4 seeds with AK4, rows 5+ use alpha*AK + (1-alpha)*prev_AQ
# Alpha comes from $BC$2 (same as other EMA columns)
_aq_updated = 0
_AQ_COL = 43
_AQ_LETTER = 'AQ'
_AK_LETTER = 'AK'
_DATA_START = 4   # first row with AK data
_DATA_END = max(_orig_rows.keys())  # last row in oxiline XML

for row_num, row_elem in _orig_rows.items():
    if row_num < _DATA_START or row_num > _DATA_END:
        continue
    for c_elem in row_elem.findall(f'{{{_ns}}}c'):
        ref = c_elem.get('r', '')
        if ref != f'{_AQ_LETTER}{row_num}':
            continue
        # Build EMA formula
        if row_num == _DATA_START:
            ema_formula = f'{_AK_LETTER}{row_num}'
        else:
            ema_formula = f'$BC$2*{_AK_LETTER}{row_num}+(1-$BC$2)*{_AQ_LETTER}{row_num - 1}'
        # Set formula, strip shared attrs
        f_elem = c_elem.find(f'{{{_ns}}}f')
        if f_elem is not None:
            for attr in ['t', 'ref', 'si']:
                if attr in f_elem.attrib:
                    del f_elem.attrib[attr]
            f_elem.text = ema_formula
        else:
            f_elem = _ET.SubElement(c_elem, f'{{{_ns}}}f')
            f_elem.text = ema_formula
        # Remove cached value
        v_elem = c_elem.find(f'{{{_ns}}}v')
        if v_elem is not None:
            c_elem.remove(v_elem)
        # Remove type attr so Excel treats as number
        if 't' in c_elem.attrib:
            del c_elem.attrib['t']
        _aq_updated += 1
        break

print(f"Replaced AQ with EMA of AK: {_aq_updated} cells updated (rows {_DATA_START}-{_DATA_END})")

# Update dimension ref
_max_col_letter = get_column_letter(total_end)
_dim = _orig_root.find(f'{{{_ns}}}dimension')
if _dim is not None:
    _dim.set('ref', f'A1:{_max_col_letter}588')

# --- Write modified oxiline XML ---
_out_bytes = _ET.tostring(_orig_root, xml_declaration=True, encoding='UTF-8', standalone=True)
with open(_orig_path, 'wb') as f:
    f.write(_out_bytes)
print(f"Wrote modified oxiline sheet ({len(_out_bytes):,} bytes)")

# --- Add DEXA weekly average columns to Week tab XML ---
# Find Week sheet file in the extracted ZIP
_wk_rid_match = re.search(r'name="Week"[^>]*r:id="(rId\d+)"', _wb_xml) or \
    re.search(r'name="Week"[^>]*id="(rId\d+)"', _wb_xml, re.IGNORECASE)
if _wk_rid_match:
    _wk_rid = _wk_rid_match.group(1)
    _wk_tgt_match = re.search(f'Id="{_wk_rid}"[^>]*Target="([^"]+)"', _rels_xml) or \
        re.search(f'Target="([^"]+)"[^>]*Id="{_wk_rid}"', _rels_xml)
    _wk_sheet_path = _wk_tgt_match.group(1).lstrip('/')
    if _wk_sheet_path.startswith('xl/'): _wk_sheet_path = _wk_sheet_path[3:]
    _wk_path = os.path.join(_inject_dir, 'xl', _wk_sheet_path)

    _wk_tree = _ET.parse(_wk_path)
    _wk_root = _wk_tree.getroot()
    _wk_ns = _wk_root.nsmap.get(None, _SS)
    _wk_sheet_data = _wk_root.find(f'{{{_wk_ns}}}sheetData')
    _wk_rows = {int(r.get('r')): r for r in _wk_sheet_data.findall(f'{{{_wk_ns}}}row')}

    # Define columns to add: (header, oxiline_col_number)
    _wk_dexa_cols = []
    _dx_metric_names = ['dx_Weight', 'dx_Fat_pct', 'dx_F_mas', 'dx_Lb_mas', 'dx_V_fat',
                        'dx_R_a_f', 'dx_L_a_f', 'dx_T_f', 'dx_R_l_f', 'dx_L_l_f',
                        'dx_R_a_m', 'dx_L_a_m', 'dx_T_m', 'dx_R_l_m', 'dx_L_l_m']
    for i, name in enumerate(_dx_metric_names):
        _wk_dexa_cols.append((name, DX_START + i))
    _derived_names = ['dx_Muscle', 'dx_ArmMuscle', 'dx_LegMuscle', 'dx_ArmMass',
                      'dx_LegMass', 'dx_TrunkMass', 'dx_ALMI', 'dx_ALMass']
    for i, name in enumerate(_derived_names):
        _wk_dexa_cols.append((name, RESULT_START + i))
    _chg_names = ['chg_Muscle', 'chg_ArmMuscle', 'chg_LegMuscle', 'chg_ArmMass',
                  'chg_LegMass', 'chg_TrunkMuscle', 'chg_TrunkMass', 'chg_BodyFat',
                  'chg_BF_pct', 'chg_ViscFat', 'chg_ALMI']
    for i, name in enumerate(_chg_names):
        _wk_dexa_cols.append((name, RESULT_START + len(derived_defs) + i))

    _WK_DX_START = 51  # col AY
    _WK_FIRST_DATA = 2
    _WK_LAST_DATA = 52
    _OXI_DATA_START = 4

    # Remove any existing cells in columns >= _WK_DX_START (from previous runs)
    _wk_removed = 0
    for _rn, _r_elem in _wk_rows.items():
        for c_elem in list(_r_elem.findall(f'{{{_wk_ns}}}c')):
            ref = c_elem.get('r', '')
            col_str = ''.join(ch for ch in ref if ch.isalpha())
            col_idx = 0
            for ch in col_str:
                col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
            if col_idx >= _WK_DX_START:
                _r_elem.remove(c_elem)
                _wk_removed += 1
    if _wk_removed:
        print(f"  Removed {_wk_removed} existing DEXA cells from Week tab (previous run)")

    # Remove any existing <col> entries that overlap with our new columns
    _wk_cols_existing = _wk_root.find(f'{{{_wk_ns}}}cols')
    if _wk_cols_existing is not None:
        for col_el in list(_wk_cols_existing.findall(f'{{{_wk_ns}}}col')):
            cmin = int(col_el.get('min', '0'))
            cmax = int(col_el.get('max', '0'))
            if cmax >= _WK_DX_START:
                _wk_cols_existing.remove(col_el)

    # Add header row (row 1)
    if 1 not in _wk_rows:
        _r1 = _ET.SubElement(_wk_sheet_data, f'{{{_wk_ns}}}row')
        _r1.set('r', '1')
        _wk_rows[1] = _r1
    for j, (header, _) in enumerate(_wk_dexa_cols):
        col_num = _WK_DX_START + j
        col_letter = get_column_letter(col_num)
        cell_ref = f'{col_letter}1'
        c = _ET.SubElement(_wk_rows[1], f'{{{_wk_ns}}}c')
        c.set('r', cell_ref)
        c.set('t', 'inlineStr')
        is_elem = _ET.SubElement(c, f'{{{_wk_ns}}}is')
        t_elem = _ET.SubElement(is_elem, f'{{{_wk_ns}}}t')
        t_elem.text = header

    # Add data rows with AVERAGE formulas
    _wk_cells_added = 0
    for row in range(_WK_FIRST_DATA, _WK_LAST_DATA + 1):
        if row not in _wk_rows:
            _r_elem = _ET.SubElement(_wk_sheet_data, f'{{{_wk_ns}}}row')
            _r_elem.set('r', str(row))
            _wk_rows[row] = _r_elem

        for j, (header, oxi_col) in enumerate(_wk_dexa_cols):
            col_num = _WK_DX_START + j
            col_letter = get_column_letter(col_num)
            oxi_letter = get_column_letter(oxi_col)
            cell_ref = f'{col_letter}{row}'

            oxi_start_row = _OXI_DATA_START + (row - _WK_FIRST_DATA) * 7
            oxi_end_row = oxi_start_row + 6
            rng = f'oxiline!${oxi_letter}${oxi_start_row}:${oxi_letter}${oxi_end_row}'
            formula = f'IF(COUNT({rng})=0,"",AVERAGE({rng}))'

            c = _ET.SubElement(_wk_rows[row], f'{{{_wk_ns}}}c')
            c.set('r', cell_ref)
            f_elem = _ET.SubElement(c, f'{{{_wk_ns}}}f')
            f_elem.text = formula
            _wk_cells_added += 1

    # --- Add "0.00" number format and style to styles.xml for Week DEXA cells ---
    _wk_styles_path = os.path.join(_inject_dir, 'xl', 'styles.xml')
    with open(_wk_styles_path, 'r', encoding='utf-8') as f:
        _wk_sty = f.read()
    # Find or create numFmtId for "0.00"
    _wk_nf_match = re.search(r'numFmtId="(\d+)"[^>]*formatCode="0\.00"', _wk_sty)
    if _wk_nf_match:
        _wk_nfid = _wk_nf_match.group(1)
    else:
        # Allocate a new numFmtId
        _wk_all_nfids = [int(x) for x in re.findall(r'numFmtId="(\d+)"', _wk_sty)]
        _wk_nfid = str(max(_wk_all_nfids, default=163) + 1)
        _nf_entry = f'<numFmt numFmtId="{_wk_nfid}" formatCode="0.00"/>'
        if '<numFmts' in _wk_sty:
            # Increment count and append entry
            _wk_sty = re.sub(r'(<numFmts\s+count=")(\d+)(")', lambda m: m.group(1) + str(int(m.group(2)) + 1) + m.group(3), _wk_sty)
            _wk_sty = _wk_sty.replace('</numFmts>', _nf_entry + '</numFmts>')
        else:
            _wk_sty = _wk_sty.replace('<fonts', f'<numFmts count="1">{_nf_entry}</numFmts><fonts')
    # Add a cellXf entry referencing this numFmtId
    _wk_xf_entry = f'<xf numFmtId="{_wk_nfid}" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
    _wk_xf_count_match = re.search(r'(<cellXfs\s+count=")(\d+)(")', _wk_sty)
    _wk_new_style_id = int(_wk_xf_count_match.group(2))  # 0-based index = old count
    _wk_sty = re.sub(r'(<cellXfs\s+count=")(\d+)(")', lambda m: m.group(1) + str(int(m.group(2)) + 1) + m.group(3), _wk_sty)
    _wk_sty = _wk_sty.replace('</cellXfs>', _wk_xf_entry + '</cellXfs>')
    with open(_wk_styles_path, 'w', encoding='utf-8') as f:
        f.write(_wk_sty)
    print(f"  Week DEXA style: numFmtId={_wk_nfid}, cellXf s=\"{_wk_new_style_id}\" (format 0.00)")

    # Apply style to all DEXA data cells (not headers)
    for row in range(_WK_FIRST_DATA, _WK_LAST_DATA + 1):
        for c_elem in _wk_rows[row].findall(f'{{{_wk_ns}}}c'):
            ref = c_elem.get('r', '')
            # Check if it's one of our new columns (AY onwards = col 51+)
            col_str = ''.join(ch for ch in ref if ch.isalpha())
            col_idx = 0
            for ch in col_str:
                col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
            if col_idx >= _WK_DX_START:
                c_elem.set('s', str(_wk_new_style_id))

    # --- Add narrow column widths for the new DEXA columns ---
    _wk_cols = _wk_root.find(f'{{{_wk_ns}}}cols')
    if _wk_cols is None:
        _wk_cols = _ET.SubElement(_wk_root, f'{{{_wk_ns}}}cols')
        # Insert <cols> before <sheetData>
        _wk_root.remove(_wk_cols)
        _wk_root.insert(list(_wk_root).index(_wk_sheet_data), _wk_cols)
    _wk_first_new = _WK_DX_START
    _wk_last_new = _WK_DX_START + len(_wk_dexa_cols) - 1
    _col_elem = _ET.SubElement(_wk_cols, f'{{{_wk_ns}}}col')
    _col_elem.set('min', str(_wk_first_new))
    _col_elem.set('max', str(_wk_last_new))
    _col_elem.set('width', '9')
    _col_elem.set('customWidth', '1')

    _wk_bytes = _ET.tostring(_wk_root, xml_declaration=True, encoding='UTF-8', standalone=True)
    with open(_wk_path, 'wb') as f:
        f.write(_wk_bytes)
    print(f"Added {len(_wk_dexa_cols)} DEXA columns to Week tab ({_wk_cells_added} formula cells)")
else:
    print("WARNING: Week tab not found in workbook")

# Repackage DST
with zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED) as z_new:
    for root, dirs, files in os.walk(_inject_dir):
        for fn in files:
            abs_path = os.path.join(root, fn)
            arc_name = os.path.relpath(abs_path, _inject_dir)
            z_new.write(abs_path, arc_name)
print("Updated oxiline formulas in-place (original cell structure preserved)")

# --- Clean calcChain.xml: remove entries for DX cells that are now values ---
# DX columns (71-85) rows 4 to (4+_n_valid-1) are now static values, not formulas.
# Leaving them in calcChain causes Excel repair errors.
_cc_path = os.path.join(_inject_dir, 'xl', 'calcChain.xml')
if os.path.exists(_cc_path):
    _cc_tree = _ET.parse(_cc_path)
    _cc_root = _cc_tree.getroot()
    _cc_ns = _cc_root.nsmap.get(None, _SS)

    # Build set of cell refs that are now value-only (DX cols, data rows)
    _dx_value_refs = set()
    for i, m in enumerate(metrics):
        col_num = DX_START + i
        cl = get_column_letter(col_num)
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + _n_valid):
            _dx_value_refs.add(f'{cl}{row}')

    # Remove matching <c> elements from calcChain
    _cc_removed = 0
    for c_elem in list(_cc_root.findall(f'{{{_cc_ns}}}c')):
        ref = c_elem.get('r', '')
        if ref in _dx_value_refs:
            _cc_root.remove(c_elem)
            _cc_removed += 1

    if _cc_removed > 0:
        _cc_bytes = _ET.tostring(_cc_root, xml_declaration=True, encoding='UTF-8', standalone=True)
        with open(_cc_path, 'wb') as f:
            f.write(_cc_bytes)
        print(f"Cleaned calcChain.xml: removed {_cc_removed} entries for value-only DX cells")

    # Repackage DST again with cleaned calcChain
    with zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED) as z_new:
        for root, dirs, files in os.walk(_inject_dir):
            for fn in files:
                abs_path = os.path.join(root, fn)
                arc_name = os.path.relpath(abs_path, _inject_dir)
                z_new.write(abs_path, arc_name)

# ============================================================
# STEP 5: Fix chart ranges + inject ExerciseSummary (XML surgery on saved file)
# ============================================================
# openpyxl saves the file correctly but chart XML still has old range refs.
# We operate directly on the ZIP/XML to update chart ranges and add a new sheet.
import zipfile
import tempfile
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta

# --- 5a: Compute data extents and build chart range replacement map ---
# Approach: scan chart XMLs from the saved file for all range references.
# For each sheet, the Jan 1 start row is known. The end row comes from either:
#   - oxiline: numpy data extent (since we may have added new data rows)
#   - All others: max end row found in existing chart references (authoritative)

oxi_last = valid.index[-1] + 4  # actual last Excel row with data (index 0 = row 4)
oxi_jan1 = 23
fi_jan1 = 21
wy_jan1 = 21
wk_jan1 = 5   # first full week in Jan is row 5 in Week sheet
mo_jan1 = 3   # Jan is row 3 in Month sheet

JAN1_ROWS = {'oxiline': oxi_jan1, 'Fitindex': fi_jan1, 'Wyze': wy_jan1, 'Week': wk_jan1, 'Month': mo_jan1}

# Scan chart refs from saved file (openpyxl may use <f> instead of <c:f>)
with zipfile.ZipFile(DST, 'r') as z:
    chart_files_list = [f for f in z.namelist() if 'charts/chart' in f and f.endswith('.xml')]
    all_range_refs = set()
    for cf in chart_files_list:
        content = z.read(cf).decode('utf-8')
        # Match both <c:f> and <f> tags (openpyxl may strip namespace)
        refs = re.findall(r'<(?:c:)?f>([^<]+)</(?:c:)?f>', content)
        for r in refs:
            if ':' in r:
                all_range_refs.add(r)

# Find max end row per sheet from chart references
sheet_max_end = {}
for ref in all_range_refs:
    m_ref = re.match(r"(.+)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", ref)
    if m_ref:
        sheet = m_ref.group(1)
        end_row = int(m_ref.group(5))
        sheet_max_end[sheet] = max(sheet_max_end.get(sheet, 0), end_row)

# Override end rows with actual data extent from numpy
# oxiline starts at row 4, Fitindex/Wyze start at row 2 (same data count)
n_data = oxi_last - 3  # number of data rows
sheet_max_end['oxiline'] = oxi_last
sheet_max_end['Fitindex'] = 1 + n_data  # header row 1 + n_data rows
sheet_max_end['Wyze'] = 1 + n_data

SHEET_CONFIG = {}
for sheet in JAN1_ROWS:
    if sheet in sheet_max_end:
        SHEET_CONFIG[sheet] = (JAN1_ROWS[sheet], sheet_max_end[sheet])

print(f"\nChart range targets:")
for s, (j, l) in SHEET_CONFIG.items():
    print(f"  {s}: row {j} to {l}")

# Build replacement map — only update END row, preserve original start row.
# Some charts (e.g. scatter plots) are designed to show broader historical data
# with start rows before Jan 1. Forcing all charts to start at Jan 1 truncates them.
range_replacement = {}
for ref in all_range_refs:
    m_ref = re.match(r"(.+)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", ref)
    if m_ref:
        sheet, col_s, row_s, col_e, row_e = m_ref.groups()
        if sheet in SHEET_CONFIG:
            _, new_e = SHEET_CONFIG[sheet]
            old_e = int(row_e)
            # Only EXPAND chart ranges, never shrink them.
            # Some charts (like chart6) have large pre-set ranges (e.g., row 1021)
            # designed to accommodate future data. Shrinking them breaks the chart.
            if new_e > old_e:
                new_ref = f"{sheet}!${col_s}${row_s}:${col_e}${new_e}"
                range_replacement[ref] = new_ref

print(f"Found {len(all_range_refs)} chart range refs, {len(range_replacement)} need updating")

# --- 5c: Build DEXA body comp lookup from numpy arrays (already in memory) ---
# Use raw DEXA-calibrated values (dx_uncorrected) to match change columns,
# which compare against raw DEXA baselines. Weight-decorrelated values (dx_corrected)
# are used for Section 3 charts only.
dexa_by_date = {}
muscle_arr = dx_uncorrected['R_a_m'] + dx_uncorrected['L_a_m'] + dx_uncorrected['R_l_m'] + dx_uncorrected['L_l_m']
almi_arr = muscle_arr / 2.20462 / (HEIGHT_M ** 2)

for i in range(len(valid)):
    dt = valid[2].iloc[i]
    if pd.isna(dt):
        continue
    dexa_by_date[dt] = {
        'weight': dx_uncorrected['Weight'][i],
        'fat_pct': dx_uncorrected['Fat_pct'][i],
        'lean_mass': dx_uncorrected['Lb_mas'][i],
        'muscle': muscle_arr[i],
        'almi': almi_arr[i],
    }

# --- 5d: Read exercise data from summary sheet ---
wb_sum = load_workbook(DST, data_only=True)
ws_sum = wb_sum['summary']

# Column mappings from Weight2026.xlsx summary sheet row 2
# Core/bodyweight (single-column, no weight):
#   col 10=hang, 11=plank, 12=knee lift, 25=push ups, 38=TRX pullup, 39=neg pull up
# PUSH exercises (wt_col/rep_col pairs):
#   13/14=bench, 15/16=incl.bench, 17/18=mil press, 19/20=HS sh press
#   21/22=pec fly, 23/24=Lat lift, 26/27=tric pull, 28/29=skull crush
# PULL exercises:
#   30/31=deadlift, 32/33=Row, 34/35=bent row, 36/37=lat pull dn
#   40/41=face pull, 42/43=cable curl, 44/45=curl, 46/47=hammer curl
#   48/49=back ext
# LEGS exercises:
#   50/51=squats, 52/53=land mine squat, 54/55=lunge walk, 56/57=leg press
#   58/59=hip thrust, 60/61=Hstring curl, 62/63=weight step
STRENGTH_EXERCISES = OrderedDict([
    # PUSH
    ('bench', {'wt_col': 13, 'rep_col': 14, 'cat': 'PUSH'}),
    ('incl bench', {'wt_col': 15, 'rep_col': 16, 'cat': 'PUSH'}),
    ('mil press', {'wt_col': 17, 'rep_col': 18, 'cat': 'PUSH'}),
    ('HS sh press', {'wt_col': 19, 'rep_col': 20, 'cat': 'PUSH'}),
    ('pec fly', {'wt_col': 21, 'rep_col': 22, 'cat': 'PUSH'}),
    ('lat lift', {'wt_col': 23, 'rep_col': 24, 'cat': 'PUSH'}),
    ('push ups', {'wt_col': 25, 'rep_col': 25, 'cat': 'PUSH'}),
    ('tric pull', {'wt_col': 26, 'rep_col': 27, 'cat': 'PUSH'}),
    ('skull crush', {'wt_col': 28, 'rep_col': 29, 'cat': 'PUSH'}),
    # PULL
    ('deadlift', {'wt_col': 30, 'rep_col': 31, 'cat': 'PULL'}),
    ('row', {'wt_col': 32, 'rep_col': 33, 'cat': 'PULL'}),
    ('bent row', {'wt_col': 34, 'rep_col': 35, 'cat': 'PULL'}),
    ('lat pull dn', {'wt_col': 36, 'rep_col': 37, 'cat': 'PULL'}),
    ('face pull', {'wt_col': 40, 'rep_col': 41, 'cat': 'PULL'}),
    ('cable curl', {'wt_col': 42, 'rep_col': 43, 'cat': 'PULL'}),
    ('curl', {'wt_col': 44, 'rep_col': 45, 'cat': 'PULL'}),
    ('hammer curl', {'wt_col': 46, 'rep_col': 47, 'cat': 'PULL'}),
    ('back ext', {'wt_col': 48, 'rep_col': 49, 'cat': 'CORE'}),
    # LEGS
    ('squats', {'wt_col': 50, 'rep_col': 51, 'cat': 'LEGS'}),
    ('land mine squat', {'wt_col': 52, 'rep_col': 53, 'cat': 'LEGS'}),
    ('lunge walk', {'wt_col': 54, 'rep_col': 55, 'cat': 'LEGS'}),
    ('leg press', {'wt_col': 56, 'rep_col': 57, 'cat': 'LEGS'}),
    ('hip thrust', {'wt_col': 58, 'rep_col': 59, 'cat': 'LEGS'}),
    ('Hstring curl', {'wt_col': 60, 'rep_col': 61, 'cat': 'LEGS'}),
    ('weight step', {'wt_col': 62, 'rep_col': 63, 'cat': 'LEGS'}),
])

# Core/bodyweight exercises (single column, reps or time only)
CORE_EXERCISES = OrderedDict([
    ('hang', {'col': 10, 'cat': 'CORE'}),
    ('plank', {'col': 11, 'cat': 'CORE'}),
    ('knee lift', {'col': 12, 'cat': 'CORE'}),
    ('TRX pullup', {'col': 38, 'cat': 'PULL'}),
    ('neg pull up', {'col': 39, 'cat': 'PULL'}),
])

CARDIO = {6: ('run', 7), 8: ('vest walk', 9)}

# Scan for custom exercises added by import_gym_log (columns after static exercises, before notes)
CUSTOM_START_COL_1 = 64  # 1-indexed, first column after weight step reps (col 63)
_custom_count = 0
_wb_scan = load_workbook(SRC, data_only=True)
_ws_scan = _wb_scan['summary']
_scan_col = CUSTOM_START_COL_1
while True:
    header = _ws_scan.cell(row=2, column=_scan_col).value
    if header is None or str(header).strip() == "" or str(header).strip().lower() == 'notes':
        break
    log_name = str(header).strip()
    reps_header = _ws_scan.cell(row=2, column=_scan_col + 1).value
    has_reps = reps_header is not None and "reps" in str(reps_header).lower()
    if has_reps:
        STRENGTH_EXERCISES[log_name] = {'wt_col': _scan_col, 'rep_col': _scan_col + 1, 'cat': 'CUSTOM'}
        _scan_col += 2
    else:
        CORE_EXERCISES[log_name] = {'col': _scan_col, 'cat': 'CUSTOM'}
        _scan_col += 1
    _custom_count += 1
_wb_scan.close()
if _custom_count:
    print(f"Found {_custom_count} custom exercise(s) in summary sheet")

# Use the actual date range from the data
START_DATE = datetime(2026, 1, 1)
last_date = valid[2].iloc[-1]
END_DATE = last_date if isinstance(last_date, datetime) else datetime(2026, 12, 31)

# re is already imported at the top

def _parse_single_weight(s):
    """Parse a single weight value. Returns (weight, sets_from_weight).
    Handles: 60, '2x27.5' (dumbbell pair), '15 ktl', '15lb'."""
    s = s.strip()
    # NxW format (e.g., '2x27.5' = pair of 27.5 lb dumbbells)
    m = re.match(r'(\d+)\s*[xX]\s*([\d.]+)', s)
    if m:
        return float(m.group(2)), int(m.group(1))
    # Plain number with optional suffix
    m = re.match(r'^([\d.]+)', s)
    if m:
        return float(m.group(1)), 1
    return 0, 1

def _parse_single_reps(s):
    """Parse a single rep value. Returns (reps, sets).
    Handles: '10x3', '12', '3x10 sl'."""
    s = s.strip()
    m = re.match(r'(\d+)\s*[xX]\s*(\d+)', s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r'^(\d+)', s)
    if m:
        return int(m.group(1)), 1
    return 0, 1

def parse_sets_reps(wt_val, rep_val):
    """Parse weight and rep columns from summary sheet.
    Supports comma-separated pairs: '140,150' with '15x3,10x3'
    means 3 sets of 15 at 140 + 3 sets of 10 at 150.
    Returns list of (weight, sets, reps) tuples."""
    if wt_val is None and rep_val is None:
        return [(0, 1, 0)]

    wt_str = str(wt_val).strip() if wt_val is not None else '0'
    rep_str = str(rep_val).strip() if rep_val is not None else '0'

    # Split by comma or semicolon to handle multi-weight/rep entries
    # Semicolons avoid Excel interpreting "140,150" as the integer 140150
    wt_parts = [p.strip() for p in re.split(r'[,;]', wt_str) if p.strip()]
    rep_parts = [p.strip() for p in re.split(r'[,;]', rep_str) if p.strip()]

    # If only one part on either side, replicate to match the other
    if len(wt_parts) == 1 and len(rep_parts) > 1:
        wt_parts = wt_parts * len(rep_parts)
    elif len(rep_parts) == 1 and len(wt_parts) > 1:
        rep_parts = rep_parts * len(wt_parts)
    elif len(wt_parts) != len(rep_parts):
        # Mismatch — fall back to pairing what we can
        max_len = max(len(wt_parts), len(rep_parts))
        while len(wt_parts) < max_len: wt_parts.append(wt_parts[-1] if wt_parts else '0')
        while len(rep_parts) < max_len: rep_parts.append(rep_parts[-1] if rep_parts else '0')

    results = []
    for wp, rp in zip(wt_parts, rep_parts):
        weight, wt_sets = _parse_single_weight(wp)
        reps, rep_sets = _parse_single_reps(rp)
        sets = max(wt_sets, rep_sets)
        if weight > 0 or reps > 0:
            results.append((weight, sets, reps))

    return results if results else [(0, 1, 0)]

def parse_pace_seconds(pace_str, distance_str=None):
    """Parse a run/walk time string into total seconds.
    Supports: MM:SS, M:SS, HH:MM:SS (Excel time format), NmNs, N min, N/mi.
    If distance is a 5K, converts total time to per-mile pace."""
    if pace_str is None: return None
    s = str(pace_str).strip()
    total_sec = None

    # MM:SS or M:SS (primary format, e.g. "8:41", "29:30")
    m = re.match(r'^(\d{1,2}):(\d{2})$', s)
    if m:
        total_sec = int(m.group(1)) * 60 + int(m.group(2))

    # HH:MM:SS — Excel time format (e.g. "08:41:00", "00:29:30")
    # Treat as MM:SS:00 when first field < 60 (it's minutes, not hours)
    if total_sec is None:
        m = re.match(r'^(\d{1,2}):(\d{2}):(\d{2})$', s)
        if m:
            a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if a < 60 and c == 0:
                # e.g. "08:41:00" -> 8 min 41 sec
                total_sec = a * 60 + b
            else:
                # e.g. "00:29:30" -> 29 min 30 sec
                total_sec = a * 3600 + b * 60 + c

    # NmNs format (e.g. "29m40s", "29m1s")
    if total_sec is None:
        m = re.match(r'(\d+)m\s*(\d+)s', s)
        if m:
            total_sec = int(m.group(1)) * 60 + int(m.group(2))

    # N min (e.g. "10 min")
    if total_sec is None:
        m = re.match(r'([\d.]+)\s*min', s)
        if m:
            total_sec = float(m.group(1)) * 60

    # N/mi (e.g. "8.5/mi") — already a per-mile pace
    if total_sec is None:
        m = re.match(r'([\d.]+)\s*/?\s*mi', s)
        if m:
            return float(m.group(1)) * 60  # already per-mile, return directly

    if total_sec is None:
        return None

    # If the distance is a 5K (or multi-mile), convert total time to per-mile pace
    if distance_str:
        dist_mi = parse_distance(distance_str)
        if dist_mi and dist_mi > 1.5:
            return total_sec / dist_mi  # per-mile pace in seconds

    return total_sec

def parse_distance(d):
    s = str(d).strip().lower()
    m = re.match(r'([\d.]+)\s*k', s)
    if m: return float(m.group(1)) * 0.621
    m = re.match(r'([\d.]+)\s*mi', s)
    if m: return float(m.group(1))
    m = re.match(r'([\d.]+)\s*m$', s)
    if m: return float(m.group(1))
    try: return float(s)
    except ValueError: return 0

def generate_commentary(week_key, wk_strength, runs, walks, walk_paces, run_paces,
                       avg_wt, prev_weight, avg_lean, prev_lean, prev_strength,
                       avg_fat, prev_fat, avg_muscle, prev_muscle, avg_almi, prev_almi,
                       training_days):
    """Generate a comprehensive weekly summary covering strength, body comp, cardio, and training patterns."""
    sections = []

    # --- 1. Strength progression (PRs, volume, notable lifts) ---
    total_sets = sum(sum(d['sets'] for d in ex_data) for ex_data in wk_strength.values())
    if total_sets == 0:
        sections.append("No strength training this week.")
    else:
        # PRs vs previous week
        prs = []
        if prev_strength and wk_strength:
            for ex_name, ex_data in wk_strength.items():
                if ex_name in prev_strength:
                    prev_max = max(d['weight'] for d in prev_strength[ex_name])
                    curr_max = max(d['weight'] for d in ex_data)
                    diff = curr_max - prev_max
                    if diff > 0:
                        prs.append(f"{ex_name} +{diff:.0f}lb ({curr_max:.0f}lb)")
                    elif diff < -5:
                        prs.append(f"{ex_name} -{abs(diff):.0f}lb ({curr_max:.0f}lb)")
        # Top lifts this week
        top_lifts = []
        for ex_name, ex_data in wk_strength.items():
            max_wt = max(d['weight'] for d in ex_data)
            best_reps = max(d['reps'] for d in ex_data if d['weight'] == max_wt)
            top_lifts.append((ex_name, max_wt, best_reps))
        top_lifts.sort(key=lambda x: x[1], reverse=True)

        strength_parts = []
        if prs:
            strength_parts.append(f"PRs: {', '.join(prs[:3])}")
        strength_parts.append(f"{total_sets} total sets across {len(wk_strength)} exercises")
        # Top 3 heaviest lifts
        top3 = [f"{n} {w:.0f}lbx{r:.0f}" for n, w, r in top_lifts[:3]]
        strength_parts.append(f"Top lifts: {', '.join(top3)}")

        # Missing muscle groups — use category from STRENGTH_EXERCISES/CORE_EXERCISES
        has_push = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'PUSH'])
        has_pull = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'PULL'])
        has_legs = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'LEGS'])
        missing = []
        if not has_legs: missing.append('LEGS')
        if not has_push: missing.append('PUSH')
        if not has_pull: missing.append('PULL')
        if missing:
            strength_parts.append(f"Missing: {', '.join(missing)}")
        sections.append(' | '.join(strength_parts))

    # --- 2. Body composition trends ---
    comp_parts = []
    if avg_wt is not None:
        wt_str = f"Avg weight: {avg_wt:.1f}lb"
        if prev_weight:
            wt_diff = avg_wt - prev_weight
            direction = "up" if wt_diff > 0 else "down"
            wt_str += f" ({direction} {abs(wt_diff):.1f}lb)"
        comp_parts.append(wt_str)

    if avg_lean is not None and prev_lean:
        lean_diff = avg_lean - prev_lean
        comp_parts.append(f"Lean mass: {avg_lean:.1f}lb ({'+' if lean_diff >= 0 else ''}{lean_diff:.1f}lb)")

    if avg_fat is not None:
        fat_str = f"Body fat: {avg_fat*100:.1f}%"
        if prev_fat:
            fat_diff = (avg_fat - prev_fat) * 100
            fat_str += f" ({'+' if fat_diff >= 0 else ''}{fat_diff:.1f}pp)"
        comp_parts.append(fat_str)

    if avg_muscle is not None and prev_muscle:
        musc_diff = avg_muscle - prev_muscle
        comp_parts.append(f"Muscle: {avg_muscle:.1f}lb ({'+' if musc_diff >= 0 else ''}{musc_diff:.1f}lb)")

    if avg_almi is not None:
        comp_parts.append(f"ALMI: {avg_almi:.2f}")

    if comp_parts:
        # Interpretation
        if avg_wt and prev_weight and avg_lean and prev_lean:
            wt_diff = avg_wt - prev_weight
            lean_diff = avg_lean - prev_lean
            fat_diff_lb = wt_diff - lean_diff
            if lean_diff > 0.3 and fat_diff_lb < 0.2:
                comp_parts.append("-> Recomposition progressing well")
            elif wt_diff > 0.5 and lean_diff > 0.3:
                comp_parts.append("-> Gaining weight with lean mass - bulk on track")
            elif wt_diff > 0.5 and lean_diff <= 0.3:
                comp_parts.append("-> Weight up without lean gain - possible excess calories")
            elif wt_diff < -0.5 and lean_diff >= -0.1:
                comp_parts.append("-> Cutting fat while preserving muscle")
            elif wt_diff < -0.5 and lean_diff < -0.3:
                comp_parts.append("-> Losing both weight and lean mass - increase protein/volume")
        sections.append(' | '.join(comp_parts))

    # --- 3. Cardio summary ---
    cardio_parts = []
    if runs:
        run_mi = sum(parse_distance(r) for r in runs)
        cardio_parts.append(f"{len(runs)} run(s), {run_mi:.1f} mi total")
        if run_paces:
            best = min(run_paces)
            avg_pace = sum(run_paces) / len(run_paces)
            cardio_parts.append(f"best {int(best//60)}:{int(best%60):02d}/mi, avg {int(avg_pace//60)}:{int(avg_pace%60):02d}/mi")
    if walks:
        walk_mi = sum(parse_distance(w) for w in walks)
        cardio_parts.append(f"{len(walks)} walk(s), {walk_mi:.1f} mi")
    if not runs and not walks:
        cardio_parts.append("No cardio logged")
    if cardio_parts:
        sections.append("Cardio: " + ' | '.join(cardio_parts))

    # --- 4. Training frequency ---
    n_days = len(training_days)
    day_list = ', '.join(sorted(training_days))
    sections.append(f"Trained {n_days} day(s): {day_list}" if n_days > 0 else "Rest week - no training logged")

    return ' || '.join(sections)

# Collect weekly data
weekly_strength = defaultdict(lambda: defaultdict(list))
weekly_cardio = defaultdict(lambda: {'runs': [], 'run_paces': [], 'walks': [], 'walk_paces': []})
weekly_training_days = defaultdict(set)
weekly_weight = defaultdict(list)
weekly_lean_mass = defaultdict(list)
weekly_fat_pct = defaultdict(list)
weekly_muscle = defaultdict(list)
weekly_almi = defaultdict(list)

for r in range(3, ws_sum.max_row + 1):
    date_val = ws_sum.cell(r, 3).value
    if not isinstance(date_val, datetime):
        continue
    if date_val < START_DATE or date_val > END_DATE:
        continue
    monday = date_val - timedelta(days=date_val.weekday())
    week_key = monday.strftime('%Y-%m-%d')

    day_has_exercise = False
    for ex_name, info in STRENGTH_EXERCISES.items():
        wt_val = ws_sum.cell(r, info['wt_col']).value
        if wt_val is not None:
            if info['wt_col'] == info['rep_col']:
                # Single-column bodyweight exercise (e.g., push ups: '3x10 sl')
                # Treat NxM as sets x reps (no weight)
                s_val = str(wt_val).strip()
                m_bw = re.match(r'(\d+)\s*[xX]\s*(\d+)', s_val)
                if m_bw:
                    s, rp = int(m_bw.group(1)), int(m_bw.group(2))
                else:
                    m_bw = re.match(r'^(\d+)', s_val)
                    rp = int(m_bw.group(1)) if m_bw else 1
                    s = 1
                w = 0
            else:
                rep_val = ws_sum.cell(r, info['rep_col']).value
                parsed_list = parse_sets_reps(wt_val, rep_val)
                for w, s, rp in parsed_list:
                    if w > 0 or (rp > 0 and s > 0):
                        weekly_strength[week_key][ex_name].append({'weight': w, 'sets': s, 'reps': rp, 'date': date_val})
                day_has_exercise = True
                continue
            # For push-ups and bodyweight exercises, weight may be 0 but we still count sets
            if w > 0 or (rp > 0 and s > 0):
                weekly_strength[week_key][ex_name].append({'weight': w, 'sets': s, 'reps': rp, 'date': date_val})
            day_has_exercise = True

    # Core/bodyweight exercises (single column — reps or time, no separate weight)
    for ex_name, info in CORE_EXERCISES.items():
        val = ws_sum.cell(r, info['col']).value
        if val is not None:
            s_val = str(val).strip()
            # Parse simple rep counts or NxS format
            m = re.match(r'(\d+)\s*[xX]\s*(\d+)', s_val)
            if m:
                rp, s = int(m.group(1)), int(m.group(2))
            else:
                m = re.match(r'^(\d+)', s_val)
                rp = int(m.group(1)) if m else 1
                s = 1
            weekly_strength[week_key][ex_name].append({'weight': 0, 'sets': s, 'reps': rp, 'date': date_val})
            day_has_exercise = True

    for col_idx, (cardio_name, pace_col) in CARDIO.items():
        dist = ws_sum.cell(r, col_idx).value
        pace = ws_sum.cell(r, pace_col).value
        if dist:
            pace_sec = parse_pace_seconds(pace, str(dist))
            if cardio_name == 'run':
                weekly_cardio[week_key]['runs'].append(str(dist))
                if pace_sec: weekly_cardio[week_key]['run_paces'].append(pace_sec)
            else:
                weekly_cardio[week_key]['walks'].append(str(dist))
                if pace_sec: weekly_cardio[week_key]['walk_paces'].append(pace_sec)
            day_has_exercise = True

    # Also check lunge walk (54) and weight step (62) for day_has_exercise
    for c_idx in [54, 62]:
        if ws_sum.cell(r, c_idx).value: day_has_exercise = True

    if day_has_exercise:
        weekly_training_days[week_key].add(date_val.strftime('%a'))

# DEXA body comp from numpy arrays (already computed above)
for dt, vals in dexa_by_date.items():
    if dt < START_DATE or dt > END_DATE: continue
    monday = dt - timedelta(days=dt.weekday())
    wk = monday.strftime('%Y-%m-%d')
    if not np.isnan(vals['weight']): weekly_weight[wk].append(vals['weight'])
    if not np.isnan(vals['fat_pct']): weekly_fat_pct[wk].append(vals['fat_pct'])
    if not np.isnan(vals['lean_mass']): weekly_lean_mass[wk].append(vals['lean_mass'])
    if not np.isnan(vals['muscle']): weekly_muscle[wk].append(vals['muscle'])
    if not np.isnan(vals['almi']): weekly_almi[wk].append(vals['almi'])

wb_sum.close()

# --- 5e: Build ExerciseSummary in temp workbook ---
# Reuse style imports under different names to avoid clashing with build_v4's styles
Font2 = Font
Align2 = Alignment
PF2 = PatternFill
Bdr2 = Border
Side2 = Side

temp_wb = Workbook()
es = temp_wb.active
es.title = 'ExerciseSummary'

# Styles
ES_HEADER_FILL = PF2(start_color='2F5496', end_color='2F5496', fill_type='solid')
ES_SUBHEADER_FILL = PF2(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
ES_PUSH_FILL = PF2(start_color='C55A11', end_color='C55A11', fill_type='solid')
ES_PULL_FILL = PF2(start_color='548235', end_color='548235', fill_type='solid')
ES_LEGS_FILL = PF2(start_color='7030A0', end_color='7030A0', fill_type='solid')
ES_CARDIO_FILL = PF2(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
ES_BODY_FILL = PF2(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
ES_UP_FILL = PF2(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
ES_DOWN_FILL = PF2(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
ES_NEUTRAL_FILL = PF2(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
ES_WHITE_FONT = Font2(bold=True, color='FFFFFF', size=10)
ES_HEADER_FONT = Font2(bold=True, color='FFFFFF', size=11)
ES_THIN_BORDER = Bdr2(
    left=Side2(style='thin', color='D0D0D0'), right=Side2(style='thin', color='D0D0D0'),
    top=Side2(style='thin', color='D0D0D0'), bottom=Side2(style='thin', color='D0D0D0'),
)
CAT_FILLS = {'PUSH': ES_PUSH_FILL, 'PULL': ES_PULL_FILL, 'LEGS': ES_LEGS_FILL}

# Layout
col = 1
es.cell(1, 1, 'OVERVIEW').font = ES_HEADER_FONT
es.cell(1, 1).fill = ES_HEADER_FILL
es.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
for i, h in enumerate(['Week Of', 'Days'], 1):
    c = es.cell(2, i, h)
    c.font = Font2(bold=True, size=9); c.fill = ES_SUBHEADER_FILL; c.border = ES_THIN_BORDER
    c.alignment = Align2(horizontal='center', wrap_text=True)
col = 3

# BODY COMP
es.cell(1, col, 'BODY COMP').font = ES_WHITE_FONT
es.cell(1, col).fill = ES_BODY_FILL
es.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 4)
for h in ['DEXA Wt', 'Lean Mass', 'DEXA Fat %', 'Muscle', 'ALMI']:
    c = es.cell(2, col, h)
    c.font = Font2(bold=True, size=9); c.fill = ES_SUBHEADER_FILL; c.border = ES_THIN_BORDER
    c.alignment = Align2(horizontal='center', wrap_text=True)
    col += 1

# CARDIO
cardio_start = col
es.cell(1, col, 'CARDIO').font = ES_WHITE_FONT
es.cell(1, col).fill = ES_CARDIO_FILL
es.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
for h in ['Runs', 'Best Pace', 'Walks', 'Walk Mi']:
    c = es.cell(2, col, h)
    c.font = Font2(bold=True, size=9); c.fill = ES_SUBHEADER_FILL; c.border = ES_THIN_BORDER
    c.alignment = Align2(horizontal='center', wrap_text=True)
    col += 1

# STRENGTH (including core/bodyweight exercises grouped by their category)
ES_CORE_FILL = PF2(start_color='0070C0', end_color='0070C0', fill_type='solid')
ES_CUSTOM_FILL = PF2(start_color='808080', end_color='808080', fill_type='solid')
CAT_FILLS['CORE'] = ES_CORE_FILL
CAT_FILLS['CUSTOM'] = ES_CUSTOM_FILL
strength_col_map = {}
for cat in ['PUSH', 'PULL', 'LEGS', 'CORE', 'CUSTOM']:
    cat_exercises = [(n, i) for n, i in STRENGTH_EXERCISES.items() if i['cat'] == cat]
    # Also include core/bodyweight exercises in their respective category
    cat_exercises += [(n, i) for n, i in CORE_EXERCISES.items() if i['cat'] == cat]
    if not cat_exercises: continue
    start_col = col
    for ex_name, _ in cat_exercises:
        c = es.cell(2, col, ex_name)
        c.font = Font2(bold=True, size=8); c.fill = ES_SUBHEADER_FILL; c.border = ES_THIN_BORDER
        c.alignment = Align2(horizontal='center', wrap_text=True, vertical='bottom')
        strength_col_map[ex_name] = col
        col += 1
    es.cell(1, start_col, cat).font = ES_WHITE_FONT
    es.cell(1, start_col).fill = CAT_FILLS[cat]
    es.cell(1, start_col).alignment = Align2(horizontal='center', wrap_text=True)
    if col - 1 > start_col:
        es.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)

# COMMENTARY
commentary_col = col
es.cell(1, col, 'COMMENTARY').font = ES_HEADER_FONT
es.cell(1, col).fill = ES_HEADER_FILL
es.cell(2, col, 'Progress Notes').font = Font2(bold=True, size=9)
es.cell(2, col).fill = ES_SUBHEADER_FILL; es.cell(2, col).border = ES_THIN_BORDER
es.cell(2, col).alignment = Align2(horizontal='left', wrap_text=True)
col += 1

# ASSESSMENT
assess_col = col
es.cell(1, col, 'ASSESSMENT').font = ES_HEADER_FONT
es.cell(1, col).fill = ES_HEADER_FILL
es.cell(2, col, 'Weekly Notes').font = Font2(bold=True, size=9)
es.cell(2, col).fill = ES_SUBHEADER_FILL; es.cell(2, col).border = ES_THIN_BORDER
es.cell(2, col).alignment = Align2(horizontal='center', wrap_text=True)

# Write data rows
all_mondays = []
d = START_DATE - timedelta(days=START_DATE.weekday())
while d <= END_DATE:
    all_mondays.append(d.strftime('%Y-%m-%d'))
    d += timedelta(days=7)

row = 3
prev_strength = {}
prev_weight_es = None
prev_lean_es = None
prev_fat_es = None
prev_muscle_es = None
prev_almi_es = None

for week_key in all_mondays:
    training_days = len(weekly_training_days.get(week_key, set()))
    if training_days == 0 and not weekly_strength.get(week_key) and not weekly_cardio.get(week_key, {}).get('runs'):
        continue

    monday_date = datetime.strptime(week_key, '%Y-%m-%d')
    week_label = monday_date.strftime('%b %d')

    es.cell(row, 1, week_label).font = Font2(bold=True, size=10)
    es.cell(row, 1).border = ES_THIN_BORDER; es.cell(row, 1).alignment = Align2(horizontal='center', wrap_text=True)
    es.cell(row, 2, training_days).border = ES_THIN_BORDER; es.cell(row, 2).alignment = Align2(horizontal='center', wrap_text=True)

    # Body Comp
    avg_wt = sum(weekly_weight[week_key]) / len(weekly_weight[week_key]) if weekly_weight.get(week_key) else None
    avg_lean = sum(weekly_lean_mass[week_key]) / len(weekly_lean_mass[week_key]) if weekly_lean_mass.get(week_key) else None
    avg_fat = sum(weekly_fat_pct[week_key]) / len(weekly_fat_pct[week_key]) if weekly_fat_pct.get(week_key) else None
    avg_muscle = sum(weekly_muscle[week_key]) / len(weekly_muscle[week_key]) if weekly_muscle.get(week_key) else None
    avg_almi = sum(weekly_almi[week_key]) / len(weekly_almi[week_key]) if weekly_almi.get(week_key) else None

    if avg_wt is not None:
        c = es.cell(row, 3, round(avg_wt, 1)); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True); c.number_format = '0.0'
        if prev_weight_es:
            diff = avg_wt - prev_weight_es
            c.fill = ES_DOWN_FILL if diff < -0.5 else (ES_UP_FILL if diff > 0.5 else ES_NEUTRAL_FILL)
    if avg_lean is not None:
        c = es.cell(row, 4, round(avg_lean, 1)); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True); c.number_format = '0.0'
        if prev_lean_es:
            diff = avg_lean - prev_lean_es
            c.fill = ES_UP_FILL if diff > 0.3 else (ES_DOWN_FILL if diff < -0.3 else ES_NEUTRAL_FILL)
    if avg_fat is not None:
        c = es.cell(row, 5, round(avg_fat * 100, 1)); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True); c.number_format = '0.0'
    if avg_muscle is not None:
        c = es.cell(row, 6, round(avg_muscle, 1)); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True); c.number_format = '0.0'
    if avg_almi is not None:
        c = es.cell(row, 7, round(avg_almi, 2)); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True); c.number_format = '0.00'

    # Cardio
    wk_cardio = weekly_cardio.get(week_key, {})
    runs = wk_cardio.get('runs', [])
    run_paces = wk_cardio.get('run_paces', [])
    walks = wk_cardio.get('walks', [])

    if runs:
        run_mi = sum(parse_distance(r) for r in runs)
        c = es.cell(row, 8, f"{len(runs)} x {run_mi:.1f}"); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True)
        if run_paces:
            best = min(run_paces)
            c = es.cell(row, 9, f"{int(best//60)}:{int(best%60):02d}/mi"); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True)

    if walks:
        walk_mi = sum(parse_distance(w) for w in walks)
        c = es.cell(row, 10, f"{len(walks)}"); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True)
        c = es.cell(row, 11, f"{walk_mi:.1f}"); c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='center', wrap_text=True)

    # Strength — volume load (W = weight × reps × sets) per exercise
    wk_strength = weekly_strength.get(week_key, {})
    ex_details = []  # For commentary: (name, volume, breakdown_str)
    for ex_name, col_idx in strength_col_map.items():
        if ex_name in wk_strength:
            ex_data = wk_strength[ex_name]

            # Group by (weight, reps) to merge identical sessions
            from collections import OrderedDict as _OD
            combo_groups = _OD()  # (weight, reps) -> total_sets
            for d_item in ex_data:
                key = (d_item['weight'], d_item['reps'])
                combo_groups[key] = combo_groups.get(key, 0) + d_item['sets']

            # Compute volume load: sum of (weight × reps × sets) for each group
            volume = 0
            parts = []
            has_weight = any(wt > 0 for wt, _ in combo_groups)
            for (wt, rp), s in combo_groups.items():
                volume += wt * rp * s
                wt_str = f"{wt:g}"
                if has_weight and wt > 0:
                    parts.append(f"{s}x{rp}@{wt_str}")
                else:
                    parts.append(f"{s}x{rp}")

            # Cell: show volume as compact number
            if has_weight and volume > 0:
                if volume >= 10000:
                    label = f"{volume/1000:.1f}k"
                else:
                    label = f"{volume:,.0f}"
            else:
                # Bodyweight/core: show total reps (volume = reps × sets)
                total_reps = sum(rp * s for (_, rp), s in combo_groups.items())
                label = f"{total_reps}r"

            # Color: green if up from last week, red if down
            c = es.cell(row, col_idx, label)
            c.border = ES_THIN_BORDER
            c.alignment = Align2(horizontal='center', wrap_text=True, vertical='center')
            c.font = Font2(size=9, bold=True)
            c.number_format = '@'  # Force text

            if ex_name in prev_strength and prev_strength[ex_name]:
                prev_vol = sum(d['weight'] * d['reps'] * d['sets'] for d in prev_strength[ex_name])
                if prev_vol > 0:
                    pct = (volume - prev_vol) / prev_vol
                    if pct > 0.05:
                        c.fill = ES_UP_FILL
                    elif pct < -0.05:
                        c.fill = ES_DOWN_FILL

            # Save detail for commentary
            breakdown = ','.join(parts)
            ex_details.append((ex_name, volume, breakdown))

    # Commentary — include exercise breakdown details
    commentary_text = generate_commentary(
        week_key, wk_strength, runs, walks, wk_cardio.get('walk_paces', []), run_paces,
        avg_wt, prev_weight_es, avg_lean, prev_lean_es, prev_strength,
        avg_fat, prev_fat_es, avg_muscle, prev_muscle_es, avg_almi, prev_almi_es,
        weekly_training_days.get(week_key, set())
    )
    # Append exercise details to commentary
    if ex_details:
        detail_lines = []
        for name, vol, breakdown in ex_details:
            if vol > 0:
                detail_lines.append(f"{name}: {vol:,.0f}lb ({breakdown})")
            else:
                detail_lines.append(f"{name}: {breakdown}")
        commentary_text += " || Detail: " + " | ".join(detail_lines)

    c = es.cell(row, commentary_col, commentary_text)
    c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='left', wrap_text=True, vertical='top'); c.font = Font2(size=8)

    # Assessment — structured weekly grade
    assessment_parts = []

    # Volume rating — use total volume load (lbs moved)
    total_vol = sum(d['weight'] * d['reps'] * d['sets'] for ex_data in wk_strength.values() for d in ex_data)
    total_sets = sum(d['sets'] for ex_data in wk_strength.values() for d in ex_data)
    n_exercises = len(wk_strength)
    n_train_days = len(weekly_training_days.get(week_key, set()))
    if total_vol > 0:
        vol_str = f"{total_vol/1000:.1f}k lbs" if total_vol >= 1000 else f"{total_vol:,.0f} lbs"
        if total_sets >= 20:
            assessment_parts.append(f"Volume: HIGH ({vol_str}, {total_sets}s, {n_exercises} ex, {n_train_days}d)")
        elif total_sets >= 10:
            assessment_parts.append(f"Volume: MOD ({vol_str}, {total_sets}s, {n_exercises} ex, {n_train_days}d)")
        else:
            assessment_parts.append(f"Volume: LOW ({vol_str}, {total_sets}s, {n_exercises} ex, {n_train_days}d)")
    else:
        assessment_parts.append(f"Volume: REST WEEK ({n_train_days} active days, cardio only)")

    # Body comp direction with detail
    if avg_wt and prev_weight_es and avg_lean and prev_lean_es:
        wt_diff = avg_wt - prev_weight_es
        lean_diff = avg_lean - prev_lean_es
        fat_change_lb = wt_diff - lean_diff
        if lean_diff > 0.3 and fat_change_lb < 0.2:
            assessment_parts.append(f"Comp: RECOMP (+{lean_diff:.1f}lb lean, {fat_change_lb:+.1f}lb fat)")
        elif wt_diff > 0.5 and lean_diff > 0.3:
            assessment_parts.append(f"Comp: LEAN BULK (+{lean_diff:.1f}lb lean, +{fat_change_lb:.1f}lb fat)")
        elif wt_diff > 0.5 and lean_diff <= 0.3:
            assessment_parts.append(f"Comp: FAT GAIN (+{wt_diff:.1f}lb weight, only +{lean_diff:.1f}lb lean)")
        elif wt_diff < -0.5 and lean_diff >= -0.1:
            assessment_parts.append(f"Comp: CUT ({wt_diff:+.1f}lb weight, lean preserved)")
        elif wt_diff < -0.5 and lean_diff < -0.3:
            assessment_parts.append(f"Comp: MUSCLE LOSS ({wt_diff:+.1f}lb weight, {lean_diff:+.1f}lb lean)")
        else:
            assessment_parts.append(f"Comp: MAINTENANCE (wt {wt_diff:+.1f}lb, lean {lean_diff:+.1f}lb)")

    # Training balance — derive from exercise category definitions
    has_push = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'PUSH'])
    has_pull = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'PULL'])
    has_legs = any(ex in wk_strength for ex in [n for n, i in list(STRENGTH_EXERCISES.items()) + list(CORE_EXERCISES.items()) if i.get('cat') == 'LEGS'])
    has_cardio = bool(runs) or bool(walks)
    balance_parts = []
    if has_push: balance_parts.append("PUSH")
    if has_pull: balance_parts.append("PULL")
    if has_legs: balance_parts.append("LEGS")
    if has_cardio: balance_parts.append("CARDIO")
    if balance_parts:
        grade = "BALANCED" if has_push and has_pull and has_legs else "INCOMPLETE"
        assessment_parts.append(f"Split: {grade} ({'+'.join(balance_parts)})")

    c = es.cell(row, assess_col, " | ".join(assessment_parts))
    c.border = ES_THIN_BORDER; c.alignment = Align2(horizontal='left', wrap_text=True, vertical='top'); c.font = Font2(size=8)

    # Update prev for next iteration
    if avg_wt: prev_weight_es = avg_wt
    if avg_lean: prev_lean_es = avg_lean
    if avg_fat: prev_fat_es = avg_fat
    if avg_muscle: prev_muscle_es = avg_muscle
    if avg_almi: prev_almi_es = avg_almi
    prev_strength = dict(wk_strength)
    row += 1

# Column widths — narrow columns with word wrap to fit on screen
es.column_dimensions['A'].width = 7      # Week Of
es.column_dimensions['B'].width = 4.5    # Days
for i in range(3, 8):                    # Body Comp (Wt, Lean, Fat%, Muscle, ALMI)
    es.column_dimensions[get_column_letter(i)].width = 7
for i in range(8, 12):                   # Cardio (Runs, Pace, Walks, Walk Mi)
    es.column_dimensions[get_column_letter(i)].width = 7.5
for i in range(12, commentary_col):      # Strength exercise columns
    es.column_dimensions[get_column_letter(i)].width = 7
es.column_dimensions[get_column_letter(commentary_col)].width = 40
es.column_dimensions[get_column_letter(assess_col)].width = 35

# Set row height for header row to accommodate wrapped text
es.row_dimensions[2].height = 30

temp_path = os.path.join(tempfile.mkdtemp(), 'temp_es.xlsx')
temp_wb.save(temp_path)
# Resave to clean format
temp2 = load_workbook(temp_path)
temp2_path = temp_path.replace('.xlsx', '_clean.xlsx')
temp2.save(temp2_path)
temp2.close()

print(f"Built ExerciseSummary: {row - 3} data rows, {col} columns")

# --- 5f: XML injection: chart range fixes + ExerciseSummary + repackage ---
temp_dir = tempfile.mkdtemp()
with zipfile.ZipFile(DST, 'r') as z:
    z.extractall(temp_dir)

# --- Remove existing ExerciseSummary if present (from a previous pipeline run) ---
_wb_xml_pre_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
with open(_wb_xml_pre_path, 'r', encoding='utf-8') as f:
    _wb_pre = f.read()

_old_es = re.search(r'<sheet\s+name="ExerciseSummary"[^/]*r:id="(rId\d+)"[^/]*/>', _wb_pre)
if not _old_es:
    _old_es = re.search(r'<sheet[^>]*name="ExerciseSummary"[^>]*r:id="(rId\d+)"[^>]*/>', _wb_pre)
if _old_es:
    _old_es_rid = _old_es.group(1)
    # Remove sheet tag from workbook.xml
    _wb_pre = _wb_pre.replace(_old_es.group(0), '')
    with open(_wb_xml_pre_path, 'w', encoding='utf-8') as f:
        f.write(_wb_pre)

    # Find and remove the worksheet file via rels
    _rels_pre_path = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
    with open(_rels_pre_path, 'r', encoding='utf-8') as f:
        _rels_pre = f.read()
    _old_target_m = re.search(f'Id="{_old_es_rid}"[^>]*Target="([^"]+)"', _rels_pre) or \
        re.search(f'Target="([^"]+)"[^>]*Id="{_old_es_rid}"', _rels_pre)
    if _old_target_m:
        _old_sheet_file = _old_target_m.group(1).lstrip('/')
        if not _old_sheet_file.startswith('xl/'):
            _old_sheet_path = os.path.join(temp_dir, 'xl', _old_sheet_file)
        else:
            _old_sheet_path = os.path.join(temp_dir, _old_sheet_file)
        if os.path.exists(_old_sheet_path):
            os.remove(_old_sheet_path)
            print(f"Removed old ExerciseSummary: {_old_sheet_file}")
    # Remove rel entry
    _old_rel_tag = re.search(f'<Relationship[^>]*Id="{_old_es_rid}"[^>]*/>', _rels_pre)
    if _old_rel_tag:
        _rels_pre = _rels_pre.replace(_old_rel_tag.group(0), '')
        with open(_rels_pre_path, 'w', encoding='utf-8') as f:
            f.write(_rels_pre)

    # Remove from [Content_Types].xml
    _ct_path = os.path.join(temp_dir, '[Content_Types].xml')
    if os.path.exists(_ct_path):
        with open(_ct_path, 'r', encoding='utf-8') as f:
            _ct = f.read()
        if _old_target_m:
            _part_name = '/' + ('xl/' + _old_target_m.group(1).lstrip('/') if not _old_target_m.group(1).startswith('xl/') else _old_target_m.group(1).lstrip('/'))
            _ct = re.sub(f'<Override[^>]*PartName="{re.escape(_part_name)}"[^>]*/>', '', _ct)
            with open(_ct_path, 'w', encoding='utf-8') as f:
                f.write(_ct)

# Find a safe sheet number that doesn't collide with worksheets OR chartsheets
_existing_nums = set()
for _f in os.listdir(os.path.join(temp_dir, 'xl', 'worksheets')):
    _m = re.match(r'sheet(\d+)\.xml', _f)
    if _m: _existing_nums.add(int(_m.group(1)))
_cs_dir = os.path.join(temp_dir, 'xl', 'chartsheets')
if os.path.isdir(_cs_dir):
    for _f in os.listdir(_cs_dir):
        _m = re.match(r'sheet(\d+)\.xml', _f)
        if _m: _existing_nums.add(int(_m.group(1)))
_es_sheet_num = max(_existing_nums) + 1 if _existing_nums else 1
_es_sheet_file = f'sheet{_es_sheet_num}.xml'
print(f"ExerciseSummary -> worksheets/{_es_sheet_file} (safe number, avoiding {len(_existing_nums)} existing)")

new_sheet_path = os.path.join(temp_dir, 'xl', 'worksheets', _es_sheet_file)
with zipfile.ZipFile(temp2_path, 'r') as z:
    es_xml_clean = z.read('xl/worksheets/sheet1.xml')
    try: es_strings = z.read('xl/sharedStrings.xml')
    except KeyError: es_strings = None

with open(new_sheet_path, 'wb') as f:
    f.write(es_xml_clean)

# Merge shared strings
orig_ss_path = os.path.join(temp_dir, 'xl', 'sharedStrings.xml')
if os.path.exists(orig_ss_path):
    with open(orig_ss_path, 'r', encoding='utf-8') as f:
        orig_ss = f.read()
    existing_count = len(re.findall(r'<si>', orig_ss))
else:
    orig_ss = None
    existing_count = 0

if es_strings:
    temp_strings = re.findall(r'<si>(.*?)</si>', es_strings.decode('utf-8'), re.DOTALL)
    offset = existing_count
    if orig_ss:
        orig_ss_mod = orig_ss.replace('</sst>', '')
        for ts in temp_strings:
            orig_ss_mod += f'<si>{ts}</si>'
        new_unique = existing_count + len(temp_strings)
        # Only update the <sst> tag — count is total references, uniqueCount is unique strings
        # We add new unique strings, so uniqueCount increases. count also increases by same amount
        # since each new string is referenced at least once by the ExerciseSummary sheet.
        def _fix_sst_counts(m):
            tag = m.group(0)
            # Update count: add the number of new strings to existing count
            old_count = int(re.search(r'count="(\d+)"', tag).group(1))
            tag = re.sub(r'count="\d+"', f'count="{old_count + len(temp_strings)}"', tag)
            tag = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{new_unique}"', tag)
            return tag
        orig_ss_mod = re.sub(r'<sst\b[^>]*>', _fix_sst_counts, orig_ss_mod, count=1)
        orig_ss_mod += '</sst>'
        with open(orig_ss_path, 'w', encoding='utf-8') as f:
            f.write(orig_ss_mod)
    else:
        with open(orig_ss_path, 'wb') as f:
            f.write(es_strings)
        offset = 0
else:
    offset = 0

# --- Merge styles: replace original styles.xml with temp workbook's styles.xml ---
# The simplest robust approach: the temp workbook's styles.xml is self-consistent with the
# ExerciseSummary sheet's s= attributes. Just replace the original styles.xml entirely,
# then remap the OTHER sheets' s= attributes to use the combined style set.
#
# Actually even simpler: just copy the temp styles.xml over the original. The other sheets
# in the workbook already have their own s= values baked into their XML, and those reference
# the ORIGINAL styles.xml indices. So we need to merge.
#
# Robust approach: append ALL temp style sections (fonts, fills, borders, numFmts, cellXfs)
# to the original, then offset s= in the ExerciseSummary sheet XML.

_orig_styles_path = os.path.join(temp_dir, 'xl', 'styles.xml')
with open(_orig_styles_path, 'r', encoding='utf-8') as f:
    _orig_styles = f.read()

with zipfile.ZipFile(temp2_path, 'r') as z:
    _temp_styles = z.read('xl/styles.xml').decode('utf-8')
    _temp_sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')

# Find max s= used in temp sheet
_temp_s_vals = [int(x) for x in re.findall(r'\bs="(\d+)"', _temp_sheet_xml)]
_temp_max_s = max(_temp_s_vals) if _temp_s_vals else 0

def _count_section_entries(xml, parent_tag, child_tag):
    """Count child entries using declared count attribute, fallback to counting tags"""
    count_match = re.search(f'<{parent_tag}\\s+count="(\\d+)"', xml)
    if count_match:
        return int(count_match.group(1))
    section = re.search(f'<{parent_tag}[^>]*>(.*?)</{parent_tag}>', xml, re.DOTALL)
    if not section: return 0
    return len(re.findall(f'<{child_tag}[\\s>]', section.group(1)))

def _get_section_entries(xml, parent_tag, child_tag):
    """Get all child entries from a section (handles both self-closing and with children)"""
    section = re.search(f'<{parent_tag}[^>]*>(.*?)</{parent_tag}>', xml, re.DOTALL)
    if not section: return []
    # Match self-closing first, then with-children (order matters for correct parsing)
    entries = re.findall(f'(<{child_tag}\\b[^/>]*/>|<{child_tag}\\b[^>]*>.*?</{child_tag}>)', section.group(1), re.DOTALL)
    return entries

def _append_to_section(xml, parent_tag, child_tag, new_entries):
    """Append entries to a section and update count"""
    section = re.search(f'<{parent_tag}[^>]*>(.*?)</{parent_tag}>', xml, re.DOTALL)
    if not section: return xml
    # Use declared count if available
    count_match = re.search(f'<{parent_tag}\\s+count="(\\d+)"', xml)
    existing_count = int(count_match.group(1)) if count_match else len(re.findall(f'<{child_tag}[\\s>]', section.group(1)))
    new_count = existing_count + len(new_entries)
    new_content = section.group(1) + '\n'.join(new_entries)
    return xml[:section.start()] + f'<{parent_tag} count="{new_count}">{new_content}</{parent_tag}>' + xml[section.end():]

# Count originals
_orig_font_count = _count_section_entries(_orig_styles, 'fonts', 'font')
_orig_fill_count = _count_section_entries(_orig_styles, 'fills', 'fill')
_orig_border_count = _count_section_entries(_orig_styles, 'borders', 'border')
_orig_cellxfs_section = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', _orig_styles, re.DOTALL)
_orig_xf_count = len(re.findall(r'<xf\b', _orig_cellxfs_section.group(1))) if _orig_cellxfs_section else 0

# Get temp entries
_temp_fonts = _get_section_entries(_temp_styles, 'fonts', 'font')
_temp_fills = _get_section_entries(_temp_styles, 'fills', 'fill')
_temp_borders = _get_section_entries(_temp_styles, 'borders', 'border')
_temp_cellxfs_section = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', _temp_styles, re.DOTALL)
_temp_xf_entries = re.findall(r'(<xf\b[^/>]*/>|<xf\b[^>]*>.*?</xf>)', _temp_cellxfs_section.group(1), re.DOTALL) if _temp_cellxfs_section else []

# Merge numFmts (temp workbook may have its own, e.g. "0.0", "0.00")
_temp_numfmts = _get_section_entries(_temp_styles, 'numFmts', 'numFmt')
_orig_numfmt_ids = set(re.findall(r'numFmtId="(\d+)"', re.search(r'<numFmts[^>]*>(.*?)</numFmts>', _orig_styles, re.DOTALL).group(1))) if '<numFmts' in _orig_styles else set()
# Remap temp numFmt IDs to avoid collisions
# Build a map of format code -> ID in the original
_orig_fc_to_id = {}
for _nf_m in re.finditer(r'numFmtId="(\d+)"[^>]*formatCode="([^"]*)"', _orig_styles):
    _orig_fc_to_id[_nf_m.group(2)] = _nf_m.group(1)

_numfmt_remap = {}
_next_numfmt_id = max((int(x) for x in _orig_numfmt_ids), default=163) + 1
_new_numfmts = []
for nf in _temp_numfmts:
    nf_id = re.search(r'numFmtId="(\d+)"', nf)
    fc_match = re.search(r'formatCode="([^"]*)"', nf)
    if nf_id and fc_match:
        old_id = nf_id.group(1)
        fc = fc_match.group(1)
        if fc in _orig_fc_to_id:
            # Same format code exists in original — remap to that ID
            _numfmt_remap[old_id] = _orig_fc_to_id[fc]
        elif old_id in _orig_numfmt_ids:
            # ID collision but different format code — assign new ID
            _numfmt_remap[old_id] = str(_next_numfmt_id)
            new_nf = re.sub(r'numFmtId="\d+"', f'numFmtId="{_next_numfmt_id}"', nf)
            _new_numfmts.append(new_nf)
            _next_numfmt_id += 1
        else:
            # No collision — keep the same ID but add it to original
            _new_numfmts.append(nf)

if _new_numfmts:
    _orig_styles = _append_to_section(_orig_styles, 'numFmts', 'numFmt', _new_numfmts)

# Append sections
_orig_styles = _append_to_section(_orig_styles, 'fonts', 'font', _temp_fonts)
_orig_styles = _append_to_section(_orig_styles, 'fills', 'fill', _temp_fills)
_orig_styles = _append_to_section(_orig_styles, 'borders', 'border', _temp_borders)

# Remap fontId, fillId, borderId, numFmtId in temp xf entries
_remapped_xfs = []
for xf in _temp_xf_entries:
    xf = re.sub(r'fontId="(\d+)"', lambda m: f'fontId="{int(m.group(1)) + _orig_font_count}"', xf)
    xf = re.sub(r'fillId="(\d+)"', lambda m: f'fillId="{int(m.group(1)) + _orig_fill_count}"', xf)
    xf = re.sub(r'borderId="(\d+)"', lambda m: f'borderId="{int(m.group(1)) + _orig_border_count}"', xf)
    for old_nf_id, new_nf_id in _numfmt_remap.items():
        xf = xf.replace(f'numFmtId="{old_nf_id}"', f'numFmtId="{new_nf_id}"')
    _remapped_xfs.append(xf)

# Append to cellXfs
_orig_styles = _append_to_section(_orig_styles, 'cellXfs', 'xf', _remapped_xfs)

with open(_orig_styles_path, 'w', encoding='utf-8') as f:
    f.write(_orig_styles)

_style_offset = _orig_xf_count
_total_xf = _orig_xf_count + len(_remapped_xfs)

# Fix string indices and remap style IDs in injected sheet XML
sheet_xml = es_xml_clean.decode('utf-8')
def fix_cell(match):
    cell_xml = match.group(0)
    if 't="s"' in cell_xml:
        cell_xml = re.sub(r'<v>(\d+)</v>', lambda m_inner: f'<v>{int(m_inner.group(1)) + offset}</v>', cell_xml)
    # Remap style ID by adding offset
    cell_xml = re.sub(r'\bs="(\d+)"', lambda m_inner: f's="{int(m_inner.group(1)) + _style_offset}"', cell_xml)
    return cell_xml
sheet_xml = re.sub(r'<c [^>]*>.*?</c>', fix_cell, sheet_xml, flags=re.DOTALL)
sheet_xml = re.sub(r'<c [^>]*/>', fix_cell, sheet_xml)
# Remap row-level style refs too
sheet_xml = re.sub(r'(<row [^>]*?)\bs="(\d+)"', lambda m: f'{m.group(1)}s="{int(m.group(2)) + _style_offset}"', sheet_xml)
# Strip customFormat from rows (temp workbook artifact)
sheet_xml = re.sub(r'\s+customFormat="1"', '', sheet_xml)
if not sheet_xml.startswith('<?xml'):
    sheet_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + sheet_xml
with open(new_sheet_path, 'w', encoding='utf-8') as f:
    f.write(sheet_xml)

# Verify: max remapped s= should be < total xf count
_remapped_s_vals = [int(x) for x in re.findall(r'\bs="(\d+)"', sheet_xml)]
_max_remapped = max(_remapped_s_vals) if _remapped_s_vals else 0
print(f"Merged styles: {len(_remapped_xfs)} xf entries appended (total {_total_xf}), max s={_max_remapped}"
      f"{' OK' if _max_remapped < _total_xf else ' ERROR: out of range!'}")

# Update workbook.xml
wb_xml_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
with open(wb_xml_path, 'r', encoding='utf-8') as f:
    wb_xml_content = f.read()

# Ensure the relationships namespace is declared (openpyxl may strip it)
R_NS = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
if 'xmlns:r=' not in wb_xml_content:
    wb_xml_content = wb_xml_content.replace(
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" ' + R_NS)

sheet_ids = [int(x) for x in re.findall(r'sheetId="(\d+)"', wb_xml_content)]
new_sheet_id = max(sheet_ids) + 1 if sheet_ids else 1

rels_path = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
with open(rels_path, 'r', encoding='utf-8') as f:
    rels_xml = f.read()
rid_nums = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels_xml)]
new_rid = f'rId{max(rid_nums) + 1 if rid_nums else 1}'

# Insert ExerciseSummary right after the summary tab (position 1)
_summary_sheet_tag = re.search(r'<sheet name="summary"[^/]*/>', wb_xml_content)
if _summary_sheet_tag:
    _insert_pos = _summary_sheet_tag.end()
    _es_tag = f'<sheet name="ExerciseSummary" sheetId="{new_sheet_id}" r:id="{new_rid}"/>'
    wb_xml_content = wb_xml_content[:_insert_pos] + _es_tag + wb_xml_content[_insert_pos:]
else:
    # Fallback: append at end
    wb_xml_content = wb_xml_content.replace('</sheets>',
        f'<sheet name="ExerciseSummary" sheetId="{new_sheet_id}" r:id="{new_rid}"/></sheets>')

# Force recalculation
if 'calcPr' in wb_xml_content:
    wb_xml_content = re.sub(r'\s*forceFullCalc="[^"]*"', '', wb_xml_content)
    wb_xml_content = re.sub(r'<calcPr([^/]*)/>', r'<calcPr\1 forceFullCalc="1"/>', wb_xml_content)
else:
    wb_xml_content = wb_xml_content.replace('</workbook>', '<calcPr forceFullCalc="1"/></workbook>')

with open(wb_xml_path, 'w', encoding='utf-8') as f:
    f.write(wb_xml_content)

# Update relationships
new_rel = f'<Relationship Id="{new_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/{_es_sheet_file}"/>'
rels_xml = rels_xml.replace('</Relationships>', f'{new_rel}</Relationships>')
with open(rels_path, 'w', encoding='utf-8') as f:
    f.write(rels_xml)

# Update Content_Types
ct_path = os.path.join(temp_dir, '[Content_Types].xml')
with open(ct_path, 'r', encoding='utf-8') as f:
    ct_xml = f.read()
ct_xml = ct_xml.replace('</Types>',
    f'<Override PartName="/xl/worksheets/{_es_sheet_file}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>')
with open(ct_path, 'w', encoding='utf-8') as f:
    f.write(ct_xml)

# Apply chart range replacements AND clear stale caches
chart_dir = os.path.join(temp_dir, 'xl', 'charts')
fixed_count = 0
cache_cleared = 0
for fname in os.listdir(chart_dir):
    if not fname.endswith('.xml') or 'color' in fname or 'style' in fname:
        continue
    fpath = os.path.join(chart_dir, fname)
    with open(fpath, 'r', encoding='utf-8') as f:
        content = f.read()
    new_content = content
    for old_ref, new_ref in range_replacement.items():
        for tag in ['f', 'c:f']:
            old_str = f'<{tag}>{old_ref}</{tag}>'
            new_str = f'<{tag}>{new_ref}</{tag}>'
            if old_str in new_content:
                fixed_count += new_content.count(old_str)
                new_content = new_content.replace(old_str, new_str)
    if new_content != content:
        # Update ptCount in numCache/strCache to match new range sizes,
        # but KEEP the cached data points intact. Stripping caches entirely
        # causes Excel to lose drop lines and up/down bars on chart open.
        # Excel will refresh stale cached values on recalc anyway.
        for old_ref, new_ref in range_replacement.items():
            # Extract old and new row numbers to compute ptCount delta
            import re as _re
            old_m = _re.search(r'\$(\d+)$', old_ref)
            new_m = _re.search(r'\$(\d+)$', new_ref)
            if old_m and new_m:
                old_end = int(old_m.group(1))
                new_end = int(new_m.group(1))
                start_m = _re.search(r'\$(\d+):', old_ref)
                if start_m:
                    old_count = old_end - int(start_m.group(1)) + 1
                    new_count = new_end - int(start_m.group(1)) + 1
                    # Update ptCount values that match the old count
                    old_pt = f'<c:ptCount val="{old_count}"/>'
                    new_pt = f'<c:ptCount val="{new_count}"/>'
                    if old_pt in new_content:
                        new_content = new_content.replace(old_pt, new_pt)
                        cache_cleared += 1
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(new_content)
print(f"Updated {fixed_count} chart references, cleared caches in {cache_cleared} charts")

# Repackage
final_path = DST + '.tmp'
with zipfile.ZipFile(final_path, 'w', zipfile.ZIP_DEFLATED) as zout:
    for root, dirs, files in os.walk(temp_dir):
        for fn in files:
            fp = os.path.join(root, fn)
            arcname = os.path.relpath(fp, temp_dir)
            zout.write(fp, arcname)

os.replace(final_path, DST)
os.system(f'cp "{DST}" "{_output_file}" && chmod 644 "{_output_file}"')
# Clean up temp file
if os.path.exists(DST) and DST != _output_file:
    os.remove(DST)
print(f"\nDone! Output: {_output_file}")
print(f"Charts: {fixed_count} refs updated, ExerciseSummary: {row - 3} weeks")
