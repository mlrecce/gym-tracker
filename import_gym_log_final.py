#!/usr/bin/env python3
"""
Import gym tracker output into Weight2026_DEXA_calibratedc.xlsx

Usage:
  python3 import_gym_log.py <clipboard_file.txt> <excel_file.xlsx>

The clipboard file should contain lines in this format (from the "Copy for Excel" button):
  2026-03-17|bench|95|10x3|notes text
  2026-03-17|run|1mi|08:45|
  2026-03-17|incl. bench|2x27.5|10x3|felt good
  2026-03-17|wall ankle stretch|✓||
  2026-03-17|foam roll thoracic|✓||felt tight

Stretch exercises (from stretches.json) get single columns (just ✓).
Custom exercises not in the column map are auto-added as new columns.
"""

import sys
import os
import re
import zipfile
import tempfile
from datetime import datetime
from openpyxl import load_workbook

# Column mapping: logName -> (weight_col_index, reps_col_index) 0-indexed
COLUMN_MAP = {
    "run":           (5, 6),
    "vest walk":     (7, 8),
    "hang":          (9, None),
    "plank":         (10, None),
    "knee lift":     (11, None),
    "bench":         (12, 13),
    "incl. bench":   (14, 15),
    "mil press":     (16, 17),
    "HS sh press":   (18, 19),
    "pec fly":       (20, 21),
    "Lat lift":      (22, 23),
    "push ups":      (24, None),
    "tric pull":     (25, 26),
    "skull crush":   (27, 28),
    "deadlift":      (29, 30),
    "Row":           (31, 32),
    "bent row":      (33, 34),
    "lat pull dn":   (35, 36),
    "TRX pullup":    (37, None),
    "neg pull up":   (38, None),
    "face pull":     (39, 40),
    "cable curl":    (41, 42),
    "curl":          (43, 44),
    "hammer curl":   (45, 46),
    "back ext":      (47, 48),
    "squats":        (49, 50),
    "land mine squat": (51, 52),
    "lunge walk":    (53, 54),
    "leg press":     (55, 56),
    "hip thrust":    (57, 58),
    "Hstring curl":  (59, 60),
    "weight step":   (61, 62),
}

# Alias map: alternative names from the gym tracker → canonical column header.
# When the clipboard uses one of these names, the import will look for (or create)
# the column under the canonical name so duplicates don't appear.
ALIAS_MAP = {
    "elliptical": "Ellip",
}

# Stretch exercise logNames — single-column (just ✓/done, no reps).
# Kept in sync with stretches.json logNames used by monday_setup.py.
STRETCH_NAMES = {
    "wall ankle stretch", "banded ankle mob", "calf stretch",
    "foam roll thoracic", "open book rotation", "wall slides",
    "doorway pec stretch", "90/90 hip stretch", "hip flexor stretch",
    "adductor stretch", "SL stand eyes closed", "tandem stance",
    "SL head turns", "SL towel stand", "SL RDL", "clock reach",
    "SL BOSU stand", "lunge pause", "SL squat bench",
    "pigeon stretch", "full body stretch",
}

# First column available for dynamically added custom exercises (0-indexed)
# Custom exercises go right after the last static exercise column (62)
CUSTOM_START_COL = 63

# Notes column is DYNAMIC — always 2 columns after the last exercise column
# (or after the last custom exercise column if any were added)
# This ensures notes always stay at the far right
def get_notes_col(ws):
    """Find the notes column, or determine where it should be.
    Scans headers from CUSTOM_START_COL to find 'notes', or returns
    the first empty column after all custom exercises + 1 gap."""
    # First check if notes header already exists somewhere
    for c in range(1, ws.max_column + 2):
        h = ws.cell(row=2, column=c).value
        if h and str(h).strip().lower() == 'notes':
            return c - 1  # return 0-indexed
    # Not found — place it after the last custom exercise column
    col = CUSTOM_START_COL
    while True:
        header = ws.cell(row=2, column=col + 1).value
        if header is None or str(header).strip() == "":
            break
        col += 2  # skip weight + reps columns
    return col + 1  # one gap after last exercise


def parse_clipboard(filepath):
    entries = []
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = line.split('|')
            if len(parts) < 3:
                print(f"  Skipping malformed line: {line}")
                continue
            date_str = parts[0].strip()
            log_name = parts[1].strip()
            val1 = parts[2].strip() if len(parts) > 2 else ""
            val2 = parts[3].strip() if len(parts) > 3 else ""
            note = parts[4].strip() if len(parts) > 4 else ""
            entries.append((date_str, log_name, val1, val2, note))
    return entries


def find_date_row(ws, target_date):
    for row in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val is None:
            continue
        if isinstance(cell_val, datetime):
            if cell_val.strftime('%Y-%m-%d') == target_date:
                return row
        elif isinstance(cell_val, str):
            if target_date in cell_val:
                return row
    return None


def find_next_custom_col(ws):
    """Find the next available column for a custom exercise.
    Scans header row 2 starting from CUSTOM_START_COL to find the first empty column.
    Skips the notes column if it exists."""
    col = CUSTOM_START_COL
    while True:
        header = ws.cell(row=2, column=col + 1).value
        if header is None or str(header).strip() == "":
            return col
        if str(header).strip().lower() == 'notes':
            return col  # insert before notes
        col += 2  # each exercise gets weight + reps columns


def scan_existing_custom_cols(ws):
    """Scan the header row for any previously added custom exercise columns.
    Handles both 2-column (weight+reps) and 1-column (stretch/single) exercises."""
    col = CUSTOM_START_COL
    found = {}
    while True:
        header = ws.cell(row=2, column=col + 1).value
        if header is None or str(header).strip() == "":
            break
        log_name = str(header).strip()
        if log_name.lower() == 'notes':
            break
        reps_header = ws.cell(row=2, column=col + 2).value
        has_reps = reps_header is not None and "reps" in str(reps_header).lower()
        found[log_name] = (col, col + 1 if has_reps else None)
        col += 2 if has_reps else 1
    return found


def add_custom_columns(ws, log_name):
    """Add header columns for a new custom exercise. Returns (weight_col, reps_col) 0-indexed.
    If the notes column is in the way, shift it right by 2."""
    col = find_next_custom_col(ws)

    # Check if notes header is at this position — if so, we need to move it
    existing_header = ws.cell(row=2, column=col + 1).value
    if existing_header and str(existing_header).strip().lower() == 'notes':
        # Move notes data 2 columns to the right (to make room)
        old_notes_col_1indexed = col + 1
        new_notes_col_1indexed = col + 3  # shift right by 2
        # Move header
        ws.cell(row=2, column=new_notes_col_1indexed, value="notes")
        ws.cell(row=2, column=old_notes_col_1indexed, value=None)
        # Move all data in the notes column
        for r in range(3, ws.max_row + 1):
            old_val = ws.cell(r, old_notes_col_1indexed).value
            if old_val is not None:
                ws.cell(r, new_notes_col_1indexed, value=old_val)
                ws.cell(r, old_notes_col_1indexed, value=None)
        print(f"  \u27a1 Moved notes column from {old_notes_col_1indexed} to {new_notes_col_1indexed}")

    # Write headers in row 2
    ws.cell(row=2, column=col + 1, value=log_name)
    ws.cell(row=2, column=col + 2, value=f"{log_name} reps")
    print(f"  \u2795 Added new columns for '{log_name}' at columns {col + 1}-{col + 2}")
    return (col, col + 1)


def import_data(clipboard_path, excel_path):
    entries = parse_clipboard(clipboard_path)
    if not entries:
        print("No entries found in clipboard file.")
        return

    print(f"Found {len(entries)} entries to import.")

    # Single-pass: read the file, compute all cell updates, then do ZIP surgery
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['summary']

    # Scan for any previously added custom exercise columns
    custom_cols = scan_existing_custom_cols(ws)

    # Track the next available column for new custom exercises (0-indexed)
    next_custom_0 = CUSTOM_START_COL
    while True:
        h = ws.cell(row=2, column=next_custom_0 + 1).value
        if h is None or str(h).strip() == "" or str(h).strip().lower() == 'notes':
            break
        # Check if next column is a reps column — if so, skip 2; otherwise skip 1
        reps_h = ws.cell(row=2, column=next_custom_0 + 2).value
        has_reps = reps_h is not None and "reps" in str(reps_h).lower()
        next_custom_0 += 2 if has_reps else 1

    # Process entries and collect all cell updates as (row, col_1indexed, value)
    cell_updates = []
    imported = 0
    skipped = 0
    row_notes = {}
    new_custom = {}  # log_name -> (col0, col0+1) for exercises added this run

    for date_str, log_name, val1, val2, note in entries:
        # Apply alias mapping (e.g. "elliptical" → "Ellip")
        log_name = ALIAS_MAP.get(log_name, log_name)

        row = find_date_row(ws, date_str)
        if row is None:
            print(f"  \u2717 No row found for date {date_str}")
            skipped += 1
            continue

        # Determine columns: static map -> existing custom -> new custom
        if log_name in COLUMN_MAP:
            col1, col2 = COLUMN_MAP[log_name]
        elif log_name in custom_cols:
            col1, col2 = custom_cols[log_name]
        elif log_name in new_custom:
            col1, col2 = new_custom[log_name]
        else:
            # Allocate new custom columns — stretches get 1 column, others get 2
            col1 = next_custom_0
            if log_name in STRETCH_NAMES:
                col2 = None
                new_custom[log_name] = (col1, None)
                next_custom_0 += 1
                print(f"  ➕ Added stretch column for '{log_name}' at column {col1 + 1}")
            else:
                col2 = next_custom_0 + 1
                new_custom[log_name] = (col1, col2)
                next_custom_0 += 2
                print(f"  ➕ Added new columns for '{log_name}' at columns {col1 + 1}-{col2 + 1}")

        # For single-column exercises (col2 is None), the gym tracker puts
        # the value in val2 with val1 empty (e.g. "hang||2x45s|").
        # Fall back: if val1 is empty and col2 is None, use val2 instead.
        effective_val1 = val1 if val1 else (val2 if col2 is None else "")
        if effective_val1:
            cell_updates.append((row, col1 + 1, effective_val1))
        if val2 and col2 is not None:
            cell_updates.append((row, col2 + 1, val2))

        if note:
            row_notes.setdefault(row, []).append(f"{log_name}: {note}")

        day_name = ws.cell(row=row, column=2).value or ""
        print(f"  \u2713 {date_str} ({day_name}) \u2192 {log_name}: {val1} / {val2}" +
              (f" [{note}]" if note else ""))
        imported += 1

    # Notes column: after all exercises (static + existing custom + new custom)
    notes_col_1 = next_custom_0 + 2  # 1-indexed, with 1 gap after last exercise
    # Collect columns occupied by new custom exercises (1-indexed)
    new_custom_cols_1 = set()
    for (c1, c2) in new_custom.values():
        new_custom_cols_1.add(c1 + 1)
        if c2 is not None:
            new_custom_cols_1.add(c2 + 1)

    # Check if notes header already exists somewhere
    cells_to_clear = []  # (row, col) pairs to clear AFTER writing custom headers
    for c in range(1, ws.max_column + 2):
        h = ws.cell(row=2, column=c).value
        if h and str(h).strip().lower() == 'notes':
            if new_custom:
                # Notes column needs to move right to make room
                # Copy old notes data to new position, mark old for clearing
                for r in range(2, ws.max_row + 1):
                    old_val = ws.cell(r, c).value
                    if old_val is not None:
                        cell_updates.append((r, notes_col_1, str(old_val)))
                        # Always clear old notes data (even if col is reused by
                        # a custom exercise — stale notes shouldn't remain there).
                        # Skip row 2 if it's a custom header col (header written later).
                        if r == 2 and c in new_custom_cols_1:
                            pass  # don't clear — custom header will overwrite
                        else:
                            cells_to_clear.append((r, c))
                print(f"  \u27a1 Moving notes column from {c} to {notes_col_1}")
            else:
                notes_col_1 = c  # keep existing position
            break

    # Custom exercise headers — written AFTER notes move so they aren't overwritten
    # Track header cells so we can apply vertical text rotation style
    header_cols = set()  # 1-indexed columns that are exercise headers in row 2
    for log_name, (c1, c2) in new_custom.items():
        cell_updates.append((2, c1 + 1, log_name))
        header_cols.add(c1 + 1)
        if c2 is not None:
            cell_updates.append((2, c2 + 1, f"{log_name} reps"))
            header_cols.add(c2 + 1)

    # Also fix vertical text rotation on any existing custom exercise headers
    # that were previously added without styling
    for log_name, (c1, c2) in custom_cols.items():
        if log_name not in new_custom:  # already handled above for new ones
            header_cols.add(c1 + 1)
            if c2 is not None:
                header_cols.add(c2 + 1)
            # Re-write existing header to trigger style application
            existing_h = ws.cell(row=2, column=c1 + 1).value
            if existing_h:
                cell_updates.append((2, c1 + 1, str(existing_h)))
            if c2 is not None:
                existing_h2 = ws.cell(row=2, column=c2 + 1).value
                if existing_h2:
                    cell_updates.append((2, c2 + 1, str(existing_h2)))

    # Notes header
    cell_updates.append((2, notes_col_1, "notes"))
    header_cols.add(notes_col_1)

    # Notes data
    for row, notes_list in row_notes.items():
        existing = ws.cell(row=row, column=notes_col_1).value
        combined = "; ".join(notes_list)
        if existing and str(existing).strip():
            combined = f"{existing}; {combined}"
        cell_updates.append((row, notes_col_1, combined))

    # Merge new custom into custom_cols for reporting
    custom_cols.update(new_custom)
    wb.close()

    if not cell_updates:
        print("\nNo cell updates to write.")
        return

    # --- Now do XML surgery on the summary sheet ---
    # IMPORTANT: We avoid lxml tree.write() because it reorders namespace
    # attributes on the root <worksheet> element (e.g. moves mc:Ignorable,
    # changes quote style in XML decl). This triggers an Excel "repair" that
    # strips chart formatting (drop lines, up/down bars, etc.).
    # Instead: parse with lxml, modify sheetData, serialize ONLY sheetData,
    # then splice it back into the original XML bytes.
    from lxml import etree as ET
    from openpyxl.utils import get_column_letter

    extract_dir = tempfile.mkdtemp()
    with zipfile.ZipFile(excel_path, 'r') as z:
        z.extractall(extract_dir)

    wb_xml = open(os.path.join(extract_dir, 'xl', 'workbook.xml'), 'r').read()
    rels_xml = open(os.path.join(extract_dir, 'xl', '_rels', 'workbook.xml.rels'), 'r').read()

    rid_match = re.search(r'name="summary"[^>]*r:id="(rId\d+)"', wb_xml) or \
                re.search(r'name="summary"[^>]*id="(rId\d+)"', wb_xml, re.IGNORECASE)
    rid = rid_match.group(1)
    target_match = re.search(f'Id="{rid}"[^>]*Target="([^"]+)"', rels_xml) or \
                   re.search(f'Target="([^"]+)"[^>]*Id="{rid}"', rels_xml)
    sheet_rel = target_match.group(1).lstrip('/')
    if sheet_rel.startswith('xl/'): sheet_rel = sheet_rel[3:]

    sheet_path = os.path.join(extract_dir, 'xl', sheet_rel)

    # Read original bytes to splice back later
    with open(sheet_path, 'rb') as f:
        orig_xml_bytes = f.read()

    tree = ET.parse(sheet_path)
    root = tree.getroot()
    ns = root.nsmap.get(None, 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')

    sheet_data = root.find(f'{{{ns}}}sheetData')
    row_elems = {int(r.get('r')): r for r in sheet_data.findall(f'{{{ns}}}row')}

    # Helper: find or create row element
    def get_row_elem(row_num):
        if row_num not in row_elems:
            row_elem = ET.SubElement(sheet_data, f'{{{ns}}}row')
            row_elem.set('r', str(row_num))
            row_elems[row_num] = row_elem
        return row_elems[row_num]

    # Helper: find cell element in a row
    def find_cell(row_elem, cell_ref):
        for c in row_elem.findall(f'{{{ns}}}c'):
            if c.get('r') == cell_ref:
                return c
        return None

    updated = 0
    for row_num, col_num, value in cell_updates:
        col_letter = get_column_letter(col_num)
        cell_ref = f"{col_letter}{row_num}"
        val_str = str(value)

        # Skip empty values — don't write empty inlineStr cells (causes corruption)
        if not val_str:
            continue

        row_elem = get_row_elem(row_num)
        cell_elem = find_cell(row_elem, cell_ref)

        if cell_elem is None:
            cell_elem = ET.SubElement(row_elem, f'{{{ns}}}c')
            cell_elem.set('r', cell_ref)

        # Remove old value and formula
        for old_v in cell_elem.findall(f'{{{ns}}}v'):
            cell_elem.remove(old_v)
        for old_f in cell_elem.findall(f'{{{ns}}}f'):
            cell_elem.remove(old_f)
        for old_is in cell_elem.findall(f'{{{ns}}}is'):
            cell_elem.remove(old_is)

        # Use inline string (t="inlineStr") to avoid shared string table issues
        cell_elem.set('t', 'inlineStr')
        is_elem = ET.SubElement(cell_elem, f'{{{ns}}}is')
        t_elem = ET.SubElement(is_elem, f'{{{ns}}}t')
        t_elem.text = val_str

        # Apply vertical text rotation style to new row-2 exercise headers.
        # Find the style from the nearest existing exercise header in row 2
        # so new headers match the look (textRotation="90", centered, etc.)
        if row_num == 2 and col_num in header_cols and cell_elem.get('s') is None:
            # Scan row 2 backwards from this column to find a styled header
            header_style = None
            row2_elem = get_row_elem(2)
            for c in row2_elem.findall(f'{{{ns}}}c'):
                s = c.get('s')
                if s:
                    header_style = s  # keep the last one found
            if header_style:
                cell_elem.set('s', header_style)

        updated += 1

    # Clear cells that need to be emptied (notes column old positions)
    # Remove the cell element entirely instead of writing empty values
    for row_num, col_num in cells_to_clear:
        col_letter = get_column_letter(col_num)
        cell_ref = f"{col_letter}{row_num}"
        if row_num in row_elems:
            row_elem = row_elems[row_num]
            cell_elem = find_cell(row_elem, cell_ref)
            if cell_elem is not None:
                row_elem.remove(cell_elem)

    # Fix cell ordering within each modified row — cells MUST be in column
    # order within a <row> element, otherwise Excel flags corruption.
    # Also update row "spans" attributes to encompass any new columns.
    from openpyxl.utils import column_index_from_string
    for row_elem in sheet_data.findall(f'{{{ns}}}row'):
        cell_elems = row_elem.findall(f'{{{ns}}}c')
        if not cell_elems:
            continue

        # Sort cells by column index
        def _col_sort_key(c):
            ref = c.get('r', '')
            col_letters = ''.join(ch for ch in ref if ch.isalpha())
            return column_index_from_string(col_letters) if col_letters else 0

        sorted_cells = sorted(cell_elems, key=_col_sort_key)

        # Check if reordering is needed
        needs_sort = False
        for i in range(1, len(sorted_cells)):
            if _col_sort_key(sorted_cells[i]) < _col_sort_key(cell_elems[i]):
                needs_sort = True
                break
            if _col_sort_key(sorted_cells[i]) > _col_sort_key(cell_elems[i]):
                needs_sort = True
                break

        if needs_sort:
            # Remove all cells, re-add in order
            for c in cell_elems:
                row_elem.remove(c)
            for c in sorted_cells:
                row_elem.append(c)

        # Update spans attribute to encompass all cells
        if row_elem.get('spans'):
            col_indices = [_col_sort_key(c) for c in sorted_cells]
            if col_indices:
                new_spans = f"{min(col_indices)}:{max(col_indices)}"
                row_elem.set('spans', new_spans)

    # Splice modified sheetData back into original XML bytes.
    # This preserves the root element's namespace declarations exactly as
    # Excel wrote them, avoiding the repair trigger.
    new_sd_bytes = ET.tostring(sheet_data)
    new_sd_str = new_sd_bytes.decode('utf-8')
    # Strip ALL namespace declarations lxml adds to the sheetData tag
    # (both default xmlns="..." and prefixed xmlns:foo="..." — they already
    # exist on the root <worksheet> element and duplicating them triggers
    # Excel's "repair" prompt)
    new_sd_str = re.sub(r'\s+xmlns(?::\w+)?="[^"]*"', '', new_sd_str)

    orig_xml_str = orig_xml_bytes.decode('utf-8')
    patched = re.sub(
        r'<sheetData[^>]*>.*?</sheetData>',
        new_sd_str,
        orig_xml_str,
        flags=re.DOTALL
    )

    # Fix dimension ref to encompass all cells (including newly added columns).
    # If we wrote cells beyond the original declared range, Excel flags corruption.
    max_col_written = 0
    max_row_written = 0
    for row_num, col_num, value in cell_updates:
        if value:  # only count non-empty writes
            max_col_written = max(max_col_written, col_num)
            max_row_written = max(max_row_written, row_num)
    dim_match = re.search(r'<dimension\s+ref="([^"]+)"\s*/>', patched)
    if dim_match:
        old_dim = dim_match.group(1)
        # Parse existing dimension to get its max row/col
        parts = old_dim.split(':')
        if len(parts) == 2:
            end_ref = parts[1]
            # Extract column letters and row number from e.g. "BO1001"
            import re as _re
            end_m = _re.match(r'^([A-Z]+)(\d+)$', end_ref)
            if end_m:
                from openpyxl.utils import column_index_from_string
                old_max_col = column_index_from_string(end_m.group(1))
                old_max_row = int(end_m.group(2))
                new_max_col = max(old_max_col, max_col_written)
                new_max_row = max(old_max_row, max_row_written)
                new_end_col_letter = get_column_letter(new_max_col)
                new_dim = f"{parts[0]}:{new_end_col_letter}{new_max_row}"
                if new_dim != old_dim:
                    patched = patched.replace(
                        f'<dimension ref="{old_dim}"/>',
                        f'<dimension ref="{new_dim}"/>'
                    )
                    print(f"  Updated dimension: {old_dim} → {new_dim}")
                else:
                    print(f"  Dimension unchanged: {old_dim}")

    with open(sheet_path, 'wb') as f:
        f.write(patched.encode('utf-8'))
    print(f"  Updated {updated} cells in summary sheet via XML surgery (namespace-safe)")

    # Repackage — preserve [Content_Types].xml first (Excel requirement)
    final_path = excel_path + '.tmp'
    all_files = []
    for root_dir, dirs, files in os.walk(extract_dir):
        for fn in files:
            fp = os.path.join(root_dir, fn)
            arcname = os.path.relpath(fp, extract_dir)
            all_files.append((arcname, fp))
    def _zip_order(item):
        n = item[0]
        if n == '[Content_Types].xml': return (0, n)
        if n.startswith('_rels/'): return (1, n)
        return (2, n)
    all_files.sort(key=_zip_order)

    with zipfile.ZipFile(final_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for arcname, fp in all_files:
            zout.write(fp, arcname)

    os.replace(final_path, excel_path)

    import shutil
    shutil.rmtree(extract_dir, ignore_errors=True)

    print(f"\nDone! {imported} entries imported, {skipped} skipped.")
    if custom_cols:
        print(f"Custom exercises with columns: {', '.join(custom_cols.keys())}")
    print(f"Saved to: {excel_path} (charts preserved)")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python3 import_gym_log.py <clipboard.txt> <Weight2026_DEXA_calibratedc.xlsx>")
        print("\nThe clipboard file should contain lines from the gym tracker's 'Copy for Excel' button.")
        sys.exit(1)

    clipboard_path = sys.argv[1]
    excel_path = sys.argv[2]

    if not os.path.exists(clipboard_path):
        print(f"Error: File not found: {clipboard_path}")
        sys.exit(1)
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    import_data(clipboard_path, excel_path)
